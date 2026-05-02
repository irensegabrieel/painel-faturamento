import os
import re
import sys
import subprocess
import time
from datetime import datetime, timedelta
import threading
import queue
import traceback
import webbrowser
from concurrent.futures import ThreadPoolExecutor, as_completed
import tkinter as tk
from tkinter import ttk, messagebox
from tkinter.scrolledtext import ScrolledText

from dotenv import load_dotenv
import pandas as pd

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter




# ================= GUI / APP =================
APP_TITULO = "Extração CWSI"

# ================= TEMA =================
COR_BG = "#0f172a"
COR_BG_CARD = "#111827"
COR_BG_CARD_2 = "#1f2937"
COR_TEXTO = "#e5e7eb"
COR_TEXTO_SUAVE = "#94a3b8"
COR_DESTAQUE = "#22d3ee"
COR_DESTAQUE_2 = "#38bdf8"
COR_BOTAO = "#1d4ed8"
COR_BOTAO_HOVER = "#2563eb"
COR_BOTAO_NEUTRO = "#334155"
COR_BORDA = "#243041"
COR_LOG_BG = "#020617"
COR_OK = "#10b981"



def esconder_console_windows():
    if not sys.platform.startswith("win"):
        return
    try:
        import ctypes
        hwnd = ctypes.windll.kernel32.GetConsoleWindow()
        if hwnd:
            ctypes.windll.user32.ShowWindow(hwnd, 0)
    except Exception:
        pass


def abrir_pasta(caminho):
    caminho_absoluto = os.path.abspath(caminho)
    try:
        if sys.platform.startswith("win"):
            os.startfile(caminho_absoluto)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", caminho_absoluto])
        else:
            subprocess.Popen(["xdg-open", caminho_absoluto])
    except Exception as e:
        print(f"Não foi possível abrir a pasta: {e}")


class QueueWriter:
    def __init__(self, fila):
        self.fila = fila

    def write(self, texto):
        if texto and texto.strip():
            self.fila.put(("log", texto.rstrip()))

    def flush(self):
        pass


class ExtratorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title(APP_TITULO)
        self.root.geometry("980x680")
        self.root.minsize(900, 620)
        self.root.configure(bg=COR_BG)

        self.fila = queue.Queue()
        self.worker = None
        self.executando = False
        self.arquivos = []
        self.agendamento_ativo = False
        self.agendamento_job = None
        self.agendamento_countdown_job = None
        self.proxima_execucao = None
        self.intervalo_var = tk.StringVar(value="30 minutos")
        self.toggle_agendamento_var = tk.BooleanVar(value=False)
        self.toggle_text_var = tk.StringVar(value="OFF")
        self.countdown_var = tk.StringVar(value="Próxima execução: não programada")

        self.status_var = tk.StringVar(value="Pronto para iniciar")
        self.substatus_var = tk.StringVar(value="Aguardando execução")
        self.arquivos_var = tk.StringVar(value="Nenhum arquivo gerado ainda")
        self.base_var = tk.StringVar(value="AMBAS")
        hoje = datetime.now().strftime("%d/%m/%Y")
        self.data_inicio_var = tk.StringVar(value=hoje)
        self.data_fim_var = tk.StringVar(value=hoje)

        self._montar()
        self._centralizar()
        self._poll()


    def _montar(self):
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass

        style.configure("Root.TFrame", background=COR_BG)
        style.configure("Card.TFrame", background=COR_BG_CARD, relief="flat")
        style.configure("SoftCard.TFrame", background=COR_BG_CARD_2, relief="flat")

        style.configure(
            "Title.TLabel",
            background=COR_BG,
            foreground=COR_TEXTO,
            font=("Segoe UI", 22, "bold")
        )
        style.configure(
            "Subtitle.TLabel",
            background=COR_BG,
            foreground=COR_TEXTO_SUAVE,
            font=("Segoe UI", 10)
        )
        style.configure(
            "Section.TLabel",
            background=COR_BG_CARD,
            foreground=COR_TEXTO,
            font=("Segoe UI", 10, "bold")
        )
        style.configure(
            "Info.TLabel",
            background=COR_BG_CARD,
            foreground=COR_TEXTO,
            font=("Segoe UI", 12)
        )
        style.configure(
            "Soft.TLabel",
            background=COR_BG_CARD,
            foreground=COR_TEXTO_SUAVE,
            font=("Segoe UI", 9)
        )
        style.configure(
            "Files.TLabel",
            background=COR_BG_CARD_2,
            foreground=COR_TEXTO,
            font=("Segoe UI", 10)
        )

        style.configure(
            "Primary.TButton",
            background=COR_BOTAO,
            foreground="#ffffff",
            borderwidth=0,
            focusthickness=0,
            focuscolor=COR_BG_CARD,
            padding=(14, 10),
            font=("Segoe UI", 10, "bold")
        )
        style.map(
            "Primary.TButton",
            background=[("active", COR_BOTAO_HOVER), ("pressed", COR_BOTAO_HOVER)],
            foreground=[("disabled", "#cbd5e1")]
        )

        style.configure(
            "Secondary.TButton",
            background=COR_BOTAO_NEUTRO,
            foreground="#ffffff",
            borderwidth=0,
            focusthickness=0,
            focuscolor=COR_BG_CARD,
            padding=(14, 10),
            font=("Segoe UI", 10, "bold")
        )
        style.map(
            "Secondary.TButton",
            background=[("active", "#475569"), ("pressed", "#475569")],
            foreground=[("disabled", "#cbd5e1")]
        )

        style.configure(
            "Premium.Horizontal.TProgressbar",
            troughcolor=COR_BG_CARD_2,
            background=COR_DESTAQUE,
            bordercolor=COR_BG_CARD_2,
            lightcolor=COR_DESTAQUE,
            darkcolor=COR_DESTAQUE_2,
            thickness=12
        )

        frame = ttk.Frame(self.root, padding=18, style="Root.TFrame")
        frame.pack(fill="both", expand=True)

        hero = tk.Frame(
            frame,
            bg=COR_BG_CARD,
            highlightthickness=1,
            highlightbackground=COR_BORDA,
            bd=0
        )
        hero.pack(fill="x", pady=(0, 14))

        topo = tk.Frame(hero, bg=COR_BG_CARD)
        topo.pack(fill="x", padx=18, pady=18)

        icone = tk.Label(
            topo,
            text="⚡",
            bg=COR_BG_CARD,
            fg=COR_DESTAQUE,
            font=("Segoe UI Emoji", 26)
        )
        icone.pack(side="left", padx=(0, 12))

        textos = tk.Frame(topo, bg=COR_BG_CARD)
        textos.pack(side="left", fill="x", expand=True)

        ttk.Label(textos, text="Extração CWSI", style="Title.TLabel").pack(anchor="w")
        ttk.Label(
            textos,
            text="Parcial automatizada de leitura Bases Americana e Piracicaba.",
            style="Subtitle.TLabel"
        ).pack(anchor="w", pady=(3, 0))

        badge = tk.Button(
            topo,
            text="Suporte/Dúvidas",
            bg="#0f766e",
            fg="#ccfbf1",
            activebackground="#115e59",
            activeforeground="#ecfeff",
            relief="flat",
            bd=0,
            padx=12,
            pady=6,
            font=("Segoe UI", 9, "bold"),
            cursor="hand2",
            command=self.abrir_suporte
        )
        badge.pack(side="right")

        cards = ttk.Frame(frame, style="Root.TFrame")
        cards.pack(fill="x", pady=(0, 14))
        cards.columnconfigure(0, weight=3)
        cards.columnconfigure(1, weight=2)

        status_card = tk.Frame(
            cards, bg=COR_BG_CARD, highlightthickness=1, highlightbackground=COR_BORDA, bd=0
        )
        status_card.grid(row=0, column=0, sticky="nsew", padx=(0, 8))

        status_inner = tk.Frame(status_card, bg=COR_BG_CARD)
        status_inner.pack(fill="both", expand=True, padx=16, pady=14)

        ttk.Label(status_inner, text="Status da automação", style="Section.TLabel").pack(anchor="w")
        ttk.Label(status_inner, textvariable=self.status_var, style="Info.TLabel").pack(anchor="w", pady=(6, 2))
        ttk.Label(status_inner, textvariable=self.substatus_var, style="Soft.TLabel").pack(anchor="w")

        self.progress = ttk.Progressbar(status_inner, mode="indeterminate", style="Premium.Horizontal.TProgressbar")
        self.progress.pack(fill="x", pady=(14, 0))

        files_card = tk.Frame(
            cards, bg=COR_BG_CARD_2, highlightthickness=1, highlightbackground=COR_BORDA, bd=0
        )
        files_card.grid(row=0, column=1, sticky="nsew", padx=(8, 0))

        files_inner = tk.Frame(files_card, bg=COR_BG_CARD_2)
        files_inner.pack(fill="both", expand=True, padx=16, pady=14)

        tk.Label(
            files_inner,
            text="Últimas planilhas",
            bg=COR_BG_CARD_2,
            fg=COR_TEXTO,
            font=("Segoe UI", 10, "bold")
        ).pack(anchor="w")
        ttk.Label(files_inner, textvariable=self.arquivos_var, style="Files.TLabel", wraplength=300).pack(
            anchor="w", pady=(8, 0)
        )

        
        selecao_frame = tk.Frame(frame, bg=COR_BG_CARD)
        selecao_frame.pack(fill="x", pady=(0, 10))

        tk.Label(
            selecao_frame,
            text="Bases:",
            bg=COR_BG_CARD,
            fg=COR_TEXTO,
            font=("Segoe UI", 10, "bold")
        ).pack(side="left", padx=(10,10))

        for txt, val in [("Ambas","AMBAS"),("Americana","AMERICANA"),("Piracicaba","PIRACICABA")]:
            tk.Radiobutton(
                selecao_frame,
                text=txt,
                variable=self.base_var,
                value=val,
                bg=COR_BG_CARD,
                fg=COR_TEXTO,
                selectcolor=COR_BG_CARD,
                activebackground=COR_BG_CARD
            ).pack(side="left", padx=5)

        datas_frame = tk.Frame(frame, bg=COR_BG_CARD, highlightthickness=1, highlightbackground=COR_BORDA, bd=0)
        datas_frame.pack(fill="x", pady=(0, 10))

        tk.Label(
            datas_frame,
            text="Período para exportar:",
            bg=COR_BG_CARD,
            fg=COR_TEXTO,
            font=("Segoe UI", 10, "bold")
        ).pack(side="left", padx=(10, 10), pady=10)

        tk.Label(datas_frame, text="De", bg=COR_BG_CARD, fg=COR_TEXTO_SUAVE, font=("Segoe UI", 10)).pack(side="left")
        self.entry_data_inicio = tk.Entry(
            datas_frame,
            textvariable=self.data_inicio_var,
            width=12,
            bg=COR_BG_CARD_2,
            fg=COR_TEXTO,
            insertbackground=COR_TEXTO,
            relief="flat",
            font=("Segoe UI", 10)
        )
        self.entry_data_inicio.pack(side="left", padx=(6, 14), ipady=5)

        tk.Label(datas_frame, text="Até", bg=COR_BG_CARD, fg=COR_TEXTO_SUAVE, font=("Segoe UI", 10)).pack(side="left")
        self.entry_data_fim = tk.Entry(
            datas_frame,
            textvariable=self.data_fim_var,
            width=12,
            bg=COR_BG_CARD_2,
            fg=COR_TEXTO,
            insertbackground=COR_TEXTO,
            relief="flat",
            font=("Segoe UI", 10)
        )
        self.entry_data_fim.pack(side="left", padx=(6, 10), ipady=5)

        tk.Label(
            datas_frame,
            text="Formato: dd/mm/aaaa. O sistema executa um dia por vez.",
            bg=COR_BG_CARD,
            fg=COR_TEXTO_SUAVE,
            font=("Segoe UI", 9)
        ).pack(side="left", padx=(10, 0))

        botoes_wrap = ttk.Frame(frame, style="Root.TFrame")
        botoes_wrap.pack(fill="x", pady=(0, 14))

        self.btn_iniciar = ttk.Button(
            botoes_wrap, text="Iniciar extração", command=self.iniciar, style="Primary.TButton"
        )
        self.btn_iniciar.pack(side="left")

        ttk.Button(
            botoes_wrap,
            text="Abrir pasta com últimas planilhas",
            command=self.abrir_pasta_saida,
            style="Secondary.TButton"
        ).pack(side="left", padx=(10, 0))

        ttk.Button(
            botoes_wrap,
            text="Abrir última planilha",
            command=self.abrir_ultimo_excel,
            style="Secondary.TButton"
        ).pack(side="left", padx=(10, 0))

        agenda_card = tk.Frame(
            frame, bg=COR_BG_CARD, highlightthickness=1, highlightbackground=COR_BORDA, bd=0
        )
        agenda_card.pack(fill="x", pady=(0, 14))

        agenda_inner = tk.Frame(agenda_card, bg=COR_BG_CARD)
        agenda_inner.pack(fill="x", padx=16, pady=14)

        agenda_top = tk.Frame(agenda_inner, bg=COR_BG_CARD)
        agenda_top.pack(fill="x")

        agenda_info = tk.Frame(agenda_top, bg=COR_BG_CARD)
        agenda_info.pack(side="left", fill="x", expand=True)

        tk.Label(
            agenda_info,
            text="Programação automática (em teste)",
            bg=COR_BG_CARD,
            fg=COR_TEXTO,
            font=("Segoe UI", 11, "bold")
        ).pack(anchor="w")

        self.lbl_countdown = tk.Label(
            agenda_info,
            textvariable=self.countdown_var,
            bg=COR_BG_CARD,
            fg=COR_DESTAQUE,
            font=("Segoe UI", 10, "bold")
        )
        self.lbl_countdown.pack(anchor="w", pady=(6, 0))

        agenda_controls = tk.Frame(agenda_top, bg=COR_BG_CARD)
        agenda_controls.pack(side="right", anchor="ne")

        self.combo_intervalo = ttk.Combobox(
            agenda_controls,
            textvariable=self.intervalo_var,
            values=["30 minutos", "1 hora", "2 horas"],
            state="readonly",
            width=12
        )
        self.combo_intervalo.grid(row=0, column=0, padx=(0, 10), pady=0)

        self.btn_toggle_agendamento = tk.Checkbutton(
            agenda_controls,
            textvariable=self.toggle_text_var,
            variable=self.toggle_agendamento_var,
            command=self.toggle_agendamento,
            indicatoron=False,
            width=8,
            padx=12,
            pady=8,
            relief="flat",
            bd=0,
            cursor="hand2",
            bg=COR_BOTAO_NEUTRO,
            fg="#ffffff",
            activebackground="#475569",
            activeforeground="#ffffff",
            selectcolor=COR_OK,
            font=("Segoe UI", 10, "bold")
        )
        self.btn_toggle_agendamento.grid(row=0, column=1, pady=0)

        self.lbl_agenda_desc = tk.Label(
            agenda_inner,
            text="Ao ativar, a primeira extração começa na hora e as próximas seguem no intervalo escolhido.",
            bg=COR_BG_CARD,
            fg=COR_TEXTO_SUAVE,
            font=("Segoe UI", 9),
            justify="left",
            anchor="w"
        )
        self.lbl_agenda_desc.pack(fill="x", pady=(10, 0))
        self.root.bind("<Configure>", self._ajustar_wrap_programacao)

        log_card = tk.Frame(
            frame, bg=COR_BG_CARD, highlightthickness=1, highlightbackground=COR_BORDA, bd=0
        )
        log_card.pack(fill="both", expand=True)

        log_top = tk.Frame(log_card, bg=COR_BG_CARD)
        log_top.pack(fill="x", padx=16, pady=(14, 8))

        tk.Label(
            log_top,
            text="Log de execução",
            bg=COR_BG_CARD,
            fg=COR_TEXTO,
            font=("Segoe UI", 10, "bold")
        ).pack(side="left")

        tk.Label(
            log_top,
            text="Tempo real",
            bg=COR_BG_CARD,
            fg=COR_OK,
            font=("Segoe UI", 9, "bold")
        ).pack(side="right")

        log_frame = tk.Frame(log_card, bg=COR_BG_CARD)
        log_frame.pack(fill="both", expand=True, padx=16, pady=(0, 16))

        self.txt_log = tk.Text(
            log_frame,
            height=20,
            font=("Consolas", 10),
            bg=COR_LOG_BG,
            fg=COR_TEXTO,
            insertbackground=COR_TEXTO,
            relief="flat",
            bd=0,
            padx=12,
            pady=10,
            wrap="none"
        )
        self.txt_log.pack(side="left", fill="both", expand=True)

        scrollbar_y = tk.Scrollbar(log_frame, command=self.txt_log.yview)
        scrollbar_y.pack(side="right", fill="y")

        scrollbar_x = tk.Scrollbar(log_frame, command=self.txt_log.xview, orient="horizontal")
        scrollbar_x.pack(side="bottom", fill="x")

        self.txt_log.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        self.txt_log.configure(state="disabled")

    def _centralizar(self):

        self.root.update_idletasks()
        w = self.root.winfo_width()
        h = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (w // 2)
        y = (self.root.winfo_screenheight() // 2) - (h // 2)
        self.root.geometry(f"{w}x{h}+{x}+{y}")
        self._ajustar_wrap_programacao()
        self._atualizar_visual_agendamento()

    def _append_log(self, texto):
        self.txt_log.configure(state="normal")
        self.txt_log.insert("end", texto + "\n")
        self.txt_log.see("end")
        self.txt_log.configure(state="disabled")

    def _poll(self):
        try:
            while True:
                tipo, valor = self.fila.get_nowait()
                if tipo == "log":
                    self._append_log(str(valor))
                elif tipo == "status":
                    self.status_var.set(str(valor))
                elif tipo == "substatus":
                    self.substatus_var.set(str(valor))
                elif tipo == "arquivos":
                    self.arquivos = list(valor)
                    self.arquivos_var.set(" | ".join(self.arquivos) if self.arquivos else "Nenhum arquivo gerado ainda")
                elif tipo == "fim":
                    self._finalizar(valor)
        except queue.Empty:
            pass

        self.root.after(120, self._poll)

    def set_status(self, texto):
        self.fila.put(("status", texto))

    def set_substatus(self, texto):
        self.fila.put(("substatus", texto))

    def log(self, texto):
        self.fila.put(("log", texto))

    def atualizar_arquivos(self, arquivos):
        self.fila.put(("arquivos", arquivos))

    def _ajustar_wrap_programacao(self, event=None):
        try:
            largura = max(320, self.root.winfo_width() - 260)
            self.lbl_agenda_desc.configure(wraplength=largura)
        except Exception:
            pass

    def _atualizar_visual_agendamento(self):
        if self.agendamento_ativo:
            self.toggle_text_var.set("ON")
            self.btn_toggle_agendamento.configure(
                bg=COR_OK,
                activebackground="#059669"
            )
        else:
            self.toggle_text_var.set("OFF")
            self.btn_toggle_agendamento.configure(
                bg=COR_BOTAO_NEUTRO,
                activebackground="#475569"
            )
            if not self.proxima_execucao:
                self.countdown_var.set("Próxima execução: não programada")

    def obter_intervalo_segundos(self):
        mapa = {
            "30 minutos": 30 * 60,
            "1 hora": 60 * 60,
            "2 horas": 2 * 60 * 60,
        }
        return mapa.get(self.intervalo_var.get(), 30 * 60)

    def toggle_agendamento(self):
        if self.agendamento_ativo:
            self.parar_agendamento()
        else:
            self.iniciar_agendamento()

    def iniciar_agendamento(self):
        self.agendamento_ativo = True
        self.toggle_agendamento_var.set(True)
        self._atualizar_visual_agendamento()
        self.combo_intervalo.configure(state="disabled")
        self.log(
            f"Programação automática ativada ({self.intervalo_var.get()}). A primeira execução será iniciada agora."
        )
        self.set_status("Programação automática ativa")
        self.set_substatus("Primeira execução agendada para agora")
        self.countdown_var.set("Primeira execução: iniciando agora")
        self.iniciar()

    def parar_agendamento(self):
        self.agendamento_ativo = False
        self.toggle_agendamento_var.set(False)
        self.proxima_execucao = None

        if self.agendamento_job is not None:
            try:
                self.root.after_cancel(self.agendamento_job)
            except Exception:
                pass
            self.agendamento_job = None

        if self.agendamento_countdown_job is not None:
            try:
                self.root.after_cancel(self.agendamento_countdown_job)
            except Exception:
                pass
            self.agendamento_countdown_job = None

        self.combo_intervalo.configure(state="readonly")
        self._atualizar_visual_agendamento()
        self.log("Programação automática desativada.")
        if not self.executando:
            self.set_status("Pronto para iniciar")
            self.set_substatus("Aguardando execução")

    def _agendar_proxima_execucao(self):
        if not self.agendamento_ativo:
            return

        intervalo = self.obter_intervalo_segundos()
        self.proxima_execucao = datetime.now() + timedelta(seconds=intervalo)

        if self.agendamento_job is not None:
            try:
                self.root.after_cancel(self.agendamento_job)
            except Exception:
                pass

        if self.agendamento_countdown_job is not None:
            try:
                self.root.after_cancel(self.agendamento_countdown_job)
            except Exception:
                pass

        self.agendamento_job = self.root.after(intervalo * 1000, self._disparar_execucao_agendada)
        self._atualizar_countdown_agendamento()
        self.log(
            f"Próxima execução agendada para {self.proxima_execucao.strftime('%d/%m/%Y %H:%M:%S')}."
        )

    def _atualizar_countdown_agendamento(self):
        if not self.agendamento_ativo or not self.proxima_execucao:
            return

        restante = int((self.proxima_execucao - datetime.now()).total_seconds())
        if restante < 0:
            restante = 0

        horas = restante // 3600
        minutos = (restante % 3600) // 60
        segundos = restante % 60

        texto_countdown = f"Próxima execução em {horas:02d}:{minutos:02d}:{segundos:02d}"
        self.countdown_var.set(texto_countdown)
        self.set_status("Programação automática ativa")
        self.set_substatus(texto_countdown)

        if restante > 0:
            self.agendamento_countdown_job = self.root.after(1000, self._atualizar_countdown_agendamento)
        else:
            self.agendamento_countdown_job = None

    def _disparar_execucao_agendada(self):
        self.agendamento_job = None
        self.agendamento_countdown_job = None
        self.proxima_execucao = None

        if not self.agendamento_ativo:
            return

        if self.executando:
            self.log("A execução agendada encontrou uma extração em andamento. Reagendando próximo ciclo.")
            self._agendar_proxima_execucao()
            return

        self.countdown_var.set("Executando extração automática agora")
        self.log("Iniciando execução automática programada.")
        self.iniciar()

    def iniciar(self):
        if self.executando:
            return

        self.executando = True
        self.arquivos = []
        self.arquivos_var.set("Nenhum arquivo gerado ainda")
        self.btn_iniciar.configure(state="disabled")
        self.progress.start(12)
        self.set_status("Iniciando automação")
        self.set_substatus("Preparando execução")
        self.log("=" * 90)
        self.log(f"Início: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        self.log("=" * 90)

        self.worker = threading.Thread(target=self._run_worker, daemon=True)
        self.worker.start()

    def _run_worker(self):
        stdout_original = sys.stdout
        stderr_original = sys.stderr
        writer = QueueWriter(self.fila)

        try:
            sys.stdout = writer
            sys.stderr = writer
            resultado = main(
                app=self,
                base_selecionada=self.base_var.get(),
                data_inicio_txt=self.data_inicio_var.get(),
                data_fim_txt=self.data_fim_var.get()
            )
            self.fila.put(("fim", {"ok": True, "resultado": resultado}))
        except Exception as e:
            erro = "".join(traceback.format_exception(type(e), e, e.__traceback__))
            self.fila.put(("log", erro))
            self.fila.put(("fim", {"ok": False, "erro": str(e)}))
        finally:
            sys.stdout = stdout_original
            sys.stderr = stderr_original

    def _finalizar(self, payload):
        self.executando = False
        self.progress.stop()
        self.btn_iniciar.configure(state="normal")

        if payload.get("ok"):
            resultado = payload.get("resultado") or {}
            arquivos = resultado.get("arquivos", [])
            self.atualizar_arquivos(arquivos)
            if self.agendamento_ativo:
                self.countdown_var.set("Execução concluída. Preparando próximo ciclo...")
                self.set_status("Execução automática concluída")
                self.set_substatus("Preparando próximo ciclo")
                self.log("Execução automática concluída com sucesso.")
            else:
                self.set_status("Finalizado com sucesso")
                self.set_substatus("Processamento concluído")
                messagebox.showinfo("Concluído", "Extração finalizada com sucesso.")
        else:
            if self.agendamento_ativo:
                self.countdown_var.set("Execução com erro. Reagendando próximo ciclo...")
                self.set_status("Execução automática com erro")
                self.set_substatus("O próximo ciclo será reagendado")
                self.log(f"Erro na execução automática: {payload.get('erro', 'Ocorreu um erro inesperado.')}")
            else:
                self.set_status("Erro na execução")
                self.set_substatus("Verifique o log abaixo")
                messagebox.showerror("Erro", payload.get("erro", "Ocorreu um erro inesperado."))

        if self.agendamento_ativo:
            self._agendar_proxima_execucao()
        else:
            self.countdown_var.set("Próxima execução: não programada")

    def abrir_suporte(self):
        url = "https://wa.me/5519991436620"
        try:
            if sys.platform.startswith("win"):
                os.startfile(url)
            elif sys.platform == "darwin":
                subprocess.Popen(["open", url])
            else:
                subprocess.Popen(["xdg-open", url])
        except Exception:
            try:
                webbrowser.open_new_tab(url)
            except Exception as e:
                try:
                    messagebox.showerror("Suporte/Dúvidas", f"Não foi possível abrir o WhatsApp.\n\n{e}")
                except Exception:
                    print(f"Não foi possível abrir o WhatsApp: {e}")

    def abrir_pasta_saida(self):
        garantir_pasta_saida()
        abrir_pasta(PASTA_SAIDA)

    def abrir_ultimo_excel(self):
        arquivos = list(self.arquivos)
        if arquivos:
            abrir_arquivo(arquivos[-1])
            return

        garantir_pasta_saida()
        candidatos = [
            os.path.join(PASTA_SAIDA, nome)
            for nome in os.listdir(PASTA_SAIDA)
            if nome.lower().endswith(".xlsx")
        ]
        if not candidatos:
            messagebox.showwarning("Aviso", "Nenhuma planilha encontrada na pasta.")
            return

        ultimo = max(candidatos, key=os.path.getmtime)
        abrir_arquivo(ultimo)


def iniciar_interface():
    esconder_console_windows()
    root = tk.Tk()
    ExtratorGUI(root)
    root.mainloop()


# ================= CONFIG =================
TEMPO_PADRAO = 30

# Pasta fixa onde as parciais de leitura ficam disponíveis para o painel local.
# Mantém o fluxo separado dos arquivos de corte/faturamento.
PASTA_SAIDA = os.path.join(os.path.expanduser("~"), "Desktop", "LEITURA", "saida")


# ================= DRIVER =================
def iniciar_driver():
    options = Options()
    # V8: volta para headless. Mantém tamanho fixo para preservar o layout da grade.
    options.add_argument("--headless=new")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-infobars")
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--disable-gpu")
    options.add_argument("--force-device-scale-factor=1")
    options.add_argument("--high-dpi-support=1")

    driver = webdriver.Chrome(options=options)
    try:
        driver.set_window_position(0, 0)
        driver.set_window_size(1920, 1080)
    except Exception:
        pass
    driver.implicitly_wait(0)
    return driver


def esperar(driver, tempo=TEMPO_PADRAO):
    return WebDriverWait(driver, tempo)


# ================= UTILS =================
def texto_limpo(valor):
    if valor is None:
        return ""
    return " ".join(str(valor).replace("\xa0", " ").split()).strip()


def garantir_pasta_saida():
    os.makedirs(PASTA_SAIDA, exist_ok=True)


def abrir_arquivo(caminho_arquivo):
    caminho_absoluto = os.path.abspath(caminho_arquivo)

    try:
        if sys.platform.startswith("win"):
            os.startfile(caminho_absoluto)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", caminho_absoluto])
        else:
            subprocess.Popen(["xdg-open", caminho_absoluto])
    except Exception as e:
        print(f"Não foi possível abrir o arquivo automaticamente: {e}")


def timestamp_str():
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def parse_data_br(valor):
    valor = texto_limpo(valor)
    try:
        return datetime.strptime(valor, "%d/%m/%Y").date()
    except ValueError:
        raise Exception(f"Data inválida: {valor}. Use o formato dd/mm/aaaa.")


def gerar_periodos_diarios(data_inicio_txt, data_fim_txt):
    inicio = parse_data_br(data_inicio_txt)
    fim = parse_data_br(data_fim_txt)

    if fim < inicio:
        raise Exception("A data final não pode ser menor que a data inicial.")

    periodos = []
    atual = inicio
    while atual <= fim:
        data_br = atual.strftime("%d/%m/%Y")
        periodos.append((data_br, data_br))
        atual += timedelta(days=1)

    return periodos


def data_para_nome_arquivo(data_br):
    return parse_data_br(data_br).strftime("%Y-%m-%d")


def gerar_posicoes(inicio, fim, passo):
    posicoes = []
    atual = inicio

    while atual <= fim:
        posicoes.append(int(round(atual)))
        atual += passo

    fim_int = int(round(fim))
    if not posicoes or posicoes[-1] != fim_int:
        posicoes.append(fim_int)

    finais = []
    vistos = set()
    for p in posicoes:
        if p not in vistos:
            vistos.add(p)
            finais.append(p)

    return finais


def eh_numero_simples(txt):
    txt = texto_limpo(txt)
    if not txt:
        return False

    txt = txt.replace(".", "").replace(",", "")
    return txt.isdigit()


def parece_nome_pessoa(txt):
    txt = texto_limpo(txt)
    if not txt:
        return False

    partes = [p for p in txt.split() if p.strip()]
    if len(partes) < 2:
        return False

    tem_letra = any(ch.isalpha() for ch in txt)
    tem_numero = any(ch.isdigit() for ch in txt)

    if not tem_letra or tem_numero:
        return False

    if len(txt) < 8:
        return False

    bloqueados = {
        "LEITURA", "AME", "P", "R", "CRIADA", "CONTROLE"
    }
    if txt.upper() in bloqueados:
        return False

    return True


def slug_base(nome_base):
    nome = texto_limpo(nome_base).upper()
    nome = nome.replace(" [B]", "")
    nome = nome.replace(" ", "_")
    nome = nome.replace("/", "_")
    return nome


def normalizar_nome_agente(nome):
    return texto_limpo(nome).upper()


# ================= LOGIN =================
def preencher_login(driver, usuario, senha):
    wait = esperar(driver, 30)

    print("Preenchendo login...")

    campo_usuario = wait.until(
        EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Usuário']"))
    )

    campo_senha = wait.until(
        EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Senha']"))
    )

    campo_usuario.clear()
    campo_usuario.send_keys(usuario)

    campo_senha.clear()
    campo_senha.send_keys(senha)

    print("Procurando botão Login...")
    botao_login = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//*[self::button or self::a or self::input][contains(., 'Login') or @value='Login']"
        ))
    )

    botao_login.click()
    print("Login enviado com sucesso.")


# ================= TELA INICIAL =================
def aguardar_tela_principal(driver):
    wait = esperar(driver, 30)

    print("Aguardando tela principal carregar...")

    wait.until(
        EC.presence_of_element_located((
            By.XPATH,
            "//*[contains(normalize-space(.), 'Portal de Leitura e Entrega') or contains(normalize-space(.), 'Planejamento / Operação')]"
        ))
    )

    wait.until(
        EC.presence_of_element_located((
            By.XPATH,
            "//h2[contains(normalize-space(.), 'Planejamento / Operação')]"
        ))
    )

    time.sleep(3)


def clicar_card_planejamento_operacao(driver):
    wait = esperar(driver, 30)

    print("Tentando clicar no card Planejamento / Operação...")

    links = wait.until(
        EC.presence_of_all_elements_located((
            By.XPATH,
            "//a[normalize-space(.)='Planejamento / Operação']"
        ))
    )

    alvo = None
    for link in links:
        try:
            if link.is_displayed():
                alvo = link
                break
        except Exception:
            pass

    if alvo is None:
        raise Exception("Não encontrei o card/link 'Planejamento / Operação'.")

    try:
        alvo.click()
        print("Card clicado com clique normal.")
    except Exception:
        driver.execute_script("arguments[0].click();", alvo)
        print("Card clicado com JavaScript.")

    time.sleep(3)


def aguardar_tela_planejamento(driver):
    wait = esperar(driver, 40)

    print("Aguardando tela de Planejamento carregar...")

    wait.until(
        lambda d: (
            len(d.find_elements(By.XPATH, "//*[contains(@placeholder,'Pesquisar Agente')]")) > 0
            or len(d.find_elements(By.XPATH, "//*[contains(@placeholder,'Pesquisar Base/Município')]")) > 0
            or len(d.find_elements(By.XPATH, "//*[contains(normalize-space(.), 'PAULISTA')]")) > 0
            or len(d.find_elements(By.XPATH, "//*[contains(normalize-space(.), 'PIRATININGA')]")) > 0
            or len(d.find_elements(By.XPATH, "//*[contains(normalize-space(.), 'RGE')]")) > 0
            or len(d.find_elements(By.XPATH, "//*[contains(normalize-space(.), 'SANTA CRUZ')]")) > 0
        )
    )

    print("Tela de Planejamento carregada com sucesso.")
    time.sleep(2)


# ================= ÁRVORE / LINHAS =================
def encontrar_linha_por_texto(driver, texto, timeout=20):
    wait = esperar(driver, timeout)

    candidatos = wait.until(
        EC.presence_of_all_elements_located((
            By.XPATH,
            f"//*[normalize-space(text())='{texto}']"
        ))
    )

    texto_el = None
    for el in candidatos:
        try:
            if el.is_displayed():
                texto_el = el
                break
        except Exception:
            pass

    if texto_el is None:
        raise Exception(f"Não encontrei o texto visível '{texto}'.")

    linha = texto_el.find_element(
        By.XPATH,
        "./ancestor::*[self::div or self::li or self::tr][1]"
    )

    driver.execute_script(
        "arguments[0].scrollIntoView({block:'center', inline:'nearest'});",
        linha
    )
    time.sleep(0.5)

    return linha, texto_el


def clicar_por_coordenada_na_linha(driver, texto, desloc_x, timeout=20):
    linha, texto_el = encontrar_linha_por_texto(driver, texto, timeout)

    rect_texto = driver.execute_script("""
        const r = arguments[0].getBoundingClientRect();
        return {
            left: r.left,
            top: r.top,
            width: r.width,
            height: r.height
        };
    """, texto_el)

    x = rect_texto["left"] + desloc_x
    y = rect_texto["top"] + (rect_texto["height"] / 2)

    print(f"Clicando na linha '{texto}' em x={x:.0f}, y={y:.0f}")

    driver.execute_script("window.scrollBy(0, -80);")
    time.sleep(0.3)

    sucesso = driver.execute_script("""
        const x = arguments[0];
        const y = arguments[1];

        const el = document.elementFromPoint(x, y);
        if (!el) return false;

        ['mousemove', 'mousedown', 'mouseup', 'click'].forEach(evtName => {
            el.dispatchEvent(new MouseEvent(evtName, {
                view: window,
                bubbles: true,
                cancelable: true,
                clientX: x,
                clientY: y,
                buttons: 1
            }));
        });

        return true;
    """, x, y)

    if not sucesso:
        raise Exception(f"Falha ao clicar na coordenada da linha '{texto}'.")

    time.sleep(1)


def clicar_seta_da_linha(driver, texto, timeout=20):
    print(f"Expandindo '{texto}'...")
    clicar_por_coordenada_na_linha(driver, texto, desloc_x=-55, timeout=timeout)


def clicar_checkbox_da_linha(driver, texto, timeout=20):
    print(f"Marcando '{texto}'...")
    clicar_por_coordenada_na_linha(driver, texto, desloc_x=-30, timeout=timeout)


def expandir_arvore_bases(driver):
    wait = esperar(driver, 30)

    print("Expandindo PAULISTA...")
    clicar_seta_da_linha(driver, "PAULISTA", 20)
    time.sleep(2)

    wait.until(
        lambda d: (
            len(d.find_elements(By.XPATH, "//*[normalize-space(text())='SUDESTE']")) > 0
            or len(d.find_elements(By.XPATH, "//*[contains(normalize-space(text()), 'AMERICANA')]")) > 0
        )
    )

    print("Expandindo SUDESTE...")
    clicar_seta_da_linha(driver, "SUDESTE", 20)
    time.sleep(2)

    wait.until(
        EC.presence_of_element_located((
            By.XPATH,
            "//*[contains(normalize-space(text()), 'AMERICANA')]"
        ))
    )

    wait.until(
        EC.presence_of_element_located((
            By.XPATH,
            "//*[contains(normalize-space(text()), 'PIRACICABA')]"
        ))
    )


def expandir_e_marcar_base(driver, nome_base):
    expandir_arvore_bases(driver)

    print(f"Marcando somente a base {nome_base}...")
    clicar_checkbox_da_linha(driver, nome_base, 20)

    print(f"Base {nome_base} selecionada com sucesso.")


# ================= DATAS DO FILTRO =================
def preencher_periodo_datas(driver, data_inicio, data_fim, timeout=20):
    """Preenche os dois campos inferiores: Data Prevista e Até."""
    esperar(driver, timeout).until(lambda d: len(d.find_elements(By.XPATH, "//input")) > 0)
    print(f"Preenchendo período na tela: {data_inicio} até {data_fim}")

    resultado = driver.execute_script(r'''
        const dataInicio = arguments[0];
        const dataFim = arguments[1];
        function visivel(el) {
            if (!el) return false;
            const st = getComputedStyle(el);
            const r = el.getBoundingClientRect();
            return st.display !== 'none' && st.visibility !== 'hidden' && r.width > 0 && r.height > 0;
        }
        function limpar(t) { return (t || '').replace(/\s+/g, ' ').trim(); }
        function setValor(input, valor) {
            const proto = Object.getPrototypeOf(input);
            const desc = Object.getOwnPropertyDescriptor(proto, 'value') || Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'value');
            if (desc && desc.set) desc.set.call(input, valor); else input.value = valor;
            input.dispatchEvent(new Event('input', {bubbles: true}));
            input.dispatchEvent(new Event('change', {bubbles: true}));
            input.dispatchEvent(new Event('blur', {bubbles: true}));
        }
        function elementosComTexto(texto) {
            return [...document.querySelectorAll('label, span, div, td, th, p')]
                .filter(visivel)
                .filter(el => limpar(el.innerText || el.textContent) === texto);
        }
        function inputPertoDoLabel(textoLabel) {
            const labels = elementosComTexto(textoLabel);
            const inputs = [...document.querySelectorAll('input')].filter(visivel);
            let melhor = null;
            let melhorScore = Infinity;
            for (const label of labels) {
                const lr = label.getBoundingClientRect();
                const ly = lr.top + lr.height / 2;
                for (const input of inputs) {
                    const ir = input.getBoundingClientRect();
                    const iy = ir.top + ir.height / 2;
                    const mesmaLinha = Math.abs(iy - ly) <= 28;
                    const aDireita = ir.left >= lr.right - 5;
                    const perto = ir.left <= lr.right + 220;
                    if (!mesmaLinha || !aDireita || !perto) continue;
                    const score = Math.abs(iy - ly) + Math.max(0, ir.left - lr.right);
                    if (score < melhorScore) { melhorScore = score; melhor = input; }
                }
            }
            return melhor;
        }
        let inputInicio = inputPertoDoLabel('Data Prevista');
        let inputFim = inputPertoDoLabel('Até');
        if (!inputInicio || !inputFim || inputInicio === inputFim) {
            const todos = [...document.querySelectorAll('input')]
                .filter(visivel)
                .map(input => ({input, r: input.getBoundingClientRect(), value: input.value || ''}))
                .filter(o => o.r.top > window.innerHeight * 0.45)
                .filter(o => /\d{2}\/\d{2}\/\d{4}/.test(o.value) || o.input.type === 'text')
                .sort((a, b) => (a.r.top - b.r.top) || (a.r.left - b.r.left));
            if (todos.length >= 2) { inputInicio = todos[0].input; inputFim = todos[1].input; }
        }
        if (!inputInicio || !inputFim) return {ok:false, erro:'Não encontrei os dois campos de data inferiores.'};
        inputInicio.focus(); setValor(inputInicio, dataInicio);
        inputFim.focus(); setValor(inputFim, dataFim);
        return {ok:true, inicio:inputInicio.value, fim:inputFim.value};
    ''', data_inicio, data_fim)

    if not resultado or not resultado.get('ok'):
        raise Exception((resultado or {}).get('erro', 'Falha ao preencher período.'))
    time.sleep(0.8)
    print(f"Período preenchido: {resultado.get('inicio')} até {resultado.get('fim')}")


# ================= BOTÃO AO LADO DO "ATÉ" =================
def clicar_botao_destacado_vermelho(driver, timeout=20):
    wait = esperar(driver, timeout)

    print("Procurando o botão ao lado do texto 'Até'...")

    ate = wait.until(
        EC.presence_of_element_located((
            By.XPATH,
            "//*[normalize-space(text())='Até']"
        ))
    )

    driver.execute_script(
        "arguments[0].scrollIntoView({block:'center', inline:'nearest'});",
        ate
    )
    time.sleep(0.5)

    rect_ate = driver.execute_script("""
        const r = arguments[0].getBoundingClientRect();
        return {
            left: r.left,
            top: r.top,
            right: r.right,
            bottom: r.bottom,
            width: r.width,
            height: r.height
        };
    """, ate)

    y_centro = rect_ate["top"] + rect_ate["height"] / 2

    print(f"Texto 'Até' encontrado em x={rect_ate['left']:.0f}, y={y_centro:.0f}")

    candidatos = driver.find_elements(
        By.XPATH,
        "//*[self::button or self::a or self::div or self::span or self::i or self::svg]"
    )

    melhor = None
    melhor_score = 999999

    for el in candidatos:
        try:
            if not el.is_displayed():
                continue

            rect = driver.execute_script("""
                const r = arguments[0].getBoundingClientRect();
                return {
                    left: r.left,
                    top: r.top,
                    right: r.right,
                    bottom: r.bottom,
                    width: r.width,
                    height: r.height
                };
            """, el)

            if rect["width"] <= 0 or rect["height"] <= 0:
                continue

            centro_y = rect["top"] + rect["height"] / 2

            mesma_faixa = abs(centro_y - y_centro) <= 35
            a_direita = rect["left"] >= rect_ate["right"] + 120
            perto = rect["left"] <= rect_ate["right"] + 260
            tamanho_ok = 18 <= rect["width"] <= 90 and 18 <= rect["height"] <= 90

            if mesma_faixa and a_direita and perto and tamanho_ok:
                score = abs(centro_y - y_centro) + (rect["left"] - rect_ate["right"])
                if score < melhor_score:
                    melhor = el
                    melhor_score = score
        except Exception:
            pass

    if melhor is not None:
        print("Elemento candidato encontrado ao lado do 'Até'. Tentando clique...")

        try:
            ActionChains(driver).move_to_element(melhor).pause(0.2).click().perform()
            print("Botão clicado com ActionChains.")
            time.sleep(2)
            return
        except Exception:
            pass

        try:
            melhor.click()
            print("Botão clicado com clique normal.")
            time.sleep(2)
            return
        except Exception:
            pass

        try:
            driver.execute_script("arguments[0].click();", melhor)
            print("Botão clicado com JavaScript.")
            time.sleep(2)
            return
        except Exception:
            pass

    print("Não achei o botão por elemento. Tentando clique por coordenada...")

    x = rect_ate["right"] + 190
    y = y_centro

    sucesso = driver.execute_script("""
        const x = arguments[0];
        const y = arguments[1];

        const el = document.elementFromPoint(x, y);
        if (!el) return false;

        ['mousemove', 'mousedown', 'mouseup', 'click'].forEach(evtName => {
            el.dispatchEvent(new MouseEvent(evtName, {
                view: window,
                bubbles: true,
                cancelable: true,
                clientX: x,
                clientY: y,
                buttons: 1
            }));
        });

        return true;
    """, x, y)

    if not sucesso:
        raise Exception("Não consegui clicar no botão ao lado do texto 'Até'.")

    print(f"Botão clicado por coordenada em x={x:.0f}, y={y:.0f}")
    time.sleep(2)


# ================= DETECÇÃO DA GRADE =================
def grade_esta_visivel(driver):
    try:
        resultado = driver.execute_script("""
            function visivel(el){
                if (!el) return false;
                const st = getComputedStyle(el);
                const r = el.getBoundingClientRect();
                return st.visibility !== 'hidden' &&
                       st.display !== 'none' &&
                       r.width > 0 &&
                       r.height > 0;
            }

            const trs = [...document.querySelectorAll('tr')].filter(visivel).length;
            const tds = [...document.querySelectorAll('td')].filter(visivel).length;

            return trs >= 5 || tds >= 20;
        """)
        return bool(resultado)
    except Exception:
        return False


def aguardar_grade_pronta(driver, timeout=30):
    print("Aguardando grade ficar pronta...")
    wait = esperar(driver, timeout)
    wait.until(lambda d: grade_esta_visivel(d))
    time.sleep(1.5)
    print("Grade pronta.")


# ================= TELA CHEIA DA GRADE =================
def clicar_botao_tela_cheia_grade(driver, timeout=20):
    wait = esperar(driver, timeout)

    print("Tentando clicar no botão da grade (lado direito da barra verde)...")

    es = wait.until(
        EC.presence_of_element_located((
            By.XPATH,
            "//*[normalize-space(text())='ES']"
        ))
    )

    driver.execute_script(
        "arguments[0].scrollIntoView({block:'center', inline:'nearest'});",
        es
    )
    time.sleep(0.5)

    rect_es = driver.execute_script("""
        const r = arguments[0].getBoundingClientRect();
        return {
            left: r.left,
            top: r.top,
            right: r.right,
            bottom: r.bottom,
            width: r.width,
            height: r.height
        };
    """, es)

    largura = driver.execute_script("return window.innerWidth;")
    y_base = rect_es["top"] + (rect_es["height"] / 2)

    xs = [largura - i for i in [8, 12, 16, 20, 24, 28, 32, 36, 40, 44, 48, 52, 56, 60]]
    ys = [y_base, y_base - 6, y_base + 6, y_base - 12, y_base + 12, y_base - 18, y_base + 18]

    clicou = False

    for y in ys:
        for x in xs:
            sucesso = driver.execute_script("""
                const x = arguments[0];
                const y = arguments[1];

                const stack = document.elementsFromPoint(x, y);
                if (!stack || stack.length === 0) return false;

                function tentarClique(el, x, y) {
                    try {
                        ['mousemove', 'mousedown', 'mouseup', 'click'].forEach(evtName => {
                            el.dispatchEvent(new MouseEvent(evtName, {
                                view: window,
                                bubbles: true,
                                cancelable: true,
                                clientX: x,
                                clientY: y,
                                buttons: 1
                            }));
                        });
                        return true;
                    } catch (e) {
                        return false;
                    }
                }

                for (const el of stack) {
                    let atual = el;
                    let nivel = 0;

                    while (atual && nivel < 6) {
                        const tag = (atual.tagName || '').toLowerCase();
                        if (['button', 'a', 'div', 'span', 'i', 'svg'].includes(tag)) {
                            if (tentarClique(atual, x, y)) {
                                return true;
                            }
                        }
                        atual = atual.parentElement;
                        nivel++;
                    }
                }

                return false;
            """, x, y)

            time.sleep(0.7)

            if sucesso:
                clicou = True
                print("Clique de tela cheia disparado.")
                time.sleep(2.0)
                break
        if clicou:
            break

    if not clicou:
        raise Exception("Não consegui disparar o clique no botão de tela cheia da grade.")

    print("Clique da tela cheia executado. Aguardando grade estabilizar...")
    time.sleep(2.5)
    aguardar_grade_pronta(driver, timeout=30)


# ================= LOCALIZAR GRADE =================
def localizar_tabela_e_scroll_reais(driver):
    resultado = driver.execute_script("""
        function visivel(el){
            if (!el) return false;
            const st = getComputedStyle(el);
            const r = el.getBoundingClientRect();
            return st.visibility !== 'hidden' &&
                   st.display !== 'none' &&
                   r.width > 0 &&
                   r.height > 0;
        }

        const tables = [...document.querySelectorAll('table')].filter(visivel);

        function scoreTabela(table){
            const rows = [...table.querySelectorAll('tr')].filter(visivel);
            const cells = [...table.querySelectorAll('th, td')].filter(visivel);
            const rect = table.getBoundingClientRect();
            return rows.length * 100 + cells.length + Math.floor((rect.width * rect.height) / 1000);
        }

        let bestTable = null;
        let bestScore = -1;

        for (const table of tables) {
            const s = scoreTabela(table);
            if (s > bestScore) {
                bestScore = s;
                bestTable = table;
            }
        }

        if (!bestTable) {
            return {ok:false, erro:'Nenhuma tabela visível encontrada.'};
        }

        let scrollEl = null;
        let atual = bestTable;

        while (atual && atual != document.body) {
            const st = getComputedStyle(atual);
            const overflowY = st.overflowY;
            const overflowX = st.overflowX;

            const scrollVertical = atual.scrollHeight > atual.clientHeight + 20;
            const scrollHorizontal = atual.scrollWidth > atual.clientWidth + 20;

            const temEstiloScroll =
                overflowY === 'auto' || overflowY === 'scroll' ||
                overflowX === 'auto' || overflowX === 'scroll';

            if ((scrollVertical || scrollHorizontal) && (temEstiloScroll || scrollVertical || scrollHorizontal)) {
                scrollEl = atual;
                break;
            }

            atual = atual.parentElement;
        }

        if (!scrollEl) {
            atual = bestTable.parentElement;
            while (atual && atual !== document.body) {
                if (atual.scrollWidth > atual.clientWidth + 20 || atual.scrollHeight > atual.clientHeight + 20) {
                    scrollEl = atual;
                    break;
                }
                atual = atual.parentElement;
            }
        }

        if (!scrollEl) {
            scrollEl = document.scrollingElement || document.documentElement;
        }

        bestTable.setAttribute('data-qa-target-table', '1');
        scrollEl.setAttribute('data-qa-target-scroll', '1');

        return {
            ok: true,
            scrollTop: scrollEl.scrollTop,
            scrollLeft: scrollEl.scrollLeft,
            scrollHeight: scrollEl.scrollHeight,
            clientHeight: scrollEl.clientHeight,
            scrollWidth: scrollEl.scrollWidth,
            clientWidth: scrollEl.clientWidth
        };
    """)

    if not resultado or not resultado.get("ok"):
        raise Exception(resultado.get("erro", "Não consegui localizar a grade real."))

    print("Estrutura real da grade localizada:")
    print(resultado)
    return resultado


def obter_metricas_scroll_grade(driver):
    return driver.execute_script("""
        const scrollEl = document.querySelector('[data-qa-target-scroll="1"]');
        if (!scrollEl) return null;

        return {
            scrollTop: scrollEl.scrollTop,
            scrollLeft: scrollEl.scrollLeft,
            scrollHeight: scrollEl.scrollHeight,
            clientHeight: scrollEl.clientHeight,
            scrollWidth: scrollEl.scrollWidth,
            clientWidth: scrollEl.clientWidth
        };
    """)


def definir_scroll_grade(driver, top=None, left=None):
    driver.execute_script("""
        const scrollEl = document.querySelector('[data-qa-target-scroll="1"]');
        if (!scrollEl) return false;

        if (arguments[0] !== null && arguments[0] !== undefined) {
            scrollEl.scrollTop = arguments[0];
        }

        if (arguments[1] !== null && arguments[1] !== undefined) {
            scrollEl.scrollLeft = arguments[1];
        }

        return true;
    """, top, left)
    time.sleep(1.0)


# ================= EXTRAÇÃO =================
def coletar_amostra_linhas_visiveis(driver):
    dados = driver.execute_script("""
        function visivel(el){
            if (!el) return false;
            const st = getComputedStyle(el);
            const r = el.getBoundingClientRect();
            return st.visibility !== 'hidden' &&
                   st.display !== 'none' &&
                   r.width > 2 &&
                   r.height > 2;
        }

        function limpar(t){
            return (t || '').replace(/\\s+/g, ' ').trim();
        }

        const table = document.querySelector('[data-qa-target-table="1"]');
        if (!table) return {ok:false, erro:'Tabela alvo não encontrada.'};

        let trs = [...table.querySelectorAll('tbody tr')].filter(visivel);
        if (!trs.length) {
            trs = [...table.querySelectorAll('tr')].filter(visivel);
        }

        const rows = [];

        for (const tr of trs) {
            const tds = [...tr.children].filter(visivel);
            if (!tds.length) continue;

            const cells = tds.map((td, idx) => {
                const r = td.getBoundingClientRect();
                return {
                    idx: idx,
                    text: limpar(td.innerText || td.textContent),
                    left: r.left,
                    top: r.top,
                    width: r.width,
                    height: r.height
                };
            });

            rows.push(cells);
        }

        return {ok:true, rows};
    """)

    if not dados.get("ok"):
        raise Exception(dados.get("erro", "Falha ao coletar amostra de linhas visíveis."))

    return dados["rows"]


def descobrir_indices_colunas_por_heuristica(driver):
    amostra = coletar_amostra_linhas_visiveis(driver)

    if not amostra:
        raise Exception("Não consegui obter nenhuma linha visível da grade.")

    max_cols = max(len(r) for r in amostra if r)
    if max_cols < 8:
        raise Exception(f"A amostra retornou poucas colunas visíveis ({max_cols}).")

    melhor = None

    for i in range(max_cols - 2):
        score_nome = 0
        score_num1 = 0
        score_num2 = 0
        score_total = 0
        linhas_validas = 0
        exemplos = []

        for row in amostra:
            if len(row) <= i + 2:
                continue

            t0 = texto_limpo(row[i]["text"])
            t1 = texto_limpo(row[i + 1]["text"])
            t2 = texto_limpo(row[i + 2]["text"])

            s = 0

            if parece_nome_pessoa(t0):
                score_nome += 1
                s += 4

            if eh_numero_simples(t1):
                score_num1 += 1
                s += 2

            if eh_numero_simples(t2):
                score_num2 += 1
                s += 2

            if parece_nome_pessoa(t0) and eh_numero_simples(t1) and eh_numero_simples(t2):
                s += 4
                if len(exemplos) < 5:
                    exemplos.append((t0, t1, t2))

            if s > 0:
                linhas_validas += 1

            score_total += s

        candidato = {
            "idxAgente": i,
            "idxInstala": i + 1,
            "idxVisitada": i + 2,
            "score_nome": score_nome,
            "score_num1": score_num1,
            "score_num2": score_num2,
            "linhas_validas": linhas_validas,
            "score_total": score_total,
            "exemplos": exemplos
        }

        if melhor is None or candidato["score_total"] > melhor["score_total"]:
            melhor = candidato

    if not melhor:
        raise Exception("Não consegui montar nenhum candidato de colunas.")

    if melhor["score_total"] < 10:
        raise Exception(f"Heurística fraca demais para confiar: {melhor}")

    print("Colunas identificadas por heurística:")
    print(melhor)
    return melhor


def posicionar_grade_nas_colunas_alvo(driver):
    metricas = obter_metricas_scroll_grade(driver)
    if not metricas:
        raise Exception("Não consegui obter métricas de scroll da grade.")

    max_left = max(0, int(metricas["scrollWidth"] - metricas["clientWidth"]))
    passo = max(120, int(metricas["clientWidth"] * 0.22))
    tentativas = gerar_posicoes(0, max_left, passo)

    if max_left not in tentativas:
        tentativas.append(max_left)

    melhor_geral = None

    for left in tentativas:
        print(f"Tentando descobrir colunas por heurística em left={left}")
        definir_scroll_grade(driver, left=left, top=0)

        try:
            candidato = descobrir_indices_colunas_por_heuristica(driver)
            candidato["left"] = left

            if melhor_geral is None or candidato["score_total"] > melhor_geral["score_total"]:
                melhor_geral = candidato

            if candidato["score_total"] >= 18:
                print("Colunas alvo encontradas com boa confiança.")
                return candidato

        except Exception as e:
            print(f"Falhou em left={left}: {e}")

    if melhor_geral and melhor_geral["score_total"] >= 10:
        print("Usando melhor candidato encontrado:")
        print(melhor_geral)
        definir_scroll_grade(driver, left=melhor_geral["left"], top=0)
        return melhor_geral

    raise Exception("Não consegui identificar AGENTE / T. INSTALA / T. VISITADA por heurística.")


def ler_linhas_visiveis_somente_alvos(driver, idx_agente, idx_instala, idx_visitada):
    dados = driver.execute_script("""
        function visivel(el){
            if (!el) return false;
            const st = getComputedStyle(el);
            const r = el.getBoundingClientRect();
            return st.visibility !== 'hidden' &&
                   st.display !== 'none' &&
                   r.width > 2 &&
                   r.height > 2;
        }

        function limpar(t){
            return (t || '').replace(/\\s+/g, ' ').trim();
        }

        const idxAgente = arguments[0];
        const idxInstala = arguments[1];
        const idxVisitada = arguments[2];

        const table = document.querySelector('[data-qa-target-table="1"]');
        if (!table) return {ok:false, erro:'Tabela alvo não encontrada.'};

        let bodyRows = [...table.querySelectorAll('tbody tr')].filter(visivel);
        if (!bodyRows.length) {
            bodyRows = [...table.querySelectorAll('tr')].filter(visivel);
        }

        const rows = [];

        for (const tr of bodyRows) {
            const cells = [...tr.children].filter(visivel);
            if (!cells.length) continue;

            function ler(idx){
                if (!cells[idx]) return "";
                return limpar(cells[idx].innerText || cells[idx].textContent);
            }

            const agente = ler(idxAgente);
            const instala = ler(idxInstala);
            const visitada = ler(idxVisitada);

            if (!agente && !instala && !visitada) continue;

            rows.push({
                "AGENTE COMERCIAL": agente,
                "T. INSTALA": instala,
                "T. VISITADA": visitada
            });
        }

        return {ok:true, rows};
    """, idx_agente, idx_instala, idx_visitada)

    if not dados.get("ok"):
        raise Exception(dados.get("erro", "Falha ao ler linhas visíveis."))

    return dados["rows"]


def construir_chave_3_colunas(row):
    return " | ".join([
        texto_limpo(row.get("AGENTE COMERCIAL", "")),
        texto_limpo(row.get("T. INSTALA", "")),
        texto_limpo(row.get("T. VISITADA", "")),
    ])


def extrair_somente_tres_colunas(driver):
    print("\n================ EXTRAÇÃO DAS 3 COLUNAS ================\n")

    aguardar_grade_pronta(driver)
    localizar_tabela_e_scroll_reais(driver)

    dados_colunas = posicionar_grade_nas_colunas_alvo(driver)

    idx_agente = dados_colunas["idxAgente"]
    idx_instala = dados_colunas["idxInstala"]
    idx_visitada = dados_colunas["idxVisitada"]

    print(f"Índices escolhidos -> AGENTE={idx_agente}, INSTALA={idx_instala}, VISITADA={idx_visitada}")

    metricas = obter_metricas_scroll_grade(driver)
    if not metricas:
        raise Exception("Não consegui obter as métricas da grade.")

    max_top = max(0, int(metricas["scrollHeight"] - metricas["clientHeight"]))
    passo_vertical = max(180, int(metricas["clientHeight"] * 0.70))
    posicoes_v = gerar_posicoes(0, max_top, passo_vertical)

    print("Posições verticais:", posicoes_v)

    registros = {}
    sem_novidade = 0

    for bloco_idx, top in enumerate(posicoes_v, start=1):
        print(f"--- Bloco vertical {bloco_idx}/{len(posicoes_v)} | top={top} ---")
        definir_scroll_grade(driver, top=top)

        linhas = ler_linhas_visiveis_somente_alvos(
            driver,
            idx_agente,
            idx_instala,
            idx_visitada
        )

        print(f"Linhas visíveis: {len(linhas)}")

        qtd_antes = len(registros)

        for row in linhas:
            registro = {
                "AGENTE COMERCIAL": texto_limpo(row.get("AGENTE COMERCIAL", "")),
                "T. INSTALA": texto_limpo(row.get("T. INSTALA", "")),
                "T. VISITADA": texto_limpo(row.get("T. VISITADA", "")),
            }

            chave = construir_chave_3_colunas(registro)
            if not chave.replace(" | ", "").strip():
                continue

            registros[chave] = registro

        qtd_depois = len(registros)

        if qtd_depois == qtd_antes:
            sem_novidade += 1
        else:
            sem_novidade = 0

        if sem_novidade >= 3:
            print("Sem novidades em 3 blocos seguidos. Encerrando varredura.")
            break

    linhas_finais = list(registros.values())

    if not linhas_finais:
        raise Exception("A extração terminou sem consolidar nenhuma linha.")

    print(f"\nExtração concluída. Linhas consolidadas: {len(linhas_finais)}")
    return linhas_finais


# ================= TRATAMENTO DOS DADOS =================
def tratar_linhas(linhas):
    df = pd.DataFrame(linhas)

    colunas_base = ["AGENTE COMERCIAL", "T. INSTALA", "T. VISITADA"]
    for col in colunas_base:
        if col not in df.columns:
            df[col] = ""

    df = df[colunas_base].copy()

    df["AGENTE COMERCIAL"] = df["AGENTE COMERCIAL"].apply(texto_limpo)
    df["T. INSTALA"] = pd.to_numeric(df["T. INSTALA"], errors="coerce").fillna(0).astype(int)
    df["T. VISITADA"] = pd.to_numeric(df["T. VISITADA"], errors="coerce").fillna(0).astype(int)

    df = df[df["AGENTE COMERCIAL"] != ""].copy()

    # Normaliza para remover duplicadas de forma segura
    df["AGENTE_NORMALIZADO"] = df["AGENTE COMERCIAL"].apply(normalizar_nome_agente)

    # Agrupa duplicadas pelo agente
    df = (
        df.groupby("AGENTE_NORMALIZADO", as_index=False)
        .agg({
            "T. INSTALA": "sum",
            "T. VISITADA": "sum"
        })
    )

    # Usa o nome padronizado como nome final
    df["AGENTE COMERCIAL"] = df["AGENTE_NORMALIZADO"]

    # Cria coluna Faltam = Instala - Visitada
    df["Faltam"] = df["T. INSTALA"] - df["T. VISITADA"]

    # Evita número negativo em "Faltam"
    df["Faltam"] = df["Faltam"].clip(lower=0)

    # Cria coluna % EXECUTADO
    df["% EXECUTADO"] = 0.0
    mask = df["T. INSTALA"] > 0
    df.loc[mask, "% EXECUTADO"] = (
        (df.loc[mask, "T. VISITADA"] / df.loc[mask, "T. INSTALA"]) * 100
    ).round(2)

    # Reorganiza colunas
    df = df[["AGENTE COMERCIAL", "T. INSTALA", "T. VISITADA", "Faltam", "% EXECUTADO"]]

    # Ordena pelo nome do agente
    df = df.sort_values(by="% EXECUTADO", ascending=True).reset_index(drop=True)

    return df


def formatar_excel(caminho_xlsx):
    wb = load_workbook(caminho_xlsx)
    ws = wb.active

    fonte_negrito = Font(bold=True)
    fonte_branca_negrito = Font(bold=True, color="FFFFFF")
    alinhamento_centralizado = Alignment(horizontal="center", vertical="center")

    preenchimento_vermelho = PatternFill(fill_type="solid", fgColor="FF0000")
    preenchimento_verde = PatternFill(fill_type="solid", fgColor="00B050")

    borda_fina = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )

    col_idx_percentual = None

    for cell in ws[1]:
        if cell.value == "% EXECUTADO":
            col_idx_percentual = cell.column
            break

    for row in ws.iter_rows():
        for cell in row:
            cell.font = fonte_negrito
            cell.alignment = alinhamento_centralizado
            cell.border = borda_fina

    if col_idx_percentual is not None:
        for row in range(2, ws.max_row + 1):
            celula_percentual = ws.cell(row=row, column=col_idx_percentual)
            celula_percentual.number_format = '0"%"'

            valor = celula_percentual.value
            try:
                percentual = float(valor)
            except (TypeError, ValueError):
                continue

            if percentual == 0:
                for col in range(1, ws.max_column + 1):
                    celula = ws.cell(row=row, column=col)
                    celula.fill = preenchimento_vermelho
                    celula.font = fonte_branca_negrito
                    celula.border = borda_fina
            elif percentual == 100:
                for col in range(1, ws.max_column + 1):
                    celula = ws.cell(row=row, column=col)
                    celula.fill = preenchimento_verde
                    celula.font = fonte_negrito
                    celula.border = borda_fina

    # Ajuste automático de largura de colunas com folga extra
    for coluna in ws.columns:
        max_len = 0
        coluna_letra = get_column_letter(coluna[0].column)

        for cell in coluna:
            valor = "" if cell.value is None else str(cell.value)
            tamanho = len(valor)
            if tamanho > max_len:
                max_len = tamanho

        largura_ajustada = max_len + 4
        ws.column_dimensions[coluna_letra].width = largura_ajustada

    wb.save(caminho_xlsx)


# ================= SALVAR RESULTADO =================
def salvar_resultado(linhas, sufixo=None, periodo_inicio=None, periodo_fim=None, abrir_excel=False):
    garantir_pasta_saida()

    df = tratar_linhas(linhas)

    detalhe_periodo = ""
    if periodo_inicio and periodo_fim:
        ini_nome = data_para_nome_arquivo(periodo_inicio)
        fim_nome = data_para_nome_arquivo(periodo_fim)
        detalhe_periodo = f"_{ini_nome}" if ini_nome == fim_nome else f"_{ini_nome}_a_{fim_nome}"

    if sufixo == "AMERICANA":
        nome = f"Parcial_Americana{detalhe_periodo}"
    elif sufixo == "PIRACICABA":
        nome = f"Parcial_Piracicaba{detalhe_periodo}"
    else:
        nome = f"planejamento_tratado{detalhe_periodo}_{timestamp_str()}"

    caminho_xlsx = os.path.join(PASTA_SAIDA, f"{nome}.xlsx")

    # Salva somente o Excel tratado
    df.to_excel(caminho_xlsx, index=False)

    # Tenta formatar o Excel sem impedir a continuidade do fluxo
    try:
        formatar_excel(caminho_xlsx)
    except Exception as e:
        print(f"Aviso: não foi possível formatar o Excel '{caminho_xlsx}': {e}")

    print("\nArquivo salvo com sucesso:")
    print(f"Excel tratado: {os.path.abspath(caminho_xlsx)}")

    if abrir_excel:
        abrir_arquivo(caminho_xlsx)

    return caminho_xlsx


# ================= EXECUÇÃO POR BASE =================
def executar_fluxo_base(url, usuario, senha, nome_base, app=None, periodo_inicio=None, periodo_fim=None):
    print("\n" + "=" * 80)
    print(f"INICIANDO FLUXO DA BASE: {nome_base}")
    print("=" * 80)

    if app:
        app.set_status(f"Processando base {nome_base}")
        app.set_substatus("Abrindo navegador")

    driver = iniciar_driver()

    try:
        print("Abrindo portal...")
        driver.get(url)

        if app:
            app.set_substatus("Fazendo login")

        preencher_login(driver, usuario, senha)
        aguardar_tela_principal(driver)

        if app:
            app.set_substatus("Entrando em Planejamento / Operação")

        clicar_card_planejamento_operacao(driver)
        aguardar_tela_planejamento(driver)

        if app:
            app.set_substatus(f"Selecionando base {nome_base}")

        expandir_e_marcar_base(driver, nome_base)

        if periodo_inicio and periodo_fim:
            if app:
                app.set_substatus(f"Ajustando data {periodo_inicio} até {periodo_fim}")
            preencher_periodo_datas(driver, periodo_inicio, periodo_fim)

        if app:
            app.set_substatus("Confirmando filtro e preparando grade")

        clicar_botao_destacado_vermelho(driver)
        clicar_botao_tela_cheia_grade(driver)

        if app:
            app.set_substatus("Extraindo dados da grade")

        linhas = extrair_somente_tres_colunas(driver)

        if app:
            app.set_substatus("Gerando Excel")

        sufixo = slug_base(nome_base)
        caminho_xlsx = salvar_resultado(
            linhas,
            sufixo=sufixo,
            periodo_inicio=periodo_inicio,
            periodo_fim=periodo_fim
        )

        enviar_para_github(caminho_xlsx)

        print("\nFluxo concluído com sucesso.")
        print(f"Base: {nome_base}")
        print(f"Arquivo Excel: {os.path.abspath(caminho_xlsx)}")

        return {
            "base": nome_base,
            "periodo_inicio": periodo_inicio,
            "periodo_fim": periodo_fim,
            "xlsx": caminho_xlsx,
            "linhas": len(linhas)
        }

    finally:
        driver.quit()



import shutil
from datetime import datetime

def enviar_para_github(caminho_arquivo):
    try:
        repo_path = r"C:\Users\user\Desktop\trata_csv\painel-faturamento"

        destino_pasta = os.path.join(repo_path, "dashboard", "leitura")
        os.makedirs(destino_pasta, exist_ok=True)

        # 🔥 cria nome com timestamp (igual sistema de corte)
        nome_base = os.path.basename(caminho_arquivo).replace(".xlsx","")
        agora = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome = f"{nome_base}_{agora}.xlsx"

        destino = os.path.join(destino_pasta, nome)

        shutil.copy2(caminho_arquivo, destino)

        # 🔥 git completo
        subprocess.run(["git", "pull", "--rebase"], cwd=repo_path)
        subprocess.run(["git", "add", "."], cwd=repo_path)
        subprocess.run(["git", "commit", "-m", f"Auto leitura {nome}"], cwd=repo_path)
        subprocess.run(["git", "push"], cwd=repo_path)

        print(f"✅ Enviado para GitHub: {nome}")

    except Exception as e:
        print(f"Erro geral: {e}")


# ================= MAIN =================
def main(app=None, base_selecionada="AMBAS", data_inicio_txt=None, data_fim_txt=None):
    load_dotenv()

    url = os.getenv("CPFL_URL")
    usuario = os.getenv("CPFL_USUARIO")
    senha = os.getenv("CPFL_SENHA")

    if not url or not usuario or not senha:
        raise Exception("Verifique se CPFL_URL, CPFL_USUARIO e CPFL_SENHA estão no arquivo .env")

    hoje = datetime.now().strftime("%d/%m/%Y")
    data_inicio_txt = data_inicio_txt or hoje
    data_fim_txt = data_fim_txt or data_inicio_txt
    periodos = gerar_periodos_diarios(data_inicio_txt, data_fim_txt)

    if base_selecionada == "AMERICANA":
        bases = ["AMERICANA [B]"]
    elif base_selecionada == "PIRACICABA":
        bases = ["PIRACICABA [B]"]
    else:
        bases = ["AMERICANA [B]", "PIRACICABA [B]"]

    print("Períodos gerados para exportação diária:")
    for ini, fim in periodos:
        print(f"- {ini} até {fim}")

    resultados = []
    total_execucoes = len(bases) * len(periodos)
    contador = 0

    # Sequencial de propósito: como o navegador fica visível, isso facilita acompanhar
    # e evita duas janelas alterando filtros ao mesmo tempo.
    for periodo_inicio, periodo_fim in periodos:
        for base in bases:
            contador += 1
            try:
                if app:
                    app.set_status(f"Processando {contador}/{total_execucoes}")
                    app.set_substatus(f"{base} | {periodo_inicio} até {periodo_fim}")

                resultado = executar_fluxo_base(
                    url, usuario, senha, base,
                    periodo_inicio=periodo_inicio,
                    periodo_fim=periodo_fim,
                    app=app
                )
                resultados.append(resultado)

            except TimeoutException as e:
                print(f"\nTempo de espera excedido na base {base} no período {periodo_inicio} até {periodo_fim}.")
                print(f"Detalhes: {e}")
            except Exception as e:
                print(f"\nOcorreu um erro durante a automação da base {base} no período {periodo_inicio} até {periodo_fim}.")
                print(f"Detalhes: {e}")

    print("\n" + "=" * 80)
    print("RESUMO FINAL")
    print("=" * 80)

    arquivos = []
    if resultados:
        for r in resultados:
            caminho_abs = os.path.abspath(r["xlsx"])
            arquivos.append(caminho_abs)
            print(f"Base: {r['base']}")
            print(f"Período: {r.get('periodo_inicio')} até {r.get('periodo_fim')}")
            print(f"Linhas extraídas antes do tratamento: {r['linhas']}")
            print(f"Excel tratado: {caminho_abs}")
            print("-" * 80)
    else:
        print("Nenhuma base foi processada com sucesso.")

    if app:
        app.atualizar_arquivos(arquivos)
        app.set_status("Execução encerrada")
        app.set_substatus("Resumo final disponível no log")

    return {"resultados": resultados, "arquivos": arquivos}


# ================= EXTRAÇÃO COMPLETA POR TAREFA (V5) =================
# Esta versão substitui a extração antiga de 3 colunas.
# Agora a chave é a coluna TAREFA, para contar cada tarefa apenas uma vez,
# inclusive quando não houver AGENTE COMERCIAL preenchido.

COLUNAS_TAREFA_DESEJADAS = [
    "TAREFA",
    "STATUS",
    "UNIDADE",
    "DESCRIÇÃO",
    "TIPO",
    "MUNICÍPIO",
    "DT PREVISTA",
    "DT LIMITE",
    "DT PLANEJA",
    "AGENTE COMERCIAL",
    "T. INSTALA",
    "T. VISITADA",
    "T. TELEMED",
    "T. DISTRIB",
]


def normalizar_header_grade(txt):
    txt = texto_limpo(txt).upper()
    txt = txt.replace("\n", " ").replace("\r", " ")
    txt = " ".join(txt.split())

    mapa = {
        "TAREFA": "TAREFA",
        "STATUS": "STATUS",
        "UNIDADE": "UNIDADE",
        "DESCRIÇÃO": "DESCRIÇÃO",
        "DESCRICAO": "DESCRIÇÃO",
        "TIPO": "TIPO",
        "MUNICÍPIO": "MUNICÍPIO",
        "MUNICIPIO": "MUNICÍPIO",
        "DT PREVISTA": "DT PREVISTA",
        "DT LIMITE": "DT LIMITE",
        "DT PLANEJA": "DT PLANEJA",
        "DT PLANEJADA": "DT PLANEJA",
        "AGENTE COMERCIAL": "AGENTE COMERCIAL",
        "T. INSTALA": "T. INSTALA",
        "T INSTALA": "T. INSTALA",
        "T. VISITADA": "T. VISITADA",
        "T VISITADA": "T. VISITADA",
        "T. TELEMED": "T. TELEMED",
        "T TELEMED": "T. TELEMED",
        "T. DISTRIB": "T. DISTRIB",
        "T DISTRIB": "T. DISTRIB",
    }

    if txt in mapa:
        return mapa[txt]

    # Normalizações por aproximação, para o caso do cabeçalho vir quebrado em linhas.
    sem_ponto = txt.replace(".", "")
    if "AGENTE" in txt and "COMERCIAL" in txt:
        return "AGENTE COMERCIAL"
    if "INSTALA" in txt:
        return "T. INSTALA"
    if "VISITADA" in txt:
        return "T. VISITADA"
    if "TELEMED" in txt:
        return "T. TELEMED"
    if "DISTRIB" in txt:
        return "T. DISTRIB"
    if "PREVISTA" in txt:
        return "DT PREVISTA"
    if "LIMITE" in txt:
        return "DT LIMITE"
    if "PLANEJA" in txt:
        return "DT PLANEJA"
    if sem_ponto == "TAREFA":
        return "TAREFA"

    return txt


def obter_mapa_headers_visiveis(driver):
    dados = driver.execute_script("""
        function visivel(el){
            if (!el) return false;
            const st = getComputedStyle(el);
            const r = el.getBoundingClientRect();
            return st.visibility !== 'hidden' && st.display !== 'none' && r.width > 2 && r.height > 2;
        }
        function limpar(t){ return (t || '').replace(/\s+/g, ' ').trim(); }

        const table = document.querySelector('[data-qa-target-table="1"]');
        if (!table) return {ok:false, erro:'Tabela alvo não encontrada.'};

        let headerRows = [...table.querySelectorAll('thead tr')].filter(visivel);
        if (!headerRows.length) {
            headerRows = [...table.querySelectorAll('tr')].filter(visivel).slice(0, 3);
        }

        const headers = [];
        for (const tr of headerRows) {
            const cells = [...tr.children].filter(visivel);
            cells.forEach((cell, idx) => {
                const txt = limpar(cell.innerText || cell.textContent);
                if (txt) headers.push({idx: idx, text: txt});
            });
        }

        return {ok:true, headers};
    """)

    if not dados.get("ok"):
        raise Exception(dados.get("erro", "Falha ao mapear cabeçalhos."))

    mapa = {}
    for h in dados.get("headers", []):
        nome = normalizar_header_grade(h.get("text", ""))
        if nome in COLUNAS_TAREFA_DESEJADAS and nome not in mapa:
            mapa[nome] = int(h.get("idx", 0))

    return mapa


def mapa_colunas_fixo_pelo_print():
    # Fallback baseado no layout visto no print depois da confirmação:
    # 0 checkbox | 1 ícones | 2 TAREFA | 3 STATUS | 4 UNIDADE | 5 DESCRIÇÃO | ...
    # Esse fallback é usado quando o cabeçalho fica fora da tabela/fora do scroll real e o Selenium
    # não consegue ler os títulos, mas as células das linhas continuam vindo na ordem correta.
    return {
        "TAREFA": 2,
        "STATUS": 3,
        "UNIDADE": 4,
        "DESCRIÇÃO": 5,
        "TIPO": 6,
        "MUNICÍPIO": 7,
        "DT PREVISTA": 8,
        "DT LIMITE": 9,
        "DT PLANEJA": 10,
        "AGENTE COMERCIAL": 11,
        "T. INSTALA": 12,
        "T. VISITADA": 13,
        "T. TELEMED": 14,
        "T. DISTRIB": 15,
    }


def amostra_valida_para_mapa_fixo(driver, mapa):
    """Confere se a tela atual parece mesmo com o layout do print.
    A validação procura número de tarefa na coluna 2 e datas/valores nas colunas seguintes.
    """
    try:
        rows = coletar_amostra_linhas_visiveis(driver)
    except Exception as e:
        print(f"Não consegui validar mapa fixo por amostra: {e}")
        return False

    import re
    linhas_com_tarefa = 0
    linhas_com_status = 0
    linhas_com_totais = 0

    for row in rows[:30]:
        if len(row) <= 13:
            continue

        tarefa = texto_limpo(row[mapa["TAREFA"]].get("text", ""))
        status = texto_limpo(row[mapa["STATUS"]].get("text", "")).upper()
        instala = texto_limpo(row[mapa["T. INSTALA"]].get("text", ""))
        visitada = texto_limpo(row[mapa["T. VISITADA"]].get("text", ""))

        if re.search(r"\d{4,}", tarefa):
            linhas_com_tarefa += 1
        if any(p in status for p in ["FINAL", "PEND", "PARC", "ABERT", "EXEC"]):
            linhas_com_status += 1
        if eh_numero_simples(instala) or eh_numero_simples(visitada):
            linhas_com_totais += 1

    print(
        "Validação do mapa fixo pelo print -> "
        f"tarefas={linhas_com_tarefa}, status={linhas_com_status}, totais={linhas_com_totais}"
    )
    return linhas_com_tarefa >= 1 and (linhas_com_status >= 1 or linhas_com_totais >= 1)


def posicionar_grade_para_tarefas(driver):
    metricas = obter_metricas_scroll_grade(driver)
    if not metricas:
        raise Exception("Não consegui obter métricas de scroll da grade.")

    # Primeiro tenta pelo caminho ideal: ler cabeçalhos reais.
    max_left = max(0, int(metricas["scrollWidth"] - metricas["clientWidth"]))
    passo = max(160, int(metricas["clientWidth"] * 0.35))
    posicoes = gerar_posicoes(0, max_left, passo)

    melhor = None
    for left in posicoes:
        definir_scroll_grade(driver, left=left, top=0)
        mapa = obter_mapa_headers_visiveis(driver)
        score = 0
        for col in ["TAREFA", "STATUS", "DT PREVISTA", "AGENTE COMERCIAL", "T. INSTALA", "T. VISITADA"]:
            if col in mapa:
                score += 1
        print(f"Cabeçalhos visíveis em left={left}: {mapa} | score={score}")

        candidato = {"left": left, "mapa": mapa, "score": score}
        if melhor is None or candidato["score"] > melhor["score"]:
            melhor = candidato
        if score >= 6:
            print("Mapa de colunas de tarefa encontrado com boa confiança.")
            return candidato

    if melhor and melhor["score"] >= 4 and "TAREFA" in melhor["mapa"]:
        print("Usando melhor mapa de colunas encontrado:")
        print(melhor)
        definir_scroll_grade(driver, left=melhor["left"], top=0)
        return melhor

    # Fallback para o layout do print: cabeçalhos não foram lidos, mas as linhas estão na ordem correta.
    print("Cabeçalhos não vieram pelo DOM. Tentando fallback pelo layout do print em left=0...")
    definir_scroll_grade(driver, left=0, top=0)
    mapa_fixo = mapa_colunas_fixo_pelo_print()
    if amostra_valida_para_mapa_fixo(driver, mapa_fixo):
        print("Usando mapa fixo validado pelo conteúdo das linhas:")
        print(mapa_fixo)
        return {"left": 0, "mapa": mapa_fixo, "score": 6, "fallback": "layout_print"}

    raise Exception("Não consegui identificar as colunas principais da grade de tarefas nem validar o fallback pelo print.")


def ler_linhas_visiveis_tarefas(driver, mapa_colunas):
    dados = driver.execute_script("""
        function visivel(el){
            if (!el) return false;
            const st = getComputedStyle(el);
            const r = el.getBoundingClientRect();
            return st.visibility !== 'hidden' && st.display !== 'none' && r.width > 2 && r.height > 2;
        }
        function limpar(t){ return (t || '').replace(/\s+/g, ' ').trim(); }

        const mapa = arguments[0];
        const table = document.querySelector('[data-qa-target-table="1"]');
        if (!table) return {ok:false, erro:'Tabela alvo não encontrada.'};

        let rows = [...table.querySelectorAll('tbody tr')].filter(visivel);
        if (!rows.length) {
            rows = [...table.querySelectorAll('tr')].filter(visivel).slice(1);
        }

        const out = [];
        for (const tr of rows) {
            const cells = [...tr.children].filter(visivel);
            if (!cells.length) continue;

            const row = {};
            for (const [nome, idx] of Object.entries(mapa)) {
                row[nome] = cells[idx] ? limpar(cells[idx].innerText || cells[idx].textContent) : "";
            }
            out.push(row);
        }
        return {ok:true, rows: out};
    """, mapa_colunas)

    if not dados.get("ok"):
        raise Exception(dados.get("erro", "Falha ao ler linhas visíveis de tarefas."))

    return dados.get("rows", [])


def chave_tarefa(row):
    tarefa = texto_limpo(row.get("TAREFA", ""))
    # A tarefa pode aparecer com ícones ou cadeado; pega o primeiro número comprido.
    import re
    m = re.search(r"\d{4,}", tarefa)
    return m.group(0) if m else tarefa


def extrair_somente_tres_colunas(driver):
    print("\n================ EXTRAÇÃO COMPLETA POR TAREFA ================\n")

    aguardar_grade_pronta(driver)
    localizar_tabela_e_scroll_reais(driver)

    dados_mapa = posicionar_grade_para_tarefas(driver)
    mapa_colunas = dados_mapa["mapa"]
    print("Mapa final de colunas:")
    print(mapa_colunas)

    if "TAREFA" not in mapa_colunas:
        raise Exception("A coluna TAREFA não foi encontrada. Não dá para deduplicar as tarefas com segurança.")

    metricas = obter_metricas_scroll_grade(driver)
    if not metricas:
        raise Exception("Não consegui obter as métricas da grade.")

    max_top = max(0, int(metricas["scrollHeight"] - metricas["clientHeight"]))
    passo_vertical = 40 if int(metricas.get("clientHeight", 0)) < 120 else max(120, int(metricas["clientHeight"] * 0.55))
    posicoes_v = gerar_posicoes(0, max_top, passo_vertical)

    print("Posições verticais para varredura completa:", posicoes_v)

    registros = {}

    for bloco_idx, top in enumerate(posicoes_v, start=1):
        print(f"--- Bloco vertical {bloco_idx}/{len(posicoes_v)} | top={top} ---")
        definir_scroll_grade(driver, top=top)

        linhas = ler_linhas_visiveis_tarefas(driver, mapa_colunas)
        print(f"Linhas visíveis lidas: {len(linhas)}")

        for row in linhas:
            registro = {col: texto_limpo(row.get(col, "")) for col in COLUNAS_TAREFA_DESEJADAS}
            chave = chave_tarefa(registro)
            if not chave or not chave.isdigit():
                continue

            # Considera cada tarefa somente uma vez. Se ela aparecer duplicada,
            # mantém a primeira ocorrência completa encontrada.
            if chave not in registros:
                registro["TAREFA"] = chave
                registros[chave] = registro

    linhas_finais = list(registros.values())

    if not linhas_finais:
        raise Exception("A extração terminou sem consolidar nenhuma tarefa.")

    print(f"\nExtração concluída. Tarefas únicas consolidadas: {len(linhas_finais)}")
    return linhas_finais


def numero_int(valor):
    txt = texto_limpo(valor)
    if not txt:
        return 0
    import re
    m = re.search(r"-?\d+", txt.replace(".", ""))
    if not m:
        return 0
    try:
        return int(m.group(0))
    except Exception:
        return 0


def classificar_situacao(status, instala, visitada):
    st = texto_limpo(status).upper()
    if instala > 0:
        if visitada >= instala:
            return "FEITA"
        if visitada > 0:
            return "PARCIAL"
        return "PENDENTE"

    if "FINAL" in st or "CONCLU" in st:
        return "FEITA"
    if "PEND" in st or "ABERT" in st:
        return "PENDENTE"
    return "SEM TOTAL"


def tratar_linhas(linhas):
    df = pd.DataFrame(linhas)

    for col in COLUNAS_TAREFA_DESEJADAS:
        if col not in df.columns:
            df[col] = ""

    df = df[COLUNAS_TAREFA_DESEJADAS].copy()
    for col in df.columns:
        df[col] = df[col].apply(texto_limpo)

    df["TAREFA"] = df["TAREFA"].apply(lambda x: chave_tarefa({"TAREFA": x}))
    df = df[df["TAREFA"] != ""].copy()
    df = df.drop_duplicates(subset=["TAREFA"], keep="first")

    for col in ["T. INSTALA", "T. VISITADA", "T. TELEMED", "T. DISTRIB"]:
        df[col] = df[col].apply(numero_int)

    df["FALTAM"] = (df["T. INSTALA"] - df["T. VISITADA"]).clip(lower=0)
    df["% EXECUTADO"] = 0.0
    mask = df["T. INSTALA"] > 0
    df.loc[mask, "% EXECUTADO"] = ((df.loc[mask, "T. VISITADA"] / df.loc[mask, "T. INSTALA"]) * 100).round(2)
    df["SITUAÇÃO LEITURA"] = df.apply(
        lambda r: classificar_situacao(r.get("STATUS", ""), int(r.get("T. INSTALA", 0)), int(r.get("T. VISITADA", 0))),
        axis=1
    )

    ordem = [
        "TAREFA", "SITUAÇÃO LEITURA", "STATUS", "UNIDADE", "DESCRIÇÃO", "TIPO", "MUNICÍPIO",
        "DT PREVISTA", "DT LIMITE", "DT PLANEJA", "AGENTE COMERCIAL",
        "T. INSTALA", "T. VISITADA", "FALTAM", "% EXECUTADO", "T. TELEMED", "T. DISTRIB"
    ]
    df = df[ordem]
    df = df.sort_values(by=["SITUAÇÃO LEITURA", "% EXECUTADO", "TAREFA"], ascending=[True, True, True]).reset_index(drop=True)
    return df


def montar_resumo_tarefas(df):
    total_tarefas = len(df)
    total_instala = int(df["T. INSTALA"].sum()) if "T. INSTALA" in df else 0
    total_visitada = int(df["T. VISITADA"].sum()) if "T. VISITADA" in df else 0
    total_faltam = int(df["FALTAM"].sum()) if "FALTAM" in df else 0

    linhas = [
        {"Indicador": "TOTAL DE TAREFAS ÚNICAS", "Valor": total_tarefas},
        {"Indicador": "TOTAL T. INSTALA", "Valor": total_instala},
        {"Indicador": "TOTAL T. VISITADA", "Valor": total_visitada},
        {"Indicador": "TOTAL FALTAM", "Valor": total_faltam},
    ]

    if "SITUAÇÃO LEITURA" in df.columns:
        cont = df["SITUAÇÃO LEITURA"].value_counts().to_dict()
        for nome in ["FEITA", "PARCIAL", "PENDENTE", "SEM TOTAL"]:
            linhas.append({"Indicador": f"TAREFAS {nome}", "Valor": int(cont.get(nome, 0))})

    return pd.DataFrame(linhas)


def formatar_excel_tarefas(caminho_xlsx):
    wb = load_workbook(caminho_xlsx)

    fonte_negrito = Font(bold=True)
    fonte_branca_negrito = Font(bold=True, color="FFFFFF")
    alinhamento_centralizado = Alignment(horizontal="center", vertical="center", wrap_text=True)
    preenchimento_vermelho = PatternFill(fill_type="solid", fgColor="FF0000")
    preenchimento_amarelo = PatternFill(fill_type="solid", fgColor="FFF2CC")
    preenchimento_verde = PatternFill(fill_type="solid", fgColor="00B050")
    preenchimento_header = PatternFill(fill_type="solid", fgColor="1F4E78")
    borda_fina = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = fonte_branca_negrito
            cell.alignment = alinhamento_centralizado
            cell.fill = preenchimento_header
            cell.border = borda_fina

        situacao_col = None
        for cell in ws[1]:
            if cell.value == "SITUAÇÃO LEITURA":
                situacao_col = cell.column
                break

        for row in ws.iter_rows(min_row=2):
            situacao = ""
            if situacao_col:
                situacao = str(ws.cell(row=row[0].row, column=situacao_col).value or "").upper()
            for cell in row:
                cell.font = fonte_negrito
                cell.alignment = alinhamento_centralizado
                cell.border = borda_fina
                if situacao == "PENDENTE":
                    cell.fill = preenchimento_vermelho
                    cell.font = fonte_branca_negrito
                elif situacao == "PARCIAL":
                    cell.fill = preenchimento_amarelo
                elif situacao == "FEITA":
                    cell.fill = preenchimento_verde

        for coluna in ws.columns:
            max_len = 0
            letra = get_column_letter(coluna[0].column)
            for cell in coluna:
                valor = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(valor))
            ws.column_dimensions[letra].width = min(max_len + 4, 42)

        ws.freeze_panes = "A2"
        try:
            ws.auto_filter.ref = ws.dimensions
        except Exception:
            pass

    wb.save(caminho_xlsx)


def salvar_resultado(linhas, sufixo=None, periodo_inicio=None, periodo_fim=None, abrir_excel=False):
    garantir_pasta_saida()

    df = tratar_linhas(linhas)
    resumo = montar_resumo_tarefas(df)

    detalhe_periodo = ""
    if periodo_inicio and periodo_fim:
        ini_nome = data_para_nome_arquivo(periodo_inicio)
        fim_nome = data_para_nome_arquivo(periodo_fim)
        detalhe_periodo = f"_{ini_nome}" if ini_nome == fim_nome else f"_{ini_nome}_a_{fim_nome}"

    if sufixo == "AMERICANA":
        nome = f"Tarefas_Americana{detalhe_periodo}"
    elif sufixo == "PIRACICABA":
        nome = f"Tarefas_Piracicaba{detalhe_periodo}"
    else:
        nome = f"tarefas_tratadas{detalhe_periodo}_{timestamp_str()}"

    caminho_xlsx = os.path.join(PASTA_SAIDA, f"{nome}.xlsx")

    with pd.ExcelWriter(caminho_xlsx, engine="openpyxl") as writer:
        resumo.to_excel(writer, index=False, sheet_name="RESUMO")
        df.to_excel(writer, index=False, sheet_name="TAREFAS")

    try:
        formatar_excel_tarefas(caminho_xlsx)
    except Exception as e:
        print(f"Aviso: não foi possível formatar o Excel '{caminho_xlsx}': {e}")

    print("\nArquivo salvo com sucesso:")
    print(f"Excel tratado: {os.path.abspath(caminho_xlsx)}")
    print(f"Tarefas únicas: {len(df)}")
    try:
        print(resumo.to_string(index=False))
    except Exception:
        pass

    if abrir_excel:
        abrir_arquivo(caminho_xlsx)

    return caminho_xlsx



# ================= EXTRAÇÃO COMPLETA POR TAREFA (V6 RÁPIDA) =================
# Otimização principal:
# - reduz espera de scroll de 1s para ~0.08s;
# - lê a grade inteira em um único execute_async_script;
# - usa passo vertical dinâmico baseado na altura real das linhas;
# - tenta usar um container de scroll maior quando o detector pega um scroll pequeno demais.

SCROLL_PAUSA_RAPIDA = 0.08


def localizar_tabela_e_scroll_reais(driver):
    resultado = driver.execute_script("""
        function visivel(el){
            if (!el) return false;
            const st = getComputedStyle(el);
            const r = el.getBoundingClientRect();
            return st.visibility !== 'hidden' && st.display !== 'none' && r.width > 0 && r.height > 0;
        }

        const tables = [...document.querySelectorAll('table')].filter(visivel);
        function scoreTabela(table){
            const rows = [...table.querySelectorAll('tr')].filter(visivel);
            const cells = [...table.querySelectorAll('th, td')].filter(visivel);
            const rect = table.getBoundingClientRect();
            return rows.length * 100 + cells.length + Math.floor((rect.width * rect.height) / 1000);
        }

        let bestTable = null;
        let bestScore = -1;
        for (const table of tables) {
            const s = scoreTabela(table);
            if (s > bestScore) { bestScore = s; bestTable = table; }
        }
        if (!bestTable) return {ok:false, erro:'Nenhuma tabela visível encontrada.'};

        // Procura todos os ancestrais com scroll e escolhe o mais útil.
        // A versão anterior às vezes pegava um scroll de cabeçalho com clientHeight muito baixo.
        const candidatos = [];
        let atual = bestTable;
        while (atual && atual !== document.body && atual !== document.documentElement) {
            const st = getComputedStyle(atual);
            const r = atual.getBoundingClientRect();
            const scrollV = atual.scrollHeight > atual.clientHeight + 20;
            const scrollH = atual.scrollWidth > atual.clientWidth + 20;
            const overflowOk = ['auto','scroll'].includes(st.overflowY) || ['auto','scroll'].includes(st.overflowX);
            if ((scrollV || scrollH) && (overflowOk || scrollV || scrollH)) {
                candidatos.push({el: atual, score: (scrollV ? atual.clientHeight * 10 : 0) + atual.scrollHeight + r.height});
            }
            atual = atual.parentElement;
        }

        let scrollEl = null;
        if (candidatos.length) {
            candidatos.sort((a,b) => b.score - a.score);
            scrollEl = candidatos[0].el;
        }
        if (!scrollEl) scrollEl = document.scrollingElement || document.documentElement;

        bestTable.setAttribute('data-qa-target-table', '1');
        scrollEl.setAttribute('data-qa-target-scroll', '1');

        return {
            ok:true,
            scrollTop: scrollEl.scrollTop,
            scrollLeft: scrollEl.scrollLeft,
            scrollHeight: scrollEl.scrollHeight,
            clientHeight: scrollEl.clientHeight,
            scrollWidth: scrollEl.scrollWidth,
            clientWidth: scrollEl.clientWidth,
            tableRows: bestTable.querySelectorAll('tr').length
        };
    """)

    if not resultado or not resultado.get('ok'):
        raise Exception(resultado.get('erro', 'Não consegui localizar a grade real.'))

    print('Estrutura real da grade localizada (modo rápido):')
    print(resultado)
    return resultado


def definir_scroll_grade(driver, top=None, left=None, pausa=SCROLL_PAUSA_RAPIDA):
    driver.execute_script("""
        const scrollEl = document.querySelector('[data-qa-target-scroll="1"]');
        if (!scrollEl) return false;
        if (arguments[0] !== null && arguments[0] !== undefined) scrollEl.scrollTop = arguments[0];
        if (arguments[1] !== null && arguments[1] !== undefined) scrollEl.scrollLeft = arguments[1];
        return true;
    """, top, left)
    time.sleep(pausa)


def ler_todas_linhas_tarefas_rapido(driver, mapa_colunas):
    dados = driver.execute_async_script("""
        const mapa = arguments[0];
        const done = arguments[arguments.length - 1];

        function visivel(el){
            if (!el) return false;
            const st = getComputedStyle(el);
            const r = el.getBoundingClientRect();
            return st.visibility !== 'hidden' && st.display !== 'none' && r.width > 2 && r.height > 2;
        }
        function limpar(t){ return (t || '').replace(/\s+/g, ' ').trim(); }
        function sleep(ms){ return new Promise(resolve => setTimeout(resolve, ms)); }

        (async () => {
            const table = document.querySelector('[data-qa-target-table="1"]');
            const scrollEl = document.querySelector('[data-qa-target-scroll="1"]');
            if (!table || !scrollEl) {
                done({ok:false, erro:'Tabela ou scroll alvo não encontrado.'});
                return;
            }

            scrollEl.scrollLeft = 0;
            await sleep(80);

            function rowsVisiveis(){
                let rows = [...table.querySelectorAll('tbody tr')].filter(visivel);
                if (!rows.length) rows = [...table.querySelectorAll('tr')].filter(visivel).slice(1);
                return rows;
            }

            function lerBloco(){
                const out = [];
                for (const tr of rowsVisiveis()) {
                    const cells = [...tr.children].filter(visivel);
                    if (!cells.length) continue;
                    const row = {};
                    for (const [nome, idx] of Object.entries(mapa)) {
                        row[nome] = cells[idx] ? limpar(cells[idx].innerText || cells[idx].textContent) : '';
                    }
                    out.push(row);
                }
                return out;
            }

            scrollEl.scrollTop = 0;
            await sleep(120);

            const primeiras = rowsVisiveis();
            let alturaLinha = 32;
            if (primeiras.length >= 2) {
                const a = primeiras[0].getBoundingClientRect();
                const b = primeiras[1].getBoundingClientRect();
                const delta = Math.abs(b.top - a.top);
                if (delta >= 12 && delta <= 80) alturaLinha = delta;
            }

            // Passo um pouco menor que a janela útil para sobrepor linhas e não perder nada.
            const janela = Math.max(alturaLinha * 3, scrollEl.clientHeight || 400);
            const passo = Math.max(alturaLinha, Math.floor(janela - (alturaLinha * 2)));
            const maxTop = Math.max(0, scrollEl.scrollHeight - scrollEl.clientHeight);

            const registros = [];
            const visitados = new Set();
            let semMudanca = 0;
            let top = 0;
            let blocos = 0;

            while (top <= maxTop + alturaLinha) {
                scrollEl.scrollTop = Math.min(top, maxTop);
                await sleep(75);
                const realTop = Math.round(scrollEl.scrollTop);
                const chaveTop = String(realTop);
                if (!visitados.has(chaveTop)) {
                    visitados.add(chaveTop);
                    registros.push(...lerBloco());
                    blocos++;
                    semMudanca = 0;
                } else {
                    semMudanca++;
                }
                if (realTop >= maxTop) break;
                if (semMudanca >= 5) break;
                top = realTop + passo;
            }

            // Garante o rodapé.
            scrollEl.scrollTop = maxTop;
            await sleep(100);
            registros.push(...lerBloco());

            done({
                ok:true,
                rows: registros,
                blocos: blocos,
                rowHeight: alturaLinha,
                passo: passo,
                maxTop: maxTop,
                clientHeight: scrollEl.clientHeight,
                scrollHeight: scrollEl.scrollHeight
            });
        })().catch(e => done({ok:false, erro:String(e && e.stack ? e.stack : e)}));
    """, mapa_colunas)

    if not dados.get('ok'):
        raise Exception(dados.get('erro', 'Falha na leitura rápida da grade.'))
    return dados


def extrair_somente_tres_colunas(driver):
    print("\n================ EXTRAÇÃO COMPLETA POR TAREFA - MODO RÁPIDO ================\n")

    aguardar_grade_pronta(driver)
    localizar_tabela_e_scroll_reais(driver)

    dados_mapa = posicionar_grade_para_tarefas(driver)
    mapa_colunas = dados_mapa['mapa']
    print('Mapa final de colunas:')
    print(mapa_colunas)

    if 'TAREFA' not in mapa_colunas:
        raise Exception('A coluna TAREFA não foi encontrada. Não dá para deduplicar as tarefas com segurança.')

    print('Lendo todos os blocos da grade em modo rápido...')
    leitura = ler_todas_linhas_tarefas_rapido(driver, mapa_colunas)
    print(
        f"Leitura rápida: blocos={leitura.get('blocos')} | "
        f"linhas brutas={len(leitura.get('rows', []))} | "
        f"rowHeight={leitura.get('rowHeight')} | passo={leitura.get('passo')} | "
        f"clientHeight={leitura.get('clientHeight')} | scrollHeight={leitura.get('scrollHeight')}"
    )

    registros = {}
    for row in leitura.get('rows', []):
        registro = {col: texto_limpo(row.get(col, '')) for col in COLUNAS_TAREFA_DESEJADAS}
        chave = chave_tarefa(registro)
        if not chave or not chave.isdigit():
            continue
        if chave not in registros:
            registro['TAREFA'] = chave
            registros[chave] = registro

    linhas_finais = list(registros.values())
    if not linhas_finais:
        raise Exception('A extração terminou sem consolidar nenhuma tarefa.')

    print(f"\nExtração concluída. Tarefas únicas consolidadas: {len(linhas_finais)}")
    return linhas_finais



# ================= CLASSIFICAÇÃO D0 / D1 / D2 POR DIAS ÚTEIS + MUNICÍPIOS (V8) =================
# D0 = dia útil atual da leitura.
# D1 = último dia útil anterior.
# D2 = dois dias úteis anteriores, e assim por diante.
# A regra abaixo pula sábados, domingos e feriados nacionais brasileiros.
# Observação: feriados municipais/regionais ainda não entram automaticamente.

from datetime import date as _date_v7

MUNICIPIOS_AMERICANA = {
    "AME": "AMERICANA",
    "COS": "COSMÓPOLIS",
    "ELF": "ELIAS FAUSTO",
    "HOR": "HORTOLÂNDIA",
    "MTM": "MONTE MOR",
    "NOO": "NOVA ODESSA",
    "PAU": "PAULÍNIA",
    "SBO": "SANTA BÁRBARA DO OESTE",
    "SUM": "SUMARÉ",
}

MUNICIPIOS_CONHECIDOS = dict(MUNICIPIOS_AMERICANA)


def nome_municipio(codigo):
    cod = texto_limpo(codigo).upper()
    return MUNICIPIOS_CONHECIDOS.get(cod, cod if cod else "SEM MUNICÍPIO")


def nome_aba_excel(txt, usados=None):
    usados = usados if usados is not None else set()
    nome = texto_limpo(txt).upper()
    nome = re.sub(r"[\\/*?:\[\]]", "-", nome)
    nome = nome[:31] if nome else "ABA"
    base = nome
    i = 2
    while nome in usados:
        sufixo = f"_{i}"
        nome = (base[:31-len(sufixo)] + sufixo)
        i += 1
    usados.add(nome)
    return nome


def _pascoa_v7(ano):
    """Retorna a data da Páscoa pelo algoritmo de Meeus/Jones/Butcher."""
    a = ano % 19
    b = ano // 100
    c = ano % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    mes = (h + l - 7 * m + 114) // 31
    dia = ((h + l - 7 * m + 114) % 31) + 1
    return _date_v7(ano, mes, dia)


def feriados_nacionais_brasil(ano):
    pascoa = _pascoa_v7(ano)
    fixos = {
        _date_v7(ano, 1, 1),
        _date_v7(ano, 4, 21),
        _date_v7(ano, 5, 1),
        _date_v7(ano, 9, 7),
        _date_v7(ano, 10, 12),
        _date_v7(ano, 11, 2),
        _date_v7(ano, 11, 15),
        _date_v7(ano, 11, 20),
        _date_v7(ano, 12, 25),
    }
    moveis = {
        pascoa - timedelta(days=48),  # Carnaval segunda
        pascoa - timedelta(days=47),  # Carnaval terça
        pascoa - timedelta(days=2),   # Sexta-feira Santa
        pascoa + timedelta(days=60),  # Corpus Christi
    }
    return fixos | moveis


def eh_dia_util_v7(data):
    return data.weekday() < 5 and data not in feriados_nacionais_brasil(data.year)


def ultimo_dia_util_ate_v7(data):
    atual = data
    while not eh_dia_util_v7(atual):
        atual -= timedelta(days=1)
    return atual


def classificar_dia_operacional_v7(data_leitura, data_referencia=None):
    if data_referencia is None:
        data_referencia = datetime.now().date()

    data_referencia = ultimo_dia_util_ate_v7(data_referencia)

    if data_leitura > data_referencia:
        return "FUTURO"

    if not eh_dia_util_v7(data_leitura):
        return "NÃO ÚTIL"

    atual = data_referencia
    indice = 0
    limite = 0
    while atual >= data_leitura and limite < 1000:
        if atual == data_leitura:
            return f"D{indice}"
        atual -= timedelta(days=1)
        while not eh_dia_util_v7(atual):
            atual -= timedelta(days=1)
        indice += 1
        limite += 1

    return f"D{indice}"


def indice_d_v7(classe):
    txt = texto_limpo(classe).upper().replace(" ", "")
    if txt.startswith("D") and txt[1:].isdigit():
        return int(txt[1:])
    if txt == "FUTURO":
        return -1
    return 9999


def tratar_linhas(linhas, periodo_inicio=None, periodo_fim=None):
    df = pd.DataFrame(linhas)

    for col in COLUNAS_TAREFA_DESEJADAS:
        if col not in df.columns:
            df[col] = ""

    df = df[COLUNAS_TAREFA_DESEJADAS].copy()
    for col in df.columns:
        df[col] = df[col].apply(texto_limpo)

    df["TAREFA"] = df["TAREFA"].apply(lambda x: chave_tarefa({"TAREFA": x}))
    df = df[df["TAREFA"] != ""].copy()
    df = df.drop_duplicates(subset=["TAREFA"], keep="first")

    for col in ["T. INSTALA", "T. VISITADA", "T. TELEMED", "T. DISTRIB"]:
        df[col] = df[col].apply(numero_int)

    df["MUNICÍPIO"] = df["MUNICÍPIO"].apply(lambda x: texto_limpo(x).upper())
    df["MUNICÍPIO NOME"] = df["MUNICÍPIO"].apply(nome_municipio)

    df["FALTAM"] = (df["T. INSTALA"] - df["T. VISITADA"]).clip(lower=0)
    df["% EXECUTADO"] = 0.0
    mask = df["T. INSTALA"] > 0
    df.loc[mask, "% EXECUTADO"] = ((df.loc[mask, "T. VISITADA"] / df.loc[mask, "T. INSTALA"]) * 100).round(2)
    df["SITUAÇÃO LEITURA"] = df.apply(
        lambda r: classificar_situacao(r.get("STATUS", ""), int(r.get("T. INSTALA", 0)), int(r.get("T. VISITADA", 0))),
        axis=1
    )

    data_base_periodo = None
    if periodo_inicio:
        try:
            data_base_periodo = parse_data_br(periodo_inicio)
        except Exception:
            data_base_periodo = None

    if data_base_periodo is None:
        def _classe_linha(row):
            for campo in ["DT PREVISTA", "DT LIMITE", "DT PLANEJA"]:
                valor = texto_limpo(row.get(campo, ""))
                if valor:
                    try:
                        return classificar_dia_operacional_v7(parse_data_br(valor[:10]))
                    except Exception:
                        pass
            return "SEM DATA"
        df["D OPERACIONAL"] = df.apply(_classe_linha, axis=1)
    else:
        classe_d = classificar_dia_operacional_v7(data_base_periodo)
        df["D OPERACIONAL"] = classe_d
        df["DATA BASE D"] = data_base_periodo.strftime("%d/%m/%Y")

    ordem = [
        "D OPERACIONAL", "DATA BASE D", "MUNICÍPIO", "MUNICÍPIO NOME", "TAREFA", "SITUAÇÃO LEITURA", "STATUS",
        "UNIDADE", "DESCRIÇÃO", "TIPO", "DT PREVISTA", "DT LIMITE", "DT PLANEJA", "AGENTE COMERCIAL",
        "T. INSTALA", "T. VISITADA", "FALTAM", "% EXECUTADO", "T. TELEMED", "T. DISTRIB"
    ]
    for col in ordem:
        if col not in df.columns:
            df[col] = ""

    df = df[ordem]
    df["_D_ORDEM"] = df["D OPERACIONAL"].apply(indice_d_v7)
    df = df.sort_values(
        by=["MUNICÍPIO", "_D_ORDEM", "SITUAÇÃO LEITURA", "% EXECUTADO", "TAREFA"],
        ascending=[True, True, True, True, True]
    ).drop(columns=["_D_ORDEM"]).reset_index(drop=True)
    return df


def montar_resumo_tarefas(df):
    total_tarefas = len(df)
    total_instala = int(df["T. INSTALA"].sum()) if "T. INSTALA" in df else 0
    total_visitada = int(df["T. VISITADA"].sum()) if "T. VISITADA" in df else 0
    total_faltam = int(df["FALTAM"].sum()) if "FALTAM" in df else 0

    linhas = [
        {"Indicador": "TOTAL DE TAREFAS ÚNICAS", "Valor": total_tarefas},
        {"Indicador": "TOTAL T. INSTALA", "Valor": total_instala},
        {"Indicador": "TOTAL T. VISITADA", "Valor": total_visitada},
        {"Indicador": "TOTAL FALTAM", "Valor": total_faltam},
    ]

    if "D OPERACIONAL" in df.columns:
        cont_d = df["D OPERACIONAL"].value_counts().to_dict()
        for classe in sorted(cont_d.keys(), key=indice_d_v7):
            linhas.append({"Indicador": f"TAREFAS {classe}", "Valor": int(cont_d.get(classe, 0))})

    if "SITUAÇÃO LEITURA" in df.columns:
        cont = df["SITUAÇÃO LEITURA"].value_counts().to_dict()
        for nome in ["FEITA", "PARCIAL", "PENDENTE", "SEM TOTAL"]:
            linhas.append({"Indicador": f"TAREFAS {nome}", "Valor": int(cont.get(nome, 0))})

    if "MUNICÍPIO NOME" in df.columns:
        linhas.append({"Indicador": "", "Valor": ""})
        linhas.append({"Indicador": "RESUMO POR MUNICÍPIO", "Valor": ""})
        for municipio, grupo in df.groupby("MUNICÍPIO NOME", dropna=False):
            linhas.append({
                "Indicador": str(municipio),
                "Valor": f"TAREFAS: {len(grupo)} | INSTALA: {int(grupo['T. INSTALA'].sum())} | VISITADA: {int(grupo['T. VISITADA'].sum())} | FALTAM: {int(grupo['FALTAM'].sum())}"
            })

    if "D OPERACIONAL" in df.columns and "SITUAÇÃO LEITURA" in df.columns:
        pivot = (
            df.pivot_table(index="D OPERACIONAL", columns="SITUAÇÃO LEITURA", values="TAREFA", aggfunc="count", fill_value=0)
            .reset_index()
        )
        pivot["_D_ORDEM"] = pivot["D OPERACIONAL"].apply(indice_d_v7)
        pivot = pivot.sort_values("_D_ORDEM").drop(columns=["_D_ORDEM"])
        linhas.append({"Indicador": "", "Valor": ""})
        linhas.append({"Indicador": "RESUMO POR D E SITUAÇÃO", "Valor": ""})
        for _, r in pivot.iterrows():
            partes = []
            for col in ["FEITA", "PARCIAL", "PENDENTE", "SEM TOTAL"]:
                if col in pivot.columns:
                    partes.append(f"{col}: {int(r.get(col, 0))}")
            linhas.append({"Indicador": str(r["D OPERACIONAL"]), "Valor": " | ".join(partes)})

    return pd.DataFrame(linhas)


def montar_resumo_municipio(df):
    if df.empty:
        return pd.DataFrame()

    idx = ["MUNICÍPIO", "MUNICÍPIO NOME", "D OPERACIONAL"]
    resumo = (
        df.pivot_table(
            index=idx,
            columns="SITUAÇÃO LEITURA",
            values="TAREFA",
            aggfunc="count",
            fill_value=0
        )
        .reset_index()
    )

    for col in ["FEITA", "PARCIAL", "PENDENTE", "SEM TOTAL"]:
        if col not in resumo.columns:
            resumo[col] = 0

    soma = df.groupby(idx, as_index=False).agg({
        "TAREFA": "count",
        "T. INSTALA": "sum",
        "T. VISITADA": "sum",
        "FALTAM": "sum",
        "T. TELEMED": "sum",
        "T. DISTRIB": "sum",
    }).rename(columns={"TAREFA": "TOTAL TAREFAS"})

    resumo = resumo.merge(soma, on=idx, how="left")
    resumo["_D_ORDEM"] = resumo["D OPERACIONAL"].apply(indice_d_v7)
    resumo = resumo.sort_values(["MUNICÍPIO", "_D_ORDEM"]).drop(columns=["_D_ORDEM"])

    return resumo[[
        "MUNICÍPIO", "MUNICÍPIO NOME", "D OPERACIONAL", "TOTAL TAREFAS",
        "FEITA", "PARCIAL", "PENDENTE", "SEM TOTAL",
        "T. INSTALA", "T. VISITADA", "FALTAM", "T. TELEMED", "T. DISTRIB"
    ]]


def salvar_resultado(linhas, sufixo=None, periodo_inicio=None, periodo_fim=None, abrir_excel=False):
    garantir_pasta_saida()

    df = tratar_linhas(linhas, periodo_inicio=periodo_inicio, periodo_fim=periodo_fim)
    resumo = montar_resumo_tarefas(df)
    resumo_municipio = montar_resumo_municipio(df)

    detalhe_periodo = ""
    if periodo_inicio and periodo_fim:
        ini_nome = data_para_nome_arquivo(periodo_inicio)
        fim_nome = data_para_nome_arquivo(periodo_fim)
        detalhe_periodo = f"_{ini_nome}" if ini_nome == fim_nome else f"_{ini_nome}_a_{fim_nome}"

    if sufixo == "AMERICANA":
        nome = f"Tarefas_Americana{detalhe_periodo}"
    elif sufixo == "PIRACICABA":
        nome = f"Tarefas_Piracicaba{detalhe_periodo}"
    else:
        nome = f"tarefas_tratadas{detalhe_periodo}_{timestamp_str()}"

    caminho_xlsx = os.path.join(PASTA_SAIDA, f"{nome}.xlsx")

    usados = set()
    with pd.ExcelWriter(caminho_xlsx, engine="openpyxl") as writer:
        resumo.to_excel(writer, index=False, sheet_name=nome_aba_excel("RESUMO", usados))
        resumo_municipio.to_excel(writer, index=False, sheet_name=nome_aba_excel("RESUMO_MUNICIPIO", usados))
        df.to_excel(writer, index=False, sheet_name=nome_aba_excel("TAREFAS", usados))

        # Abas separadas por município. Cada aba mantém o D, situação e totais por tarefa.
        for codigo, grupo in df.groupby("MUNICÍPIO", dropna=False):
            municipio_nome = nome_municipio(codigo)
            aba = nome_aba_excel(f"{codigo}_{municipio_nome}", usados)
            grupo.drop(columns=[], errors="ignore").to_excel(writer, index=False, sheet_name=aba)

    try:
        formatar_excel_tarefas(caminho_xlsx)
    except Exception as e:
        print(f"Aviso: não foi possível formatar o Excel '{caminho_xlsx}': {e}")

    print("\nArquivo salvo com sucesso:")
    print(f"Excel tratado: {os.path.abspath(caminho_xlsx)}")
    print(f"Tarefas únicas: {len(df)}")
    try:
        print(resumo.to_string(index=False))
        print("\nResumo por município:")
        print(resumo_municipio.to_string(index=False))
    except Exception:
        pass

    if abrir_excel:
        abrir_arquivo(caminho_xlsx)

    return caminho_xlsx


def gerar_resumo_d_consolidado_v7(resultados):
    if not resultados:
        return None

    linhas_todas = []
    for r in resultados:
        caminho = r.get("xlsx")
        if not caminho or not os.path.exists(caminho):
            continue
        try:
            df = pd.read_excel(caminho, sheet_name="TAREFAS")
        except Exception as e:
            print(f"Aviso: não consegui ler '{caminho}' para consolidar D: {e}")
            continue
        df["BASE"] = slug_base(r.get("base", ""))
        df["PERÍODO"] = r.get("periodo_inicio", "")
        linhas_todas.append(df)

    if not linhas_todas:
        return None

    df_all = pd.concat(linhas_todas, ignore_index=True)
    if "D OPERACIONAL" not in df_all.columns:
        return None

    resumo_d = (
        df_all.pivot_table(
            index=["BASE", "D OPERACIONAL"],
            columns="SITUAÇÃO LEITURA",
            values="TAREFA",
            aggfunc="count",
            fill_value=0
        )
        .reset_index()
    )

    for col in ["FEITA", "PARCIAL", "PENDENTE", "SEM TOTAL"]:
        if col not in resumo_d.columns:
            resumo_d[col] = 0

    resumo_d["TOTAL TAREFAS"] = resumo_d[["FEITA", "PARCIAL", "PENDENTE", "SEM TOTAL"]].sum(axis=1)
    resumo_d["_D_ORDEM"] = resumo_d["D OPERACIONAL"].apply(indice_d_v7)
    resumo_d = resumo_d.sort_values(["BASE", "_D_ORDEM"]).drop(columns=["_D_ORDEM"])
    resumo_d = resumo_d[["BASE", "D OPERACIONAL", "TOTAL TAREFAS", "FEITA", "PARCIAL", "PENDENTE", "SEM TOTAL"]]

    resumo_municipio = pd.DataFrame()
    if "MUNICÍPIO" in df_all.columns:
        resumo_municipio = montar_resumo_municipio(df_all)
        if "BASE" in df_all.columns and not resumo_municipio.empty:
            resumo_municipio = (
                df_all.pivot_table(
                    index=["BASE", "MUNICÍPIO", "MUNICÍPIO NOME", "D OPERACIONAL"],
                    columns="SITUAÇÃO LEITURA",
                    values="TAREFA",
                    aggfunc="count",
                    fill_value=0
                ).reset_index()
            )
            for col in ["FEITA", "PARCIAL", "PENDENTE", "SEM TOTAL"]:
                if col not in resumo_municipio.columns:
                    resumo_municipio[col] = 0
            soma = df_all.groupby(["BASE", "MUNICÍPIO", "MUNICÍPIO NOME", "D OPERACIONAL"], as_index=False).agg({
                "TAREFA": "count", "T. INSTALA": "sum", "T. VISITADA": "sum", "FALTAM": "sum"
            }).rename(columns={"TAREFA": "TOTAL TAREFAS"})
            resumo_municipio = resumo_municipio.merge(soma, on=["BASE", "MUNICÍPIO", "MUNICÍPIO NOME", "D OPERACIONAL"], how="left")
            resumo_municipio["_D_ORDEM"] = resumo_municipio["D OPERACIONAL"].apply(indice_d_v7)
            resumo_municipio = resumo_municipio.sort_values(["BASE", "MUNICÍPIO", "_D_ORDEM"]).drop(columns=["_D_ORDEM"])

    detalhe = timestamp_str()
    caminho = os.path.join(PASTA_SAIDA, f"Resumo_D_por_base_{detalhe}.xlsx")
    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        resumo_d.to_excel(writer, index=False, sheet_name="RESUMO_D")
        if not resumo_municipio.empty:
            resumo_municipio.to_excel(writer, index=False, sheet_name="RESUMO_MUNICIPIO")
        df_all.to_excel(writer, index=False, sheet_name="TAREFAS_CONSOLIDADAS")

    try:
        formatar_excel_tarefas(caminho)
    except Exception as e:
        print(f"Aviso: não foi possível formatar o resumo consolidado '{caminho}': {e}")

    print("\nResumo consolidado D gerado:")
    print(os.path.abspath(caminho))
    try:
        print(resumo_d.to_string(index=False))
    except Exception:
        pass
    return caminho



# ================= GITHUB / PAINEL STREAMLIT =================
def localizar_repo_painel():
    """Localiza o repositório do painel para copiar os XLSX para dashboard/leitura."""
    candidatos = []
    for var in ["PAINEL_REPO_PATH", "GITHUB_REPO_PATH", "REPO_PAINEL_PATH"]:
        valor = os.getenv(var)
        if valor:
            candidatos.append(valor)

    candidatos.extend([
        r"C:\Users\user\Desktop\trata_csv\painel-faturamento",
        os.path.join(os.path.expanduser("~"), "Desktop", "trata_csv", "painel-faturamento"),
        os.getcwd(),
    ])

    for caminho in candidatos:
        if not caminho:
            continue
        caminho = os.path.abspath(os.path.expanduser(caminho))
        if os.path.isdir(os.path.join(caminho, ".git")):
            return caminho

    # fallback: usa o primeiro caminho tradicional, mesmo que ainda não exista .git,
    # para o erro ficar claro no log.
    return os.path.abspath(os.path.expanduser(candidatos[0]))


def _rodar_git(args, cwd):
    """Executa git e mostra stdout/stderr no log da interface."""
    cmd = ["git"] + list(args)
    print("$ " + " ".join(cmd))
    proc = subprocess.run(
        cmd,
        cwd=cwd,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        shell=False,
    )
    if proc.stdout.strip():
        print(proc.stdout.strip())
    if proc.stderr.strip():
        print(proc.stderr.strip())
    return proc


def _git_abortar_operacoes_pendentes(repo_path):
    """Remove estado de merge/rebase em andamento, quando existir."""
    _rodar_git(["rebase", "--abort"], cwd=repo_path)
    _rodar_git(["merge", "--abort"], cwd=repo_path)


def _git_sincronizar_branch(repo_path, branch="main"):
    """
    Deixa o repositório local exatamente na branch remota escolhida.

    Isso resolve os problemas vistos no log:
    - detached HEAD
    - branches divergentes
    - merge/rebase em conflito

    Observação: por segurança operacional do robô, alterações locais não commitadas
    no repositório do painel são descartadas antes de copiar os novos XLSX.
    """
    branch = (branch or "main").strip()

    _git_abortar_operacoes_pendentes(repo_path)

    fetch = _rodar_git(["fetch", "origin", branch], cwd=repo_path)
    if fetch.returncode != 0:
        print(f"⚠️ Não consegui executar git fetch origin {branch}.")
        return False

    # Cria/atualiza a branch local apontando exatamente para origin/main.
    checkout = _rodar_git(["checkout", "-B", branch, f"origin/{branch}"], cwd=repo_path)
    if checkout.returncode != 0:
        print(f"⚠️ Não consegui posicionar o repositório na branch {branch}.")
        return False

    reset = _rodar_git(["reset", "--hard", f"origin/{branch}"], cwd=repo_path)
    if reset.returncode != 0:
        print(f"⚠️ Não consegui sincronizar a branch {branch} com origin/{branch}.")
        return False

    return True


def enviar_para_github(caminho_arquivo):
    """
    Copia o arquivo gerado para dashboard/leitura do repositório do painel
    e envia para o GitHub.

    Esta versão é blindada contra:
    - detached HEAD
    - branch main divergente
    - pull/rebase em conflito
    - merge pendente

    O painel no Streamlit Cloud só enxerga arquivos que foram enviados ao GitHub.
    """
    try:
        if not caminho_arquivo or not os.path.exists(caminho_arquivo):
            print(f"⚠️ Arquivo não encontrado para enviar ao GitHub: {caminho_arquivo}")
            return None

        repo_path = localizar_repo_painel()
        if not os.path.isdir(repo_path):
            print(f"⚠️ Repositório do painel não encontrado: {repo_path}")
            return None
        if not os.path.isdir(os.path.join(repo_path, ".git")):
            print(f"⚠️ Pasta encontrada, mas não parece ser um repositório Git: {repo_path}")
            return None

        branch = os.getenv("PAINEL_GIT_BRANCH") or os.getenv("GITHUB_BRANCH") or "main"
        print(f"Sincronizando repositório do painel na branch {branch}...")

        if not _git_sincronizar_branch(repo_path, branch=branch):
            print("⚠️ Não foi possível sincronizar o repositório. O arquivo será salvo localmente, mas talvez não suba ao GitHub.")

        destino_pasta = os.path.join(repo_path, "dashboard", "leitura")
        os.makedirs(destino_pasta, exist_ok=True)

        nome_base = os.path.splitext(os.path.basename(caminho_arquivo))[0]
        agora = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_destino = f"{nome_base}_{agora}.xlsx"
        destino = os.path.join(destino_pasta, nome_destino)

        shutil.copy2(caminho_arquivo, destino)
        print(f"✅ Copiado para o painel: {destino}")

        add = _rodar_git(["add", "dashboard/leitura"], cwd=repo_path)
        if add.returncode != 0:
            print("⚠️ git add falhou. Arquivo copiado localmente, mas não commitado.")
            return destino

        status = _rodar_git(["status", "--porcelain", "dashboard/leitura"], cwd=repo_path)
        if status.returncode == 0 and not status.stdout.strip():
            print("ℹ️ Nenhuma alteração nova em dashboard/leitura para commitar.")
            return destino

        commit = _rodar_git(["commit", "-m", f"Auto leitura {nome_destino}"], cwd=repo_path)
        if commit.returncode != 0:
            saida = ((commit.stdout or "") + "\n" + (commit.stderr or "")).lower()
            if "nothing to commit" in saida or "nada a declarar" in saida or "nothing added to commit" in saida:
                print("ℹ️ Nada novo para commitar.")
                return destino
            print("⚠️ git commit falhou. Verifique o log acima.")
            return destino

        push = _rodar_git(["push", "origin", branch], cwd=repo_path)
        if push.returncode == 0:
            print(f"🚀 Enviado para GitHub: dashboard/leitura/{nome_destino}")
        else:
            print("⚠️ git push origin main falhou. Tentando fallback: git push origin HEAD:main")
            push2 = _rodar_git(["push", "origin", f"HEAD:{branch}"], cwd=repo_path)
            if push2.returncode == 0:
                print(f"🚀 Enviado para GitHub via fallback: dashboard/leitura/{nome_destino}")
            else:
                print("⚠️ git push falhou. Verifique autenticação/permissão do GitHub.")

        return destino

    except Exception as e:
        print(f"Erro ao enviar para GitHub: {e}")
        return None

def main(app=None, base_selecionada="AMBAS", data_inicio_txt=None, data_fim_txt=None):
    load_dotenv()

    url = os.getenv("CPFL_URL")
    usuario = os.getenv("CPFL_USUARIO")
    senha = os.getenv("CPFL_SENHA")

    if not url or not usuario or not senha:
        raise Exception("Verifique se CPFL_URL, CPFL_USUARIO e CPFL_SENHA estão no arquivo .env")

    hoje = datetime.now().strftime("%d/%m/%Y")
    data_inicio_txt = data_inicio_txt or hoje
    data_fim_txt = data_fim_txt or data_inicio_txt
    periodos = gerar_periodos_diarios(data_inicio_txt, data_fim_txt)

    if base_selecionada == "AMERICANA":
        bases = ["AMERICANA [B]"]
    elif base_selecionada == "PIRACICABA":
        bases = ["PIRACICABA [B]"]
    else:
        bases = ["AMERICANA [B]", "PIRACICABA [B]"]

    print("Períodos gerados para exportação diária:")
    for ini, fim in periodos:
        classe_d = classificar_dia_operacional_v7(parse_data_br(ini))
        print(f"- {ini} até {fim} => {classe_d}")

    resultados = []
    total_execucoes = len(bases) * len(periodos)
    contador = 0

    for periodo_inicio, periodo_fim in periodos:
        for base in bases:
            contador += 1
            try:
                classe_d = classificar_dia_operacional_v7(parse_data_br(periodo_inicio))
                if app:
                    app.set_status(f"Processando {contador}/{total_execucoes}")
                    app.set_substatus(f"{base} | {periodo_inicio} | {classe_d}")

                resultado = executar_fluxo_base(
                    url, usuario, senha, base,
                    periodo_inicio=periodo_inicio,
                    periodo_fim=periodo_fim,
                    app=app
                )
                resultado["d_operacional"] = classe_d
                resultados.append(resultado)

            except TimeoutException as e:
                print(f"\nTempo de espera excedido na base {base} no período {periodo_inicio} até {periodo_fim}.")
                print(f"Detalhes: {e}")
            except Exception as e:
                print(f"\nOcorreu um erro durante a automação da base {base} no período {periodo_inicio} até {periodo_fim}.")
                print(f"Detalhes: {e}")

    print("\n" + "=" * 80)
    print("RESUMO FINAL")
    print("=" * 80)

    arquivos = []
    if resultados:
        for r in resultados:
            caminho_abs = os.path.abspath(r["xlsx"])
            arquivos.append(caminho_abs)
            print(f"Base: {r['base']}")
            print(f"Período: {r.get('periodo_inicio')} até {r.get('periodo_fim')} | {r.get('d_operacional')}")
            print(f"Linhas extraídas antes do tratamento: {r['linhas']}")
            print(f"Excel tratado: {caminho_abs}")
            print("-" * 80)

        consolidado = gerar_resumo_d_consolidado_v7(resultados)
        if consolidado:
            arquivos.append(os.path.abspath(consolidado))
            enviar_para_github(consolidado)
    else:
        print("Nenhuma base foi processada com sucesso.")

    if app:
        app.atualizar_arquivos(arquivos)
        app.set_status("Execução encerrada")
        app.set_substatus("Resumo final disponível no log")

    return {"resultados": resultados, "arquivos": arquivos}



# ================= RESUMO CONSOLIDADO POR BASE + MUNICÍPIO (V9) =================
def montar_resumo_base_municipio_v9(df_base):
    """Monta uma aba com TOTAL da base por D e, abaixo, o detalhamento por município e D."""
    if df_base is None or df_base.empty:
        return pd.DataFrame()

    df = df_base.copy()
    for col in ["FEITA", "PARCIAL", "PENDENTE", "SEM TOTAL"]:
        # estes nomes são situações, não colunas do df original; ficam no pivot abaixo
        pass

    def _agrupar(idx):
        pivot = (
            df.pivot_table(
                index=idx,
                columns="SITUAÇÃO LEITURA",
                values="TAREFA",
                aggfunc="count",
                fill_value=0
            )
            .reset_index()
        )
        for col in ["FEITA", "PARCIAL", "PENDENTE", "SEM TOTAL"]:
            if col not in pivot.columns:
                pivot[col] = 0

        soma = df.groupby(idx, as_index=False).agg({
            "TAREFA": "count",
            "T. INSTALA": "sum",
            "T. VISITADA": "sum",
            "FALTAM": "sum",
        }).rename(columns={"TAREFA": "TOTAL TAREFAS"})

        out = pivot.merge(soma, on=idx, how="left")
        out["_D_ORDEM"] = out["D OPERACIONAL"].apply(indice_d_v7)
        return out.sort_values(idx[:-1] + ["_D_ORDEM"]).drop(columns=["_D_ORDEM"])

    total_d = _agrupar(["BASE", "D OPERACIONAL"])
    total_d.insert(1, "NÍVEL", "TOTAL BASE")
    total_d.insert(2, "MUNICÍPIO", "TOTAL")
    total_d.insert(3, "MUNICÍPIO NOME", "TOTAL DA BASE")

    mun_d = _agrupar(["BASE", "MUNICÍPIO", "MUNICÍPIO NOME", "D OPERACIONAL"])
    mun_d.insert(1, "NÍVEL", "MUNICÍPIO")

    colunas = [
        "BASE", "NÍVEL", "MUNICÍPIO", "MUNICÍPIO NOME", "D OPERACIONAL",
        "TOTAL TAREFAS", "FEITA", "PARCIAL", "PENDENTE", "SEM TOTAL",
        "T. INSTALA", "T. VISITADA", "FALTAM"
    ]
    for c in colunas:
        if c not in total_d.columns:
            total_d[c] = ""
        if c not in mun_d.columns:
            mun_d[c] = ""

    total_d = total_d[colunas]
    mun_d = mun_d[colunas]

    # Linha em branco para facilitar leitura no Excel.
    branco = pd.DataFrame([{c: "" for c in colunas}])
    return pd.concat([total_d, branco, mun_d], ignore_index=True)


def gerar_resumo_d_consolidado_v7(resultados):
    """V9: gera consolidado com RESUMO_GERAL, DETALHE_MUNICIPIO e uma aba por base."""
    if not resultados:
        return None

    linhas_todas = []
    for r in resultados:
        caminho = r.get("xlsx")
        if not caminho or not os.path.exists(caminho):
            continue
        try:
            df = pd.read_excel(caminho, sheet_name="TAREFAS")
        except Exception as e:
            print(f"Aviso: não consegui ler '{caminho}' para consolidar D: {e}")
            continue
        df["BASE"] = slug_base(r.get("base", ""))
        df["PERÍODO"] = r.get("periodo_inicio", "")
        linhas_todas.append(df)

    if not linhas_todas:
        return None

    df_all = pd.concat(linhas_todas, ignore_index=True)
    if "D OPERACIONAL" not in df_all.columns:
        return None

    # Resumo geral por base e D, igual ao que já funcionou no teste.
    resumo_d = (
        df_all.pivot_table(
            index=["BASE", "D OPERACIONAL"],
            columns="SITUAÇÃO LEITURA",
            values="TAREFA",
            aggfunc="count",
            fill_value=0
        )
        .reset_index()
    )
    for col in ["FEITA", "PARCIAL", "PENDENTE", "SEM TOTAL"]:
        if col not in resumo_d.columns:
            resumo_d[col] = 0
    resumo_d["TOTAL TAREFAS"] = resumo_d[["FEITA", "PARCIAL", "PENDENTE", "SEM TOTAL"]].sum(axis=1)
    resumo_d["_D_ORDEM"] = resumo_d["D OPERACIONAL"].apply(indice_d_v7)
    resumo_d = resumo_d.sort_values(["BASE", "_D_ORDEM"]).drop(columns=["_D_ORDEM"])
    resumo_d = resumo_d[["BASE", "D OPERACIONAL", "TOTAL TAREFAS", "FEITA", "PARCIAL", "PENDENTE", "SEM TOTAL"]]

    # Detalhe por município e D, com instalações/visitadas/faltam.
    detalhe_municipio = pd.DataFrame()
    if all(c in df_all.columns for c in ["MUNICÍPIO", "MUNICÍPIO NOME"]):
        detalhe_municipio = (
            df_all.pivot_table(
                index=["BASE", "MUNICÍPIO", "MUNICÍPIO NOME", "D OPERACIONAL"],
                columns="SITUAÇÃO LEITURA",
                values="TAREFA",
                aggfunc="count",
                fill_value=0
            )
            .reset_index()
        )
        for col in ["FEITA", "PARCIAL", "PENDENTE", "SEM TOTAL"]:
            if col not in detalhe_municipio.columns:
                detalhe_municipio[col] = 0
        soma = df_all.groupby(["BASE", "MUNICÍPIO", "MUNICÍPIO NOME", "D OPERACIONAL"], as_index=False).agg({
            "TAREFA": "count",
            "T. INSTALA": "sum",
            "T. VISITADA": "sum",
            "FALTAM": "sum",
        }).rename(columns={"TAREFA": "TOTAL TAREFAS"})
        detalhe_municipio = detalhe_municipio.merge(
            soma,
            on=["BASE", "MUNICÍPIO", "MUNICÍPIO NOME", "D OPERACIONAL"],
            how="left"
        )
        detalhe_municipio["_D_ORDEM"] = detalhe_municipio["D OPERACIONAL"].apply(indice_d_v7)
        detalhe_municipio = detalhe_municipio.sort_values(["BASE", "MUNICÍPIO", "_D_ORDEM"]).drop(columns=["_D_ORDEM"])
        detalhe_municipio = detalhe_municipio[[
            "BASE", "MUNICÍPIO", "MUNICÍPIO NOME", "D OPERACIONAL", "TOTAL TAREFAS",
            "FEITA", "PARCIAL", "PENDENTE", "SEM TOTAL", "T. INSTALA", "T. VISITADA", "FALTAM"
        ]]

    detalhe = timestamp_str()
    caminho = os.path.join(PASTA_SAIDA, f"Resumo_D_por_base_municipio_{detalhe}.xlsx")
    usados = set()
    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        resumo_d.to_excel(writer, index=False, sheet_name=nome_aba_excel("RESUMO_GERAL", usados))
        if not detalhe_municipio.empty:
            detalhe_municipio.to_excel(writer, index=False, sheet_name=nome_aba_excel("DETALHE_MUNICIPIO", usados))

        # Uma aba para cada base: primeiro o TOTAL da base por D, depois cada município por D.
        for base_nome, df_base in df_all.groupby("BASE", dropna=False):
            aba_base = montar_resumo_base_municipio_v9(df_base)
            aba_base.to_excel(writer, index=False, sheet_name=nome_aba_excel(str(base_nome), usados))

        df_all.to_excel(writer, index=False, sheet_name=nome_aba_excel("TAREFAS_CONSOLIDADAS", usados))

    try:
        formatar_excel_tarefas(caminho)
    except Exception as e:
        print(f"Aviso: não foi possível formatar o resumo consolidado '{caminho}': {e}")

    print("\nResumo consolidado por base e município gerado:")
    print(os.path.abspath(caminho))
    try:
        print(resumo_d.to_string(index=False))
        if not detalhe_municipio.empty:
            print("\nDetalhe por município:")
            print(detalhe_municipio.to_string(index=False))
    except Exception:
        pass
    return caminho

if __name__ == "__main__":
    iniciar_interface()
