import os
import sys
import time
import queue
import threading
import traceback
import subprocess
from pathlib import Path
from datetime import datetime, timedelta
import calendar
import json
import base64
import urllib.request
import urllib.error
from io import StringIO

from dotenv import load_dotenv, set_key

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import tkinter as tk
from tkinter import ttk, messagebox
from PIL import Image, ImageTk


APP_TITULO = "Extrator de Produção"
BASE_DIR = Path(__file__).resolve().parent
ENV_PATH = BASE_DIR / ".env"
# Carrega o .env logo no início para GitHub, Oracle e outras configurações.
# Antes, algumas configs do GitHub eram lidas antes do .env carregar.
load_dotenv(ENV_PATH, override=True)

APP_ICON_PATH = BASE_DIR / "logo_app.ico"
APP_LOGO_PATH = BASE_DIR / "logo.png"

import glob
import pandas as pd
import unicodedata
import re
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


PROCESS_INPUT_FOLDER = BASE_DIR / "input"
PROCESS_OUTPUT_FOLDER = BASE_DIR / "output"
PROCESS_OUTPUT_CSV = PROCESS_OUTPUT_FOLDER / "resultado_final.csv"
PROCESS_OUTPUT_TXT = PROCESS_OUTPUT_FOLDER / "resultado_final.txt"
PROCESS_OUTPUT_INATIVIDADE = PROCESS_OUTPUT_FOLDER / "inatividade_equipes.xlsx"
DOWNLOADS_DIR = Path.home() / "Downloads"
PROCESS_CACHE_FOLDER = BASE_DIR / "cache_csvs"
UI_STATE_PATH = BASE_DIR / "ui_state.json"

# Configuração do envio automático dos CSVs do dashboard para o GitHub.
# O token NÃO fica escrito no código: ele deve ser salvo no arquivo .env ou nas variáveis do Windows.
GITHUB_DASHBOARD_REPO = os.getenv("GITHUB_DASHBOARD_REPO", "irensegabriel/painel-faturamento")
GITHUB_DASHBOARD_BRANCH = os.getenv("GITHUB_DASHBOARD_BRANCH", "main")
GITHUB_DASHBOARD_REMOTE_DIR = os.getenv("GITHUB_DASHBOARD_REMOTE_DIR", "dashboard")
GITHUB_DASHBOARD_TOKEN_ENV_NAMES = ("GITHUB_DASHBOARD_TOKEN", "GITHUB_TOKEN")



def normalizar_texto_processamento(s):
    if not isinstance(s, str):
        return ""
    s = unicodedata.normalize("NFKD", s).encode("ASCII", "ignore").decode()
    return re.sub(r"[^a-zA-Z0-9]", "", s).lower()




EA_VARIANTES_PROCESSAMENTO = {
    "AVARE": ["AVARE"],
    "ITAPETININGA": ["ITAPETININGA"],
    "JAGUARIUNA": ["JAGUARIUNA"],
    "MOCOCA": ["MOCOCA"],
    "OURINHOS": ["OURINHOS"],
    "PEDREIRA": ["PEDREIRA"],
    "PIRAJU": ["PIRAJU"],
    "SANTA CRUZ DO RIO PARDO": ["SANTA CRUZ DO RIO PARDO"],
    "CERQUEIRA CESAR": ["CERQUEIRA CESAR", "CERQUEIRA CÉSAR"],
    "CASA BRANCA": ["CASA BRANCA"],
    "CAMPO LIMPO PAULISTA": ["CAMPO LIMPO PAULISTA"],
    "EA ITUPEVA": ["EA ITUPEVA", "ITUPEVA"],
    "EA VINHEDO": ["EA VINHEDO", "VINHEDO"],
    "EA JUNDIA": ["EA JUNDIA", "JUNDIA", "JUNDIAI"],
    "EA INDAIATUBA": ["EA INDAIATUBA", "INDAIATUBA"],
    "EA SALTO": ["EA SALTO", "SALTO"],
    "SAO JOSE DO RIO PARDO": ["SAO JOSE DO RIO PARDO", "SÃO JOSÉ DO RIO PARDO"],
    "SARAPUI": ["SARAPUI", "SARAPUÍ"],
    "SAO MIGUEL ARCAJNO": ["SAO MIGUEL ARCAJNO", "SAO MIGUEL ARCANJO", "SÃO MIGUEL ARCANJO", "SAO MIGUEL ARCANJ0"]
}


def obter_variantes_ea_processamento(ea):
    variantes = EA_VARIANTES_PROCESSAMENTO.get(ea, [ea])
    normalizadas = []
    vistos = set()
    for nome in variantes:
        nome_norm = normalizar_texto_processamento(nome).replace("ea", "")
        if nome_norm and nome_norm not in vistos:
            vistos.add(nome_norm)
            normalizadas.append(nome_norm)
    return normalizadas


def extrair_data_do_nome_arquivo(nome_arquivo):
    m = re.search(r"(\d{1,2})[_\-](\d{1,2})[_\-](\d{2,4})", nome_arquivo)
    if not m:
        return None

    dia, mes, ano = map(int, m.groups())
    if ano < 100:
        ano += 2000

    try:
        return datetime(ano, mes, dia).date()
    except ValueError:
        return None



def copiar_para_cache_csv(arquivo_origem, logger=None):
    logger = logger or log
    PROCESS_CACHE_FOLDER.mkdir(exist_ok=True)
    destino_cache = PROCESS_CACHE_FOLDER / arquivo_origem.name
    try:
        if not destino_cache.exists() or arquivo_origem.stat().st_mtime > destino_cache.stat().st_mtime:
            shutil.copy2(arquivo_origem, destino_cache)
            logger(f"🗃️ Cache atualizado: {arquivo_origem.name}")
        else:
            logger(f"🗃️ Cache reutilizado: {arquivo_origem.name}")
    except Exception as e:
        logger(f"⚠️ Falha ao atualizar cache de {arquivo_origem.name}: {e}")
    return destino_cache


def iterar_csvs_disponiveis():
    vistos = set()
    if PROCESS_CACHE_FOLDER.exists():
        for f in PROCESS_CACHE_FOLDER.glob("Atividades-*Terceiro_*.csv"):
            vistos.add(f.name)
            yield f
    for f in DOWNLOADS_DIR.glob("Atividades-*Terceiro_*.csv"):
        if f.name in vistos:
            cache_file = PROCESS_CACHE_FOLDER / f.name
            try:
                if (not cache_file.exists()) or (f.stat().st_mtime > cache_file.stat().st_mtime):
                    yield f
            except Exception:
                yield f
        else:
            yield f


def obter_periodo_atual(tipo_periodo):
    hoje = datetime.now().date()

    if tipo_periodo == "dia":
        return hoje, hoje

    if tipo_periodo == "semana":
        inicio = hoje - timedelta(days=hoje.weekday())
        fim = inicio + timedelta(days=5)  # segunda a sábado
        return inicio, fim

    if tipo_periodo == "mes":
        inicio = hoje.replace(day=1)
        ultimo_dia = calendar.monthrange(hoje.year, hoje.month)[1]
        fim = hoje.replace(day=ultimo_dia)
        return inicio, fim

    if tipo_periodo == "mes_anterior":
        primeiro_dia_mes_atual = hoje.replace(day=1)
        ultimo_dia_mes_anterior = primeiro_dia_mes_atual - timedelta(days=1)
        inicio = ultimo_dia_mes_anterior.replace(day=1)
        fim = ultimo_dia_mes_anterior
        return inicio, fim

    raise ValueError(f"Tipo de período inválido: {tipo_periodo}")


def _nome_arquivo_periodo(tipo_periodo):
    if tipo_periodo == "dia":
        return {
            "csv": PROCESS_OUTPUT_CSV,
            "txt": PROCESS_OUTPUT_TXT,
            "inatividade": PROCESS_OUTPUT_INATIVIDADE,
        }

    hoje = datetime.now()
    if tipo_periodo == "semana":
        sufixo = f"semana_{hoje.strftime('%Y_%m_%d')}"
    elif tipo_periodo == "mes":
        sufixo = f"mes_{hoje.strftime('%Y_%m')}"
    elif tipo_periodo == "mes_anterior":
        primeiro_dia_mes_atual = hoje.replace(day=1)
        ultimo_dia_mes_anterior = primeiro_dia_mes_atual - timedelta(days=1)
        sufixo = f"fechamento_mes_anterior_{ultimo_dia_mes_anterior.strftime('%Y_%m')}"
    else:
        raise ValueError(f"Tipo de período inválido: {tipo_periodo}")

    return {
        "csv": PROCESS_OUTPUT_FOLDER / f"resultado_final_{sufixo}.csv",
        "txt": PROCESS_OUTPUT_FOLDER / f"resultado_final_{sufixo}.txt",
        "inatividade": PROCESS_OUTPUT_FOLDER / f"inatividade_equipes_{sufixo}.xlsx",
    }


def obter_saida_mes_anterior():
    hoje = datetime.now().date()
    primeiro_dia_mes_atual = hoje.replace(day=1)
    ultimo_dia_mes_anterior = primeiro_dia_mes_atual - timedelta(days=1)
    return {
        "txt": PROCESS_OUTPUT_FOLDER / f"resultado_final_fechamento_mes_anterior_{ultimo_dia_mes_anterior.strftime('%Y_%m')}.txt",
        "csv": PROCESS_OUTPUT_FOLDER / f"resultado_final_fechamento_mes_anterior_{ultimo_dia_mes_anterior.strftime('%Y_%m')}.csv",
        "inatividade": PROCESS_OUTPUT_FOLDER / f"inatividade_equipes_fechamento_mes_anterior_{ultimo_dia_mes_anterior.strftime('%Y_%m')}.xlsx",
    }


def obter_saida_medicao_disjuntor_jundiai(tipo_periodo):
    hoje = datetime.now().date()
    if tipo_periodo == "mes":
        sufixo = f"mes_{hoje.strftime('%Y_%m')}"
    elif tipo_periodo == "mes_anterior":
        primeiro_dia_mes_atual = hoje.replace(day=1)
        ultimo_dia_mes_anterior = primeiro_dia_mes_atual - timedelta(days=1)
        sufixo = f"fechamento_mes_anterior_{ultimo_dia_mes_anterior.strftime('%Y_%m')}"
    elif tipo_periodo == "semana":
        sufixo = f"semana_{hoje.strftime('%Y_%m_%d')}"
    elif tipo_periodo == "dia":
        sufixo = hoje.strftime('%Y_%m_%d')
    else:
        sufixo = tipo_periodo
    return PROCESS_OUTPUT_FOLDER / f"calculo_medicao_disjuntor_jundiai_{sufixo}.xlsx"


def obter_saida_medicao_por_csv(caminho_csv):
    caminho_csv = Path(caminho_csv)
    nome_base = caminho_csv.stem
    if nome_base.startswith("resultado_final_"):
        nome_base = nome_base[len("resultado_final_"):]
    elif nome_base == "resultado_final":
        nome_base = "dia"
    return PROCESS_OUTPUT_FOLDER / f"calculo_medicao_disjuntor_jundiai_{nome_base}.xlsx"




def gerar_bases_dashboard(final, tipo_periodo, data_inicio, data_fim, logger=None):
    """
    Gera os CSVs do dashboard de forma ACUMULATIVA.

    Regra principal:
    - Cada nova extração entra no notas_dashboard.csv histórico.
    - Se uma ORDEM_DE_SERVICO já existir, ela é atualizada/substituída pela versão mais nova.
    - Os demais CSVs do dashboard são recalculados a partir desse histórico completo.

    Assim, uma extração diária, semanal ou mensal não apaga o que já estava no painel.
    """
    logger = logger or log
    pasta_dashboard = PROCESS_OUTPUT_FOLDER / "dashboard"
    pasta_dashboard.mkdir(parents=True, exist_ok=True)

    caminhos = {
        "notas_dashboard": pasta_dashboard / "notas_dashboard.csv",
        "faturamento_contratos_dashboard": pasta_dashboard / "faturamento_contratos_dashboard.csv",
        "faturamento_dias_dashboard": pasta_dashboard / "faturamento_dias_dashboard.csv",
        "faturamento_carro_estimado_dashboard": pasta_dashboard / "faturamento_carro_estimado_dashboard.csv",
        "faturamento_carro_dias_dashboard": pasta_dashboard / "faturamento_carro_dias_dashboard.csv",
    }

    notas_cols = [
        "PERIODO_TIPO", "PERIODO_INICIO", "PERIODO_FIM", "ORDEM_DE_SERVICO", "GRUPO_NOTA", "RECURSO",
        "FINALIZACAO", "DATA_ENCERRAMENTO", "DATA", "DIA_SEMANA", "SEMANA_INICIO",
        "ELETRICISTA1", "ELETRICISTA2", "QTD_EXECUTORES", "RECUSA",
    ]

    def preparar_notas(df_base, periodo_tipo=None, periodo_inicio=None, periodo_fim=None):
        dfp = df_base.copy()

        for col in ["RECURSO", "RECUSA", "ELETRICISTA1", "ELETRICISTA2", "GRUPO_NOTA", "ORDEM_DE_SERVICO", "FINALIZACAO", "DATA_ENCERRAMENTO"]:
            if col not in dfp.columns:
                dfp[col] = ""
            dfp[col] = dfp[col].fillna("").astype(str).str.strip()

        dfp["DATA_ENCERRAMENTO_DT"] = pd.to_datetime(dfp["DATA_ENCERRAMENTO"], dayfirst=True, errors="coerce")
        dfp["DATA"] = dfp["DATA_ENCERRAMENTO_DT"].dt.strftime("%d/%m/%Y").fillna("")
        dfp["DIA_SEMANA_NUM"] = dfp["DATA_ENCERRAMENTO_DT"].dt.weekday
        dfp["DIA_SEMANA"] = dfp["DIA_SEMANA_NUM"].map({
            0: "Segunda", 1: "Terça", 2: "Quarta", 3: "Quinta",
            4: "Sexta", 5: "Sábado", 6: "Domingo",
        }).fillna("")
        dfp["SEMANA_INICIO"] = (
            dfp["DATA_ENCERRAMENTO_DT"] - pd.to_timedelta(dfp["DIA_SEMANA_NUM"].fillna(0), unit="D")
        ).dt.strftime("%d/%m/%Y").fillna("")

        dfp["QTD_EXECUTORES"] = (
            (dfp["ELETRICISTA1"] != "").astype(int) + (dfp["ELETRICISTA2"] != "").astype(int)
        )

        if periodo_tipo is not None:
            dfp["PERIODO_TIPO"] = periodo_tipo
        elif "PERIODO_TIPO" not in dfp.columns:
            dfp["PERIODO_TIPO"] = ""

        if periodo_inicio is not None:
            dfp["PERIODO_INICIO"] = periodo_inicio.strftime("%d/%m/%Y")
        elif "PERIODO_INICIO" not in dfp.columns:
            dfp["PERIODO_INICIO"] = ""

        if periodo_fim is not None:
            dfp["PERIODO_FIM"] = periodo_fim.strftime("%d/%m/%Y")
        elif "PERIODO_FIM" not in dfp.columns:
            dfp["PERIODO_FIM"] = ""

        for col in notas_cols:
            if col not in dfp.columns:
                dfp[col] = ""

        return dfp

    # 1) Prepara as notas novas da extração atual.
    notas_novas = preparar_notas(final, tipo_periodo, data_inicio, data_fim)

    # 2) Lê o histórico local já existente, se houver.
    caminho_historico = caminhos["notas_dashboard"]
    if caminho_historico.exists():
        try:
            notas_antigas = pd.read_csv(caminho_historico, sep=";", encoding="utf-8-sig", dtype=str)
            notas_antigas = preparar_notas(notas_antigas)
            logger(f"📚 Histórico do dashboard carregado: {len(notas_antigas)} notas")
        except Exception as e:
            logger(f"⚠️ Não consegui ler o histórico antigo do dashboard. Vou seguir só com a extração atual. Erro: {e}")
            notas_antigas = pd.DataFrame(columns=notas_cols)
    else:
        notas_antigas = pd.DataFrame(columns=notas_cols)

    # 3) Junta histórico + extração nova e remove duplicidade.
    #    A chave principal é ORDEM_DE_SERVICO. Se a OS já existir, fica a versão mais recente.
    historico = pd.concat([notas_antigas[notas_cols], notas_novas[notas_cols]], ignore_index=True)
    historico = preparar_notas(historico)

    historico["ORDEM_DE_SERVICO"] = historico["ORDEM_DE_SERVICO"].fillna("").astype(str).str.strip()
    com_os = historico[historico["ORDEM_DE_SERVICO"] != ""].copy()
    sem_os = historico[historico["ORDEM_DE_SERVICO"] == ""].copy()

    antes = len(historico)
    com_os = com_os.drop_duplicates(subset=["ORDEM_DE_SERVICO"], keep="last")
    sem_os = sem_os.drop_duplicates(subset=notas_cols, keep="last")
    df = pd.concat([com_os, sem_os], ignore_index=True)
    df = preparar_notas(df)
    depois = len(df)

    logger(f"➕ Novas notas da extração atual: {len(notas_novas)}")
    logger(f"🧹 Duplicidades removidas do dashboard: {antes - depois}")
    logger(f"📊 Total acumulado no dashboard: {depois} notas")

    def eh_disjuntor_jundiai(recurso):
        recurso_norm = str(recurso).strip().upper()
        return recurso_norm.startswith("JUN55") or recurso_norm.startswith("JUN59") or recurso_norm.startswith("SAL55")

    def eh_disjuntor_santa_cruz(recurso):
        recurso_norm = str(recurso).strip().upper()
        m = re.search(r"(\d+)", recurso_norm)
        if not m:
            return False
        primeiros_numeros = m.group(1)
        return primeiros_numeros.startswith("89") or primeiros_numeros.startswith("20")

    contratos_precos = [
        {"CONTRATO": "Disjuntor Jundiaí", "filtro": eh_disjuntor_jundiai, "CORTE": 13.72, "RELIGUE": 27.43},
        {"CONTRATO": "Disjuntor Santa Cruz", "filtro": eh_disjuntor_santa_cruz, "CORTE": 11.98, "RELIGUE": 23.97, "VERIFICACAO": 23.97},
    ]

    df_pagavel = df[df["RECUSA"] == ""].copy()
    partes_contratos = []

    for contrato in contratos_precos:
        parte = df_pagavel[df_pagavel["RECURSO"].apply(contrato["filtro"])].copy()
        if parte.empty:
            continue
        parte["CONTRATO"] = contrato["CONTRATO"]
        parte["VALOR"] = parte["GRUPO_NOTA"].map({"CORTE": contrato["CORTE"], "RELIGUE": contrato["RELIGUE"], "VERIFICACAO": contrato.get("VERIFICACAO", 0)}).fillna(0)
        partes_contratos.append(parte)

    df_contratos = pd.concat(partes_contratos, ignore_index=True) if partes_contratos else pd.DataFrame()

    if not df_contratos.empty:
        resumo = (
            df_contratos.groupby(["CONTRATO", "GRUPO_NOTA"], dropna=False)
            .agg(QTD_NOTAS=("ORDEM_DE_SERVICO", "count"), FATURAMENTO=("VALOR", "sum"))
            .reset_index()
        )
        dias = (
            df_contratos.groupby(["CONTRATO", "SEMANA_INICIO", "DIA_SEMANA_NUM", "DIA_SEMANA", "DATA"], dropna=False)
            .agg(QTD_NOTAS=("ORDEM_DE_SERVICO", "count"), FATURAMENTO=("VALOR", "sum"))
            .reset_index()
            .sort_values(["CONTRATO", "DIA_SEMANA_NUM", "DATA"])
        )
    else:
        resumo = pd.DataFrame(columns=["CONTRATO", "GRUPO_NOTA", "QTD_NOTAS", "FATURAMENTO"])
        dias = pd.DataFrame(columns=["CONTRATO", "SEMANA_INICIO", "DIA_SEMANA_NUM", "DIA_SEMANA", "DATA", "QTD_NOTAS", "FATURAMENTO"])

    mask_carro = (
        df_pagavel["RECURSO"].str.upper().str.startswith("JUN58")
        & (df_pagavel["ELETRICISTA1"] != "")
        & (df_pagavel["ELETRICISTA2"] != "")
    )
    carro = df_pagavel[mask_carro].copy()

    if not carro.empty:
        carro["CONTRATO"] = "Contrato Carro STC estimado"
        carro["VALOR_MIN_ESTIMADO"] = carro["GRUPO_NOTA"].map({"CORTE": 38.18, "RELIGUE": 36.36}).fillna(0)
        carro["VALOR_MAX_ESTIMADO"] = carro["GRUPO_NOTA"].map({"CORTE": 45.45, "RELIGUE": 50.91}).fillna(0)

        carro_resumo = (
            carro.groupby(["CONTRATO", "GRUPO_NOTA"], dropna=False)
            .agg(
                QTD_NOTAS=("ORDEM_DE_SERVICO", "count"),
                FATURAMENTO_MIN=("VALOR_MIN_ESTIMADO", "sum"),
                FATURAMENTO_MAX=("VALOR_MAX_ESTIMADO", "sum"),
            )
            .reset_index()
        )
        carro_dias = (
            carro.groupby(["CONTRATO", "SEMANA_INICIO", "DIA_SEMANA_NUM", "DIA_SEMANA", "DATA"], dropna=False)
            .agg(
                QTD_NOTAS=("ORDEM_DE_SERVICO", "count"),
                FATURAMENTO_MIN=("VALOR_MIN_ESTIMADO", "sum"),
                FATURAMENTO_MAX=("VALOR_MAX_ESTIMADO", "sum"),
            )
            .reset_index()
            .sort_values(["DIA_SEMANA_NUM", "DATA"])
        )
    else:
        carro_resumo = pd.DataFrame(columns=["CONTRATO", "GRUPO_NOTA", "QTD_NOTAS", "FATURAMENTO_MIN", "FATURAMENTO_MAX"])
        carro_dias = pd.DataFrame(columns=["CONTRATO", "SEMANA_INICIO", "DIA_SEMANA_NUM", "DIA_SEMANA", "DATA", "QTD_NOTAS", "FATURAMENTO_MIN", "FATURAMENTO_MAX"])

    df[notas_cols].to_csv(caminhos["notas_dashboard"], sep=";", index=False, encoding="utf-8-sig")
    resumo.to_csv(caminhos["faturamento_contratos_dashboard"], sep=";", index=False, encoding="utf-8-sig")
    dias.to_csv(caminhos["faturamento_dias_dashboard"], sep=";", index=False, encoding="utf-8-sig")
    carro_resumo.to_csv(caminhos["faturamento_carro_estimado_dashboard"], sep=";", index=False, encoding="utf-8-sig")
    carro_dias.to_csv(caminhos["faturamento_carro_dias_dashboard"], sep=";", index=False, encoding="utf-8-sig")

    logger(f"✅ Bases acumuladas para dashboard atualizadas em: {pasta_dashboard}")
    return {k: str(v) for k, v in caminhos.items()}

def obter_token_github_dashboard():
    """Procura o token do GitHub nas variáveis de ambiente/.env."""
    # Recarrega o .env aqui também, para pegar alterações feitas sem reiniciar tudo.
    try:
        load_dotenv(ENV_PATH, override=True)
    except Exception:
        pass
    for nome in GITHUB_DASHBOARD_TOKEN_ENV_NAMES:
        valor = os.getenv(nome, "").strip()
        if valor:
            return valor
    return ""


def _github_erro_transitorio(erro):
    """Identifica falhas temporárias do GitHub/rede que devem ser retentadas ou ignoradas sem derrubar a extração."""
    texto = str(erro).upper()
    return any(sinal in texto for sinal in [
        "HTTP 408", "HTTP 429", "HTTP 500", "HTTP 502", "HTTP 503", "HTTP 504",
        "BAD GATEWAY", "SERVER ERROR", "TIMED OUT", "TIMEOUT", "TEMPORAR",
        "CONNECTION RESET", "REMOTE END CLOSED", "URLERROR",
    ])


def _github_api_request(url, token=None, method="GET", payload=None, retries=6, timeout=60):
    """Chamada robusta à API do GitHub.

    Corrige o problema observado nas exportações: HTTP 502/Bad Gateway é erro temporário
    do lado do GitHub/rede. Antes, uma única resposta 502 no upload de um CSV derrubava
    toda a execução. Agora a chamada tenta novamente com espera progressiva antes de falhar.
    """
    headers = {
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
        "User-Agent": "extrator-gerente-dashboard",
    }
    if token:
        headers["Authorization"] = f"Bearer {token}"
    data = None
    if payload is not None:
        data = json.dumps(payload).encode("utf-8")
        headers["Content-Type"] = "application/json"

    ultima_mensagem = ""
    for tentativa in range(1, int(retries) + 1):
        req = urllib.request.Request(url, data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                corpo = resp.read().decode("utf-8")
                return json.loads(corpo) if corpo else {}
        except urllib.error.HTTPError as e:
            corpo = e.read().decode("utf-8", errors="replace")
            ultima_mensagem = f"GitHub respondeu erro HTTP {e.code}: {corpo}"

            # 404/401/403 normalmente são configuração/permissão, não adianta insistir.
            if e.code not in (408, 429, 500, 502, 503, 504):
                raise RuntimeError(ultima_mensagem) from e

            if tentativa >= int(retries):
                raise RuntimeError(
                    f"GitHub manteve erro temporário após {retries} tentativa(s): {ultima_mensagem}"
                ) from e

        except urllib.error.URLError as e:
            ultima_mensagem = f"Falha de conexão com o GitHub: {e}"
            if tentativa >= int(retries):
                raise RuntimeError(
                    f"Falha temporária de conexão com o GitHub após {retries} tentativa(s): {e}"
                ) from e

        # Backoff progressivo: 2s, 5s, 10s, 20s, 40s... limitado a 60s.
        espera = min(60, [2, 5, 10, 20, 40, 60][min(tentativa - 1, 5)])
        try:
            print(f"⚠️ GitHub instável ({ultima_mensagem}). Nova tentativa {tentativa + 1}/{retries} em {espera}s...")
        except Exception:
            pass
        time.sleep(espera)

    raise RuntimeError(ultima_mensagem or "Falha desconhecida ao chamar GitHub")

def _normalizar_repo_github(repo):
    """Aceita usuario/repositorio ou URL do GitHub e devolve usuario/repositorio."""
    repo = (repo or "").strip().strip("/")
    repo = repo.replace("https://github.com/", "").replace("http://github.com/", "")
    repo = repo.replace("github.com/", "")
    if repo.endswith(".git"):
        repo = repo[:-4]
    return repo.strip("/")


def _repo_publico_existe(repo):
    """Testa sem token se o repositório público existe. Serve para diferenciar repo errado de token sem acesso."""
    try:
        _github_api_request(f"https://api.github.com/repos/{repo}", token=None, method="GET")
        return True
    except Exception:
        return False


def enviar_dashboard_para_github(caminhos_dashboard, logger=None):
    """Envia/atualiza os CSVs do dashboard no repositório do GitHub.

    Versão resiliente: erro temporário 502/503/504/429 não derruba a extração inteira.
    A chamada tenta novamente e, se um arquivo continuar falhando, registra aviso e segue
    com os demais arquivos. Os CSVs locais continuam salvos normalmente em output/dashboard.
    """
    logger = logger or log

    # Recarrega as opções do .env na hora do envio.
    # Isso permite corrigir GITHUB_DASHBOARD_REPO no .env sem mexer no código.
    try:
        load_dotenv(ENV_PATH, override=True)
    except Exception:
        pass

    token = obter_token_github_dashboard()
    if not token:
        logger(
            "ℹ️ Upload automático para GitHub pulado: token não configurado. "
            "Configure GITHUB_DASHBOARD_TOKEN no .env para atualizar o painel sozinho."
        )
        return {"enviado": False, "motivo": "token_nao_configurado", "arquivos": []}

    repo = _normalizar_repo_github(os.getenv("GITHUB_DASHBOARD_REPO", GITHUB_DASHBOARD_REPO))
    branch = (os.getenv("GITHUB_DASHBOARD_BRANCH", GITHUB_DASHBOARD_BRANCH) or "main").strip() or "main"
    remote_dir = (os.getenv("GITHUB_DASHBOARD_REMOTE_DIR", GITHUB_DASHBOARD_REMOTE_DIR) or "dashboard").strip().strip("/") or "dashboard"

    if "/" not in repo:
        raise RuntimeError(
            "Configuração inválida: GITHUB_DASHBOARD_REPO deve estar no formato usuario/repositorio. "
            "Exemplo: irensegabriel/painel-faturamento"
        )

    api_base = f"https://api.github.com/repos/{repo}"
    logger(f"☁️ Upload GitHub iniciado: repo={repo} | branch={branch} | pasta={remote_dir}")

    repo_existe_sem_token = _repo_publico_existe(repo)

    try:
        _github_api_request(api_base, token, method="GET")
    except RuntimeError as e:
        if "HTTP 404" in str(e):
            if repo_existe_sem_token:
                raise RuntimeError(
                    "O repositório existe publicamente, mas ESSE TOKEN não tem acesso a ele. "
                    "Crie um token fine-grained com acesso ao repositório e Contents = Read and write. "
                    f"Repo usado pelo extrator: {repo}"
                ) from e
            raise RuntimeError(
                "GitHub não encontrou o repositório. Confira GITHUB_DASHBOARD_REPO no .env. "
                f"Repo configurado: {repo}"
            ) from e
        if _github_erro_transitorio(e):
            logger(f"⚠️ GitHub instável ao validar repositório. Upload pulado nesta execução: {e}")
            return {"enviado": False, "motivo": "github_instavel_validacao_repo", "arquivos": []}
        raise

    try:
        _github_api_request(f"{api_base}/branches/{branch}", token, method="GET")
    except RuntimeError as e:
        if "HTTP 404" in str(e):
            raise RuntimeError(
                f"GitHub não encontrou a branch '{branch}'. Confira se a branch é main. "
                "Se for outro nome, ajuste GITHUB_DASHBOARD_BRANCH no .env."
            ) from e
        if _github_erro_transitorio(e):
            logger(f"⚠️ GitHub instável ao validar branch. Upload pulado nesta execução: {e}")
            return {"enviado": False, "motivo": "github_instavel_validacao_branch", "arquivos": []}
        raise

    enviados = []
    falhas_temporarias = []

    for nome_base, caminho in caminhos_dashboard.items():
        caminho = Path(caminho)
        if not caminho.exists():
            logger(f"⚠️ Dashboard GitHub: arquivo não encontrado, ignorado: {caminho}")
            continue

        remote_path = f"{remote_dir}/{caminho.name}"
        url = f"{api_base}/contents/{remote_path}"
        sha_atual = None

        try:
            existente = _github_api_request(f"{url}?ref={branch}", token, method="GET")
            sha_atual = existente.get("sha")
        except RuntimeError as e:
            if "HTTP 404" not in str(e):
                if _github_erro_transitorio(e):
                    falhas_temporarias.append(remote_path)
                    logger(f"⚠️ GitHub instável ao consultar {remote_path}. Vou manter o arquivo local e seguir: {e}")
                    continue
                raise

        conteudo_b64 = base64.b64encode(caminho.read_bytes()).decode("ascii")
        payload = {
            "message": f"Atualiza dashboard: {caminho.name}",
            "content": conteudo_b64,
            "branch": branch,
        }
        if sha_atual:
            payload["sha"] = sha_atual

        try:
            _github_api_request(url, token, method="PUT", payload=payload)
        except RuntimeError as e:
            if "HTTP 404" in str(e):
                raise RuntimeError(
                    "GitHub recusou o envio com 404. Normalmente isso é token sem acesso ao repo "
                    "ou permissão Contents sem Read and write. Recrie o token fine-grained selecionando "
                    f"o repositório {repo} e Contents = Read and write."
                ) from e
            if _github_erro_transitorio(e):
                falhas_temporarias.append(remote_path)
                logger(f"⚠️ GitHub instável ao enviar {remote_path}. Arquivo local preservado; próxima execução tenta novamente. Erro: {e}")
                continue
            raise

        enviados.append(remote_path)
        logger(f"☁️ Dashboard GitHub atualizado: {remote_path}")

    if enviados:
        if falhas_temporarias:
            logger(
                "⚠️ Upload automático para GitHub concluído parcialmente: "
                f"{len(enviados)} enviado(s), {len(falhas_temporarias)} com falha temporária. "
                "A próxima execução automática tentará reenviar."
            )
            return {"enviado": True, "motivo": "parcial_github_instavel", "arquivos": enviados, "falhas": falhas_temporarias}
        logger(f"✅ Upload automático para GitHub concluído: {len(enviados)} arquivo(s).")
        return {"enviado": True, "motivo": "ok", "arquivos": enviados}

    if falhas_temporarias:
        logger(
            "⚠️ Nenhum arquivo foi enviado porque o GitHub respondeu erro temporário. "
            "Os CSVs locais foram gerados e a próxima execução tentará novamente."
        )
        return {"enviado": False, "motivo": "github_instavel", "arquivos": [], "falhas": falhas_temporarias}

    logger("ℹ️ Nenhum arquivo de dashboard foi enviado ao GitHub.")
    return {"enviado": False, "motivo": "sem_arquivos", "arquivos": []}


# ==============================
# PÓS-EXTRATOR G.Z.U.S. / SQLITE DASHBOARD
# ==============================
def _copiar_csvs_para_painel_faturamento(caminhos_dashboard, logger=None):
    """Copia os CSVs gerados em output/dashboard para o repositório local do painel.

    Por que existe:
    - O extrator salva os CSVs em output/dashboard.
    - O banco_gzus.py lê os CSVs em painel-faturamento/dashboard.
    - Sem esta cópia, o SQLite pode ser gerado com dados velhos.
    """
    logger = logger or log

    projeto_painel = Path(os.getenv("GZUS_PROJETO_PAINEL_DIR", str(BASE_DIR / "painel-faturamento")))
    destino_dashboard = projeto_painel / "dashboard"
    destino_dashboard.mkdir(parents=True, exist_ok=True)

    copiados = {}
    for nome_base, caminho_origem in (caminhos_dashboard or {}).items():
        origem = Path(caminho_origem)
        if not origem.exists():
            logger(f"⚠️ Pós-extrator: CSV não encontrado para copiar: {origem}")
            continue
        destino = destino_dashboard / origem.name
        shutil.copy2(origem, destino)
        copiados[nome_base] = str(destino)
        logger(f"📌 Pós-extrator: CSV copiado para painel local: {destino.name}")

    return projeto_painel, destino_dashboard, copiados


def executar_pos_extrator_gzus(caminhos_dashboard, logger=None):
    """Atualiza automaticamente o SQLite leve do dashboard e sobe ao GitHub.

    Fluxo:
    1) copia CSVs novos de output/dashboard para painel-faturamento/dashboard;
    2) roda `python banco_gzus.py importar` dentro de painel-faturamento;
    3) envia dashboard/gzus_dashboard.db para o GitHub usando o mesmo token já configurado.
    """
    logger = logger or log

    habilitado = str(os.getenv("GZUS_POS_EXTRATOR_AUTO", "true") or "true").strip().lower()
    if habilitado in ["0", "false", "nao", "não", "no", "off"]:
        logger("ℹ️ Pós-extrator G.Z.U.S. desligado por GZUS_POS_EXTRATOR_AUTO=false.")
        return {"ok": False, "motivo": "desligado"}

    try:
        projeto_painel, destino_dashboard, copiados = _copiar_csvs_para_painel_faturamento(caminhos_dashboard, logger=logger)

        banco_script = projeto_painel / "banco_gzus.py"
        if not banco_script.exists():
            logger(f"⚠️ Pós-extrator: banco_gzus.py não encontrado em {banco_script}")
            return {"ok": False, "motivo": "banco_gzus_nao_encontrado"}

        logger("🧠 Pós-extrator: recriando gzus_dashboard.db...")
        resultado = subprocess.run(
            [sys.executable, str(banco_script.name), "importar"],
            cwd=str(projeto_painel),
            capture_output=True,
            text=True,
            timeout=int(os.getenv("GZUS_POS_EXTRATOR_TIMEOUT", "300") or "300"),
        )

        if resultado.stdout:
            for linha in resultado.stdout.strip().splitlines():
                logger(f"   {linha}")
        if resultado.stderr:
            for linha in resultado.stderr.strip().splitlines():
                logger(f"   ⚠️ {linha}")

        if resultado.returncode != 0:
            logger(f"⚠️ Pós-extrator: banco_gzus.py importar falhou com código {resultado.returncode}.")
            return {"ok": False, "motivo": "erro_importar_banco", "codigo": resultado.returncode}

        db_leve = destino_dashboard / "gzus_dashboard.db"
        if not db_leve.exists():
            logger(f"⚠️ Pós-extrator: gzus_dashboard.db não foi criado em {db_leve}")
            return {"ok": False, "motivo": "db_nao_criado"}

        tamanho_kb = db_leve.stat().st_size / 1024
        logger(f"✅ Pós-extrator: banco leve atualizado ({tamanho_kb:,.1f} KB).")

        # Envia só o banco leve para o GitHub. Os CSVs já foram enviados pelo fluxo normal.
        upload_db = enviar_dashboard_para_github({"gzus_dashboard": str(db_leve)}, logger=logger)
        if upload_db.get("enviado"):
            logger("✅ Pós-extrator: gzus_dashboard.db enviado ao GitHub automaticamente.")
        else:
            logger(f"⚠️ Pós-extrator: gzus_dashboard.db não enviado ao GitHub. Motivo: {upload_db.get('motivo')}")

        return {
            "ok": True,
            "copiados": copiados,
            "db": str(db_leve),
            "tamanho_kb": round(tamanho_kb, 1),
            "upload_db": upload_db,
        }

    except Exception as e:
        logger(f"⚠️ Pós-extrator G.Z.U.S. falhou, mas a extração principal foi preservada: {e}")
        return {"ok": False, "motivo": "excecao", "erro": str(e)}

def gerar_planilha_medicao_disjuntor_jundiai(df_final, caminho_saida, periodo_inicio, periodo_fim, logger=None):
    logger = logger or log

    df_base = df_final.copy()
    df_base["RECURSO"] = df_base["RECURSO"].fillna("").astype(str).str.strip()
    df_base["RECUSA"] = df_base["RECUSA"].fillna("").astype(str).str.strip()
    df_base = df_base[df_base["RECUSA"] == ""].copy()

    def eh_disjuntor_jundiai(recurso):
        recurso_norm = str(recurso).strip().upper()
        return (
            recurso_norm.startswith("JUN55")
            or recurso_norm.startswith("JUN59")
            or recurso_norm.startswith("SAL55")
        )

    def eh_disjuntor_santa_cruz(recurso):
        recurso_norm = str(recurso).strip().upper()
        m = re.search(r"(\d+)", recurso_norm)
        if not m:
            return False
        primeiros_numeros = m.group(1)
        return primeiros_numeros.startswith("89") or primeiros_numeros.startswith("20")

    contratos = [
        {
            "nome": "Disjuntor Jundiaí",
            "prefixo_abas": "Jundiaí",
            "filtro": eh_disjuntor_jundiai,
            "preco_corte": 13.72,
            "preco_religue": 27.43,
        },
        {
            "nome": "Disjuntor Santa Cruz",
            "prefixo_abas": "Santa Cruz",
            "filtro": eh_disjuntor_santa_cruz,
            "preco_corte": 11.98,
            "preco_religue": 23.97,
            "preco_verificacao": 23.97,
        },
    ]

    dark_fill = PatternFill("solid", fgColor="1F1B3A")
    header_fill = PatternFill("solid", fgColor="E9D5FF")
    total_fill = PatternFill("solid", fgColor="DDD6FE")
    thin_top = Side(style="thin", color="6D28D9")
    currency_fmt = 'R$ #,##0.00;[Red](R$ #,##0.00);-'
    integer_fmt = '#,##0;[Red](#,##0);-'

    wb = Workbook()
    wb.remove(wb.active)

    registros_faturamento_dia_semana = []
    registros_faturamento_carro_dia_semana = []

    contrato_carro_estimado_gerado = False

    contratos_gerados = []
    for contrato in contratos:
        df_med = df_base[df_base["RECURSO"].apply(contrato["filtro"])].copy()
        if df_med.empty:
            logger(f"ℹ️ Nenhuma nota pagável encontrada para o contrato {contrato['nome']}.")
            continue

        df_med["CONTRATO"] = contrato["nome"]
        df_med["DATA_ENCERRAMENTO_DT"] = pd.to_datetime(
            df_med["DATA_ENCERRAMENTO"], dayfirst=True, errors="coerce"
        )
        df_med["DATA_FATURAMENTO"] = df_med["DATA_ENCERRAMENTO_DT"].dt.date
        df_med["VALOR_FATURAMENTO"] = df_med["GRUPO_NOTA"].map({
            "CORTE": contrato["preco_corte"],
            "RELIGUE": contrato["preco_religue"],
            "VERIFICACAO": contrato.get("preco_verificacao", 0),
        }).fillna(0)
        if not df_med.empty:
            registros_faturamento_dia_semana.append(
                df_med[["CONTRATO", "DATA_FATURAMENTO", "VALOR_FATURAMENTO"]].copy()
            )

        resumo = (
            df_med.groupby("RECURSO", dropna=False)["GRUPO_NOTA"]
            .value_counts()
            .unstack(fill_value=0)
            .reset_index()
        )

        if "CORTE" not in resumo.columns:
            resumo["CORTE"] = 0
        if "RELIGUE" not in resumo.columns:
            resumo["RELIGUE"] = 0

        resumo = resumo.rename(columns={"CORTE": "QTD_CORTE", "RELIGUE": "QTD_RELIGUE"})
        resumo["CONTRATO"] = contrato["nome"]

        nome_curto = contrato["prefixo_abas"]

        ws_resumo = wb.create_sheet(f"Resumo {nome_curto}")
        ws_resumo["A1"] = f"Cálculo medição {contrato['nome']}"
        ws_resumo["A1"].font = Font(color="FFFFFF", bold=True, size=13)
        ws_resumo["A1"].fill = dark_fill
        ws_resumo.merge_cells("A1:F1")

        ws_resumo["A3"] = "Período"
        ws_resumo["B3"] = f"{periodo_inicio} até {periodo_fim}"
        ws_resumo["A4"] = "Contrato"
        ws_resumo["B4"] = contrato["nome"]
        ws_resumo["A5"] = "Preço corte"
        ws_resumo["B5"] = contrato["preco_corte"]
        ws_resumo["A6"] = "Preço religue"
        ws_resumo["B6"] = contrato["preco_religue"]
        ws_resumo["B5"].number_format = currency_fmt
        ws_resumo["B6"].number_format = currency_fmt

        resumo_headers = ["Contrato", "Qtd corte", "Qtd religue", "Preço corte", "Preço religue", "Faturamento total"]
        for col_idx, header in enumerate(resumo_headers, start=1):
            cell = ws_resumo.cell(row=8, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = Font(bold=True)

        ws_resumo["A9"] = contrato["nome"]
        ws_resumo["D9"] = "=B5"
        ws_resumo["E9"] = "=B6"

        ws_equipe = wb.create_sheet(f"Equipe {nome_curto}")
        equipe_headers = [
            "Equipe", "Contrato", "Qtd corte", "Qtd religue",
            "Preço corte", "Preço religue", "Faturamento corte",
            "Faturamento religue", "Faturamento total"
        ]
        for col_idx, header in enumerate(equipe_headers, start=1):
            cell = ws_equipe.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for row_idx, (_, row) in enumerate(resumo.sort_values("RECURSO").iterrows(), start=2):
            ws_equipe.cell(row=row_idx, column=1, value=row["RECURSO"])
            ws_equipe.cell(row=row_idx, column=2, value=row["CONTRATO"])
            ws_equipe.cell(row=row_idx, column=3, value=int(row["QTD_CORTE"]))
            ws_equipe.cell(row=row_idx, column=4, value=int(row["QTD_RELIGUE"]))
            ws_equipe.cell(row=row_idx, column=5, value=contrato["preco_corte"])
            ws_equipe.cell(row=row_idx, column=6, value=contrato["preco_religue"])
            ws_equipe.cell(row=row_idx, column=7, value=f"=C{row_idx}*E{row_idx}")
            ws_equipe.cell(row=row_idx, column=8, value=f"=D{row_idx}*F{row_idx}")
            ws_equipe.cell(row=row_idx, column=9, value=f"=G{row_idx}+H{row_idx}")

        total_row = ws_equipe.max_row + 1
        ws_equipe.cell(row=total_row, column=1, value="TOTAL")
        ws_equipe.cell(row=total_row, column=3, value=f"=SUM(C2:C{total_row-1})")
        ws_equipe.cell(row=total_row, column=4, value=f"=SUM(D2:D{total_row-1})")
        ws_equipe.cell(row=total_row, column=7, value=f"=SUM(G2:G{total_row-1})")
        ws_equipe.cell(row=total_row, column=8, value=f"=SUM(H2:H{total_row-1})")
        ws_equipe.cell(row=total_row, column=9, value=f"=SUM(I2:I{total_row-1})")
        for col in range(1, 10):
            ws_equipe.cell(row=total_row, column=col).fill = total_fill
            ws_equipe.cell(row=total_row, column=col).font = Font(bold=True)
            ws_equipe.cell(row=total_row, column=col).border = Border(top=thin_top)

        ws_resumo["B9"] = f"='Equipe {nome_curto}'!C{total_row}"
        ws_resumo["C9"] = f"='Equipe {nome_curto}'!D{total_row}"
        ws_resumo["F9"] = f"='Equipe {nome_curto}'!I{total_row}"

        for cell_ref in ["B9", "C9"]:
            ws_resumo[cell_ref].number_format = integer_fmt
        for cell_ref in ["D9", "E9", "F9"]:
            ws_resumo[cell_ref].number_format = currency_fmt

        for row in ws_equipe.iter_rows(min_row=2, max_row=total_row, min_col=3, max_col=4):
            for cell in row:
                cell.number_format = integer_fmt
        for row in ws_equipe.iter_rows(min_row=2, max_row=total_row, min_col=5, max_col=9):
            for cell in row:
                cell.number_format = currency_fmt

        ws_notas = wb.create_sheet(f"Notas {nome_curto}")
        notas_headers = ["OS", "Grupo", "Equipe", "Finalização", "Data encerramento", "Recusa", "Contrato"]
        for col_idx, header in enumerate(notas_headers, start=1):
            cell = ws_notas.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = Font(bold=True)

        notas = df_med[["ORDEM_DE_SERVICO", "GRUPO_NOTA", "RECURSO", "FINALIZACAO", "DATA_ENCERRAMENTO", "RECUSA", "CONTRATO"]].copy()
        for row_idx, row in enumerate(notas.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws_notas.cell(row=row_idx, column=col_idx, value=value)

        for ws in [ws_resumo, ws_equipe, ws_notas]:
            ws.freeze_panes = "A2" if not ws.title.startswith("Resumo ") else "A8"
            widths = {1: 20, 2: 22, 3: 14, 4: 14, 5: 18, 6: 18, 7: 20, 8: 20, 9: 20}
            for col_idx, width in widths.items():
                if col_idx <= ws.max_column:
                    ws.column_dimensions[get_column_letter(col_idx)].width = width

        contratos_gerados.append(contrato["nome"])

    # Contrato do carro/STC: estimativa baseada nas notas com dois executores.
    # Os valores desse contrato variam, por isso a planilha mostra faixa mínima e máxima.
    df_carro = df_base.copy()
    for col in ["RECURSO", "ELETRICISTA1", "ELETRICISTA2", "GRUPO_NOTA"]:
        if col not in df_carro.columns:
            df_carro[col] = ""
        df_carro[col] = df_carro[col].fillna("").astype(str).str.strip()

    mask_carro_stc = df_carro["RECURSO"].str.upper().str.startswith("JUN58")
    mask_dois_executores = (df_carro["ELETRICISTA1"] != "") & (df_carro["ELETRICISTA2"] != "")
    df_carro = df_carro[mask_carro_stc & mask_dois_executores].copy()

    if df_carro.empty:
        logger("ℹ️ Nenhuma nota pagável encontrada para o contrato do carro/STC estimado.")
    else:
        contrato_carro_estimado_gerado = True
        contrato_carro_nome = "Contrato Carro STC (estimado)"
        precos_carro = {
            "CORTE": {"min": 38.18, "max": 45.45},
            "RELIGUE": {"min": 36.36, "max": 50.91},
        }

        df_carro["CONTRATO"] = contrato_carro_nome
        df_carro["DATA_ENCERRAMENTO_DT"] = pd.to_datetime(
            df_carro["DATA_ENCERRAMENTO"], dayfirst=True, errors="coerce"
        )
        df_carro["DATA_FATURAMENTO"] = df_carro["DATA_ENCERRAMENTO_DT"].dt.date
        df_carro["VALOR_MIN_ESTIMADO"] = df_carro["GRUPO_NOTA"].map({
            "CORTE": precos_carro["CORTE"]["min"],
            "RELIGUE": precos_carro["RELIGUE"]["min"],
        }).fillna(0)
        df_carro["VALOR_MAX_ESTIMADO"] = df_carro["GRUPO_NOTA"].map({
            "CORTE": precos_carro["CORTE"]["max"],
            "RELIGUE": precos_carro["RELIGUE"]["max"],
        }).fillna(0)
        registros_faturamento_carro_dia_semana.append(
            df_carro[["CONTRATO", "DATA_FATURAMENTO", "VALOR_MIN_ESTIMADO", "VALOR_MAX_ESTIMADO"]].copy()
        )

        resumo_carro = (
            df_carro.groupby("RECURSO", dropna=False)["GRUPO_NOTA"]
            .value_counts()
            .unstack(fill_value=0)
            .reset_index()
        )
        if "CORTE" not in resumo_carro.columns:
            resumo_carro["CORTE"] = 0
        if "RELIGUE" not in resumo_carro.columns:
            resumo_carro["RELIGUE"] = 0
        resumo_carro = resumo_carro.rename(columns={"CORTE": "QTD_CORTE", "RELIGUE": "QTD_RELIGUE"})

        ws_resumo_carro = wb.create_sheet("Resumo Carro Est")
        ws_resumo_carro["A1"] = "Faturamento estimado - Contrato Carro STC"
        ws_resumo_carro["A1"].font = Font(color="FFFFFF", bold=True, size=13)
        ws_resumo_carro["A1"].fill = dark_fill
        ws_resumo_carro.merge_cells("A1:I1")
        ws_resumo_carro["A3"] = "Período"
        ws_resumo_carro["B3"] = f"{periodo_inicio} até {periodo_fim}"
        ws_resumo_carro["A4"] = "Contrato"
        ws_resumo_carro["B4"] = contrato_carro_nome
        ws_resumo_carro["A5"] = "Critério"
        ws_resumo_carro["B5"] = "Notas STC/JUN58 pagáveis com dois executores."
        ws_resumo_carro.merge_cells("B5:I5")
        ws_resumo_carro["A6"] = "Observação"
        ws_resumo_carro["B6"] = "Valores estimados em faixa mínima/máxima, não valor fechado."
        ws_resumo_carro.merge_cells("B6:I6")

        resumo_carro_headers = [
            "Contrato", "Qtd corte", "Qtd religue", "Corte mín", "Corte máx",
            "Religa mín", "Religa máx", "Total estimado mín", "Total estimado máx"
        ]
        for col_idx, header in enumerate(resumo_carro_headers, start=1):
            cell = ws_resumo_carro.cell(row=8, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = Font(bold=True)

        ws_resumo_carro["A9"] = contrato_carro_nome
        ws_resumo_carro["D9"] = precos_carro["CORTE"]["min"]
        ws_resumo_carro["E9"] = precos_carro["CORTE"]["max"]
        ws_resumo_carro["F9"] = precos_carro["RELIGUE"]["min"]
        ws_resumo_carro["G9"] = precos_carro["RELIGUE"]["max"]

        ws_equipe_carro = wb.create_sheet("Equipe Carro Est")
        equipe_carro_headers = [
            "Equipe", "Contrato", "Qtd corte", "Qtd religue",
            "Corte mín", "Corte máx", "Religa mín", "Religa máx",
            "Total mín", "Total máx"
        ]
        for col_idx, header in enumerate(equipe_carro_headers, start=1):
            cell = ws_equipe_carro.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for row_idx, (_, row) in enumerate(resumo_carro.sort_values("RECURSO").iterrows(), start=2):
            ws_equipe_carro.cell(row=row_idx, column=1, value=row["RECURSO"])
            ws_equipe_carro.cell(row=row_idx, column=2, value=contrato_carro_nome)
            ws_equipe_carro.cell(row=row_idx, column=3, value=int(row["QTD_CORTE"]))
            ws_equipe_carro.cell(row=row_idx, column=4, value=int(row["QTD_RELIGUE"]))
            ws_equipe_carro.cell(row=row_idx, column=5, value=precos_carro["CORTE"]["min"])
            ws_equipe_carro.cell(row=row_idx, column=6, value=precos_carro["CORTE"]["max"])
            ws_equipe_carro.cell(row=row_idx, column=7, value=precos_carro["RELIGUE"]["min"])
            ws_equipe_carro.cell(row=row_idx, column=8, value=precos_carro["RELIGUE"]["max"])
            ws_equipe_carro.cell(row=row_idx, column=9, value=f"=C{row_idx}*E{row_idx}+D{row_idx}*G{row_idx}")
            ws_equipe_carro.cell(row=row_idx, column=10, value=f"=C{row_idx}*F{row_idx}+D{row_idx}*H{row_idx}")

        total_row_carro = ws_equipe_carro.max_row + 1
        ws_equipe_carro.cell(row=total_row_carro, column=1, value="TOTAL")
        ws_equipe_carro.cell(row=total_row_carro, column=3, value=f"=SUM(C2:C{total_row_carro-1})")
        ws_equipe_carro.cell(row=total_row_carro, column=4, value=f"=SUM(D2:D{total_row_carro-1})")
        ws_equipe_carro.cell(row=total_row_carro, column=9, value=f"=SUM(I2:I{total_row_carro-1})")
        ws_equipe_carro.cell(row=total_row_carro, column=10, value=f"=SUM(J2:J{total_row_carro-1})")
        for col in range(1, 11):
            ws_equipe_carro.cell(row=total_row_carro, column=col).fill = total_fill
            ws_equipe_carro.cell(row=total_row_carro, column=col).font = Font(bold=True)
            ws_equipe_carro.cell(row=total_row_carro, column=col).border = Border(top=thin_top)

        ws_resumo_carro["B9"] = f"='Equipe Carro Est'!C{total_row_carro}"
        ws_resumo_carro["C9"] = f"='Equipe Carro Est'!D{total_row_carro}"
        ws_resumo_carro["H9"] = f"='Equipe Carro Est'!I{total_row_carro}"
        ws_resumo_carro["I9"] = f"='Equipe Carro Est'!J{total_row_carro}"

        for cell_ref in ["B9", "C9"]:
            ws_resumo_carro[cell_ref].number_format = integer_fmt
        for cell_ref in ["D9", "E9", "F9", "G9", "H9", "I9"]:
            ws_resumo_carro[cell_ref].number_format = currency_fmt

        for row in ws_equipe_carro.iter_rows(min_row=2, max_row=total_row_carro, min_col=3, max_col=4):
            for cell in row:
                cell.number_format = integer_fmt
        for row in ws_equipe_carro.iter_rows(min_row=2, max_row=total_row_carro, min_col=5, max_col=10):
            for cell in row:
                cell.number_format = currency_fmt

        ws_notas_carro = wb.create_sheet("Notas Carro Est")
        notas_carro_headers = [
            "OS", "Grupo", "Equipe", "Executor 1", "Executor 2", "Finalização",
            "Data encerramento", "Recusa", "Valor mín estimado", "Valor máx estimado", "Contrato"
        ]
        for col_idx, header in enumerate(notas_carro_headers, start=1):
            cell = ws_notas_carro.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = Font(bold=True)

        notas_carro = df_carro[[
            "ORDEM_DE_SERVICO", "GRUPO_NOTA", "RECURSO", "ELETRICISTA1", "ELETRICISTA2",
            "FINALIZACAO", "DATA_ENCERRAMENTO", "RECUSA", "VALOR_MIN_ESTIMADO",
            "VALOR_MAX_ESTIMADO", "CONTRATO"
        ]].copy()
        for row_idx, row in enumerate(notas_carro.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws_notas_carro.cell(row=row_idx, column=col_idx, value=value)

        for ws in [ws_resumo_carro, ws_equipe_carro, ws_notas_carro]:
            ws.freeze_panes = "A2" if not ws.title.startswith("Resumo ") else "A8"
            for col_idx in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 18

        for row in ws_notas_carro.iter_rows(min_row=2, max_row=ws_notas_carro.max_row, min_col=9, max_col=10):
            for cell in row:
                cell.number_format = currency_fmt

    if registros_faturamento_dia_semana:
        df_dias = pd.concat(registros_faturamento_dia_semana, ignore_index=True)
        df_dias = df_dias[df_dias["DATA_FATURAMENTO"].notna()].copy()

        if not df_dias.empty:
            df_dias["DATA_FATURAMENTO"] = pd.to_datetime(df_dias["DATA_FATURAMENTO"])
            df_dias["INICIO_SEMANA"] = (
                df_dias["DATA_FATURAMENTO"]
                - pd.to_timedelta(df_dias["DATA_FATURAMENTO"].dt.weekday, unit="D")
            )
            df_dias["DIA_SEMANA_NUM"] = df_dias["DATA_FATURAMENTO"].dt.weekday

            dias_headers = [
                "Semana iniciada em", "Segunda", "Terça", "Quarta",
                "Quinta", "Sexta", "Sábado", "Domingo", "Total semana"
            ]

            def criar_aba_faturamento_dia_semana(df_contrato, contrato_nome, nome_aba):
                resumo_dias = (
                    df_contrato.groupby(["INICIO_SEMANA", "DIA_SEMANA_NUM", "DATA_FATURAMENTO"], dropna=False)["VALOR_FATURAMENTO"]
                    .sum()
                    .reset_index()
                    .sort_values(["INICIO_SEMANA", "DIA_SEMANA_NUM", "DATA_FATURAMENTO"])
                )

                ws_dias = wb.create_sheet(nome_aba)
                ws_dias["A1"] = f"Faturamento por dia da semana - {contrato_nome}"
                ws_dias["A1"].font = Font(color="FFFFFF", bold=True, size=13)
                ws_dias["A1"].fill = dark_fill
                ws_dias.merge_cells("A1:I1")
                ws_dias["A3"] = "Período"
                ws_dias["B3"] = f"{periodo_inicio} até {periodo_fim}"
                ws_dias["A4"] = "Contrato"
                ws_dias["B4"] = contrato_nome
                ws_dias["A5"] = "Observação"
                ws_dias["B5"] = "Somente notas pagáveis deste contrato de medição, sem recusa."
                ws_dias.merge_cells("B5:I5")

                for col_idx, header in enumerate(dias_headers, start=1):
                    cell = ws_dias.cell(row=7, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")

                semanas = list(resumo_dias["INICIO_SEMANA"].drop_duplicates())
                for row_idx, inicio_semana in enumerate(semanas, start=8):
                    ws_dias.cell(row=row_idx, column=1, value=inicio_semana.strftime("%d/%m/%Y"))
                    for dia_num in range(7):
                        dados_dia = resumo_dias[
                            (resumo_dias["INICIO_SEMANA"] == inicio_semana)
                            & (resumo_dias["DIA_SEMANA_NUM"] == dia_num)
                        ]
                        valor = 0 if dados_dia.empty else float(dados_dia["VALOR_FATURAMENTO"].sum())
                        ws_dias.cell(row=row_idx, column=dia_num + 2, value=valor)
                    ws_dias.cell(row=row_idx, column=9, value=f"=SUM(B{row_idx}:H{row_idx})")

                total_row_dias = ws_dias.max_row + 1
                ws_dias.cell(row=total_row_dias, column=1, value="TOTAL")
                for col_idx in range(2, 10):
                    letra = get_column_letter(col_idx)
                    ws_dias.cell(row=total_row_dias, column=col_idx, value=f"=SUM({letra}8:{letra}{total_row_dias-1})")

                for col in range(1, 10):
                    ws_dias.cell(row=total_row_dias, column=col).fill = total_fill
                    ws_dias.cell(row=total_row_dias, column=col).font = Font(bold=True)
                    ws_dias.cell(row=total_row_dias, column=col).border = Border(top=thin_top)

                for row in ws_dias.iter_rows(min_row=8, max_row=total_row_dias, min_col=2, max_col=9):
                    for cell in row:
                        cell.number_format = currency_fmt

                ws_dias.freeze_panes = "A8"
                ws_dias.column_dimensions["A"].width = 20
                for col_idx in range(2, 10):
                    ws_dias.column_dimensions[get_column_letter(col_idx)].width = 16

            for contrato in contratos:
                df_contrato = df_dias[df_dias["CONTRATO"] == contrato["nome"]].copy()
                if df_contrato.empty:
                    continue
                criar_aba_faturamento_dia_semana(
                    df_contrato,
                    contrato["nome"],
                    f"Dias {contrato['prefixo_abas']}"
                )


    if registros_faturamento_carro_dia_semana:
        df_carro_dias = pd.concat(registros_faturamento_carro_dia_semana, ignore_index=True)
        df_carro_dias = df_carro_dias[df_carro_dias["DATA_FATURAMENTO"].notna()].copy()

        if not df_carro_dias.empty:
            df_carro_dias["DATA_FATURAMENTO"] = pd.to_datetime(df_carro_dias["DATA_FATURAMENTO"])
            df_carro_dias["INICIO_SEMANA"] = (
                df_carro_dias["DATA_FATURAMENTO"]
                - pd.to_timedelta(df_carro_dias["DATA_FATURAMENTO"].dt.weekday, unit="D")
            )
            df_carro_dias["DIA_SEMANA_NUM"] = df_carro_dias["DATA_FATURAMENTO"].dt.weekday

            ws_carro_dias = wb.create_sheet("Dias Carro Est")
            ws_carro_dias["A1"] = "Faturamento estimado por dia da semana - Contrato Carro STC"
            ws_carro_dias["A1"].font = Font(color="FFFFFF", bold=True, size=13)
            ws_carro_dias["A1"].fill = dark_fill
            ws_carro_dias.merge_cells("A1:Q1")
            ws_carro_dias["A3"] = "Período"
            ws_carro_dias["B3"] = f"{periodo_inicio} até {periodo_fim}"
            ws_carro_dias["A4"] = "Observação"
            ws_carro_dias["B4"] = "Cada dia tem valor mínimo e máximo estimado."
            ws_carro_dias.merge_cells("B4:Q4")

            headers_carro_dias = ["Semana iniciada em"]
            for dia in ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"]:
                headers_carro_dias.extend([f"{dia} mín", f"{dia} máx"])
            headers_carro_dias.extend(["Total mín semana", "Total máx semana"])

            for col_idx, header in enumerate(headers_carro_dias, start=1):
                cell = ws_carro_dias.cell(row=7, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            resumo_carro_dias = (
                df_carro_dias.groupby(["INICIO_SEMANA", "DIA_SEMANA_NUM"], dropna=False)
                .agg(VALOR_MIN_ESTIMADO=("VALOR_MIN_ESTIMADO", "sum"), VALOR_MAX_ESTIMADO=("VALOR_MAX_ESTIMADO", "sum"))
                .reset_index()
                .sort_values(["INICIO_SEMANA", "DIA_SEMANA_NUM"])
            )

            semanas = list(resumo_carro_dias["INICIO_SEMANA"].drop_duplicates())
            for row_idx, inicio_semana in enumerate(semanas, start=8):
                ws_carro_dias.cell(row=row_idx, column=1, value=inicio_semana.strftime("%d/%m/%Y"))
                for dia_num in range(7):
                    dados_dia = resumo_carro_dias[
                        (resumo_carro_dias["INICIO_SEMANA"] == inicio_semana)
                        & (resumo_carro_dias["DIA_SEMANA_NUM"] == dia_num)
                    ]
                    valor_min = 0 if dados_dia.empty else float(dados_dia["VALOR_MIN_ESTIMADO"].sum())
                    valor_max = 0 if dados_dia.empty else float(dados_dia["VALOR_MAX_ESTIMADO"].sum())
                    ws_carro_dias.cell(row=row_idx, column=dia_num * 2 + 2, value=valor_min)
                    ws_carro_dias.cell(row=row_idx, column=dia_num * 2 + 3, value=valor_max)
                ws_carro_dias.cell(row=row_idx, column=16, value=f"=SUM(B{row_idx},D{row_idx},F{row_idx},H{row_idx},J{row_idx},L{row_idx},N{row_idx})")
                ws_carro_dias.cell(row=row_idx, column=17, value=f"=SUM(C{row_idx},E{row_idx},G{row_idx},I{row_idx},K{row_idx},M{row_idx},O{row_idx})")

            total_row_carro_dias = ws_carro_dias.max_row + 1
            ws_carro_dias.cell(row=total_row_carro_dias, column=1, value="TOTAL")
            for col_idx in range(2, 18):
                letra = get_column_letter(col_idx)
                ws_carro_dias.cell(row=total_row_carro_dias, column=col_idx, value=f"=SUM({letra}8:{letra}{total_row_carro_dias-1})")

            for col in range(1, 18):
                ws_carro_dias.cell(row=total_row_carro_dias, column=col).fill = total_fill
                ws_carro_dias.cell(row=total_row_carro_dias, column=col).font = Font(bold=True)
                ws_carro_dias.cell(row=total_row_carro_dias, column=col).border = Border(top=thin_top)

            for row in ws_carro_dias.iter_rows(min_row=8, max_row=total_row_carro_dias, min_col=2, max_col=17):
                for cell in row:
                    cell.number_format = currency_fmt

            ws_carro_dias.freeze_panes = "A8"
            ws_carro_dias.column_dimensions["A"].width = 20
            for col_idx in range(2, 18):
                ws_carro_dias.column_dimensions[get_column_letter(col_idx)].width = 15

    if not contratos_gerados and not contrato_carro_estimado_gerado:
        logger("ℹ️ Nenhuma nota pagável encontrada para os contratos de medição.")
        return None

    caminho_saida.parent.mkdir(exist_ok=True)
    wb.save(caminho_saida)
    logger(f"✅ Planilha de medição gerada: {caminho_saida}")
    try:
        abrir_arquivo_ou_pasta(caminho_saida)
    except Exception:
        pass
    return str(caminho_saida)

def carregar_dataframe_resultado(caminho_csv):
    caminho_csv = Path(caminho_csv)
    encodings = ["utf-8-sig", "utf-8", "latin-1", "cp1252"]
    for enc in encodings:
        try:
            return pd.read_csv(caminho_csv, sep=";", dtype=str, encoding=enc)
        except Exception:
            continue
    raise RuntimeError(f"Não foi possível ler o CSV tratado: {caminho_csv}")


def processar_arquivos_por_periodo(tipo_periodo="dia", eas_list=None, logger=None):
    logger = logger or log
    PROCESS_INPUT_FOLDER.mkdir(exist_ok=True)
    PROCESS_OUTPUT_FOLDER.mkdir(exist_ok=True)
    PROCESS_CACHE_FOLDER.mkdir(exist_ok=True)

    for f in PROCESS_INPUT_FOLDER.glob("*.csv"):
        try:
            f.unlink()
        except Exception:
            pass

    eas_list = eas_list or list(DEFAULT_EAS)
    data_inicio, data_fim = obter_periodo_atual(tipo_periodo)
    saidas = _nome_arquivo_periodo(tipo_periodo)

    logger("🧹 Pasta input limpa.")
    logger(
        "🔎 Procurando CSVs para o período de "
        f"{data_inicio.strftime('%d/%m/%Y')} até {data_fim.strftime('%d/%m/%Y')}..."
    )

    arquivos_csv = []
    total_eas = max(1, len(eas_list))

    for idx_ea, eas in enumerate(eas_list, start=1):
        variantes_ea = obter_variantes_ea_processamento(eas)
        melhor_por_dia = {}

        for f in iterar_csvs_disponiveis():
            nome_norm = normalizar_texto_processamento(f.name).replace("ea", "")
            if not any(variacao in nome_norm for variacao in variantes_ea):
                continue

            data_ref = extrair_data_do_nome_arquivo(f.name)
            if data_ref is None:
                logger(f"ℹ️ Ignorado sem data no nome: {f.name}")
                continue

            if not (data_inicio <= data_ref <= data_fim):
                continue

            mtime = f.stat().st_mtime
            atual = melhor_por_dia.get(data_ref)
            if atual is None or mtime > atual[0]:
                melhor_por_dia[data_ref] = (mtime, f)

        if melhor_por_dia:
            for data_ref in sorted(melhor_por_dia.keys()):
                _, arquivo = melhor_por_dia[data_ref]
                arquivo_base = arquivo
                if str(arquivo).startswith(str(DOWNLOADS_DIR)):
                    arquivo_base = copiar_para_cache_csv(arquivo, logger=logger)
                destino = PROCESS_INPUT_FOLDER / arquivo_base.name
                shutil.copy2(arquivo_base, destino)
                arquivos_csv.append(destino)
                logger(
                    f"📥 Copiado ({data_ref.strftime('%d/%m')}) [{eas}]: "
                    f"{arquivo_base.name} → {destino.name}"
                )
        else:
            logger(f"⚠️ Nenhum arquivo encontrado no período para {eas}")

        if CURRENT_APP is not None:
            try:
                CURRENT_APP.emit("progress", {"atual": idx_ea, "total": total_eas, "nome": eas})
                CURRENT_APP.emit("counter", {"exportadas": idx_ea, "total": total_eas})
            except Exception:
                pass

    if not arquivos_csv:
        raise RuntimeError(
            "Nenhum CSV encontrado na pasta Downloads para o período selecionado."
        )

    def normalize(col_name):
        if not isinstance(col_name, str):
            return ""
        s = unicodedata.normalize("NFKD", col_name).encode("ASCII", "ignore").decode()
        return re.sub(r"[^0-9a-zA-Z]", "", s).lower()

    def find_column(df, candidates):
        norm_map = {normalize(c): c for c in df.columns}
        for cand in candidates:
            if cand in norm_map:
                return norm_map[cand]
        return None

    def split_single_column_dataframe(df):
        if df is None or df.shape[1] != 1:
            return df

        col = df.columns[0]
        serie = df[col].dropna().astype(str)
        if serie.empty:
            return df

        texto = "\n".join(serie.tolist())
        for sep in [";", ",", "\t", "|"]:
            try:
                novo = pd.read_csv(StringIO(texto), sep=sep, dtype=str, engine="python")
                if novo.shape[1] > 1:
                    return novo
            except Exception:
                continue

        return df

    def try_read(file_path):
        encodings = ["utf-8-sig", "utf-8", "latin-1", "cp1252"]
        seps = [";", ",", "\t", "|"]

        for enc in encodings:
            for sep in seps:
                try:
                    df = pd.read_csv(
                        file_path,
                        sep=sep,
                        encoding=enc,
                        dtype=str,
                        engine="python",
                    )
                    df = split_single_column_dataframe(df)
                    if df.shape[1] >= 2:
                        return df
                except Exception:
                    continue

        try:
            df = pd.read_csv(file_path, sep=None, engine="python", dtype=str)
            df = split_single_column_dataframe(df)
            return df
        except Exception:
            return None

    def map_grupo(v):
        if pd.isna(v):
            return ""

        s = str(v).strip().lower()

        if "verificar instalacao auto" in s or "verificar instalação auto" in s or "auto-religada" in s or "auto religada" in s:
            return "VERIFICACAO"

        if any(
            k in s
            for k in [
                "corte",
                "progressivo",
                "perdas",
                "inativo",
                "leitura progressiva",
            ]
        ):
            return "CORTE"

        if any(k in s for k in ["relig", "religar", "religi"]):
            return "RELIGUE"

        return ""

    ordem_cands = ["ordemdeservico", "ordemdeserviço", "ordem"]
    recurso_cands = ["recurso"]
    eletricista1_cands = ["eletricista1", "eletricista", "id_eletricista"]
    eletricista2_cands = ["eletricista2", "id_eletricista2"]
    data_enc_cands = ["dataencerramento", "dataencerr", "data"]
    hora_enc_cands = ["horaencerramento", "horaencerr", "hora"]
    tipo_atividade_cands = ["tipodeatividade", "tipodeatividade_1", "tipodeatividade1"]
    motivo_rej_cands = [
        "motivodarejeicao",
        "motivodarejeição",
        "motivodarejeicao_com_retorno",
        "historicomotivodarejeicaocomretorno",
    ]

    df_list = []
    total_rows_in = 0
    files = glob.glob(str(PROCESS_INPUT_FOLDER / "*.csv"))

    for f in files:
        logger(f"📄 Lendo: {os.path.basename(f)}")
        nome_arquivo = os.path.basename(f)
        data_base = extrair_data_do_nome_arquivo(nome_arquivo)

        if data_base is None:
            data_base = datetime.fromtimestamp(Path(f).stat().st_mtime).date()

        df = try_read(f)
        if df is None:
            logger(" -> Erro ao ler arquivo. Pulando.")
            continue

        total_rows_in += len(df)

        ordem_col = find_column(df, ordem_cands)
        recurso_col = find_column(df, recurso_cands)
        eletricista1_col = find_column(df, eletricista1_cands)
        eletricista2_col = find_column(df, eletricista2_cands)
        data_col = find_column(df, data_enc_cands)
        hora_col = find_column(df, hora_enc_cands)
        tipo_col = find_column(df, tipo_atividade_cands)
        motivo_col = find_column(df, motivo_rej_cands)

        if hora_col is None or data_col is None:
            logger(" -> Sem colunas de data/hora. Pulando.")
            continue

        df = df[df[hora_col].notna() & (df[hora_col].astype(str).str.strip() != "")]
        if df.empty:
            logger(" -> Nenhuma linha válida. Pulando.")
            continue

        combined = (
            df[data_col].fillna("").astype(str).str.strip()
            + " "
            + df[hora_col].fillna("").astype(str).str.strip()
        )

        with pd.option_context("mode.chained_assignment", None):
            sample = combined.dropna().iloc[0] if not combined.dropna().empty else ""
            if re.match(r"\d{4}-\d{2}-\d{2}", sample):
                dt_parsed = pd.to_datetime(combined, dayfirst=False, errors="coerce")
            else:
                dt_parsed = pd.to_datetime(combined, dayfirst=True, errors="coerce")

        df["DATA_DT"] = dt_parsed.dt.date

        linhas_antes = len(df)
        if tipo_periodo == "dia":
            df = df[df["DATA_DT"] == data_base]
            if linhas_antes - len(df) > 0:
                logger(
                    f" -> {linhas_antes - len(df)} linha(s) descartadas "
                    f"por data diferente de {data_base.strftime('%d/%m/%Y')}."
                )
        else:
            df = df[(df["DATA_DT"] >= data_inicio) & (df["DATA_DT"] <= data_fim)]
            if linhas_antes - len(df) > 0:
                logger(
                    f" -> {linhas_antes - len(df)} linha(s) fora do período "
                    f"{data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}."
                )

        if df.empty:
            logger(" -> Nenhuma linha do período. Pulando.")
            continue

        df["DATA_ENCERRAMENTO"] = dt_parsed.dt.strftime("%d/%m/%Y %H:%M")
        df["GRUPO_NOTA"] = df[tipo_col].apply(map_grupo) if tipo_col else ""

        def map_final(v):
            if pd.isna(v) or str(v).strip() == "":
                return "FINALIZADA"
            return "REJEITADA"

        df["FINALIZACAO"] = df[motivo_col].apply(map_final) if motivo_col else "FINALIZADA"
        df["RECUSA"] = df[motivo_col].fillna("") if motivo_col else ""

        ordem_values = df[ordem_col] if ordem_col else pd.Series([""] * len(df), index=df.index)
        recurso_values = df[recurso_col] if recurso_col else pd.Series([""] * len(df), index=df.index)
        eletricista1_values = df[eletricista1_col] if eletricista1_col else pd.Series([""] * len(df), index=df.index)
        eletricista2_values = df[eletricista2_col] if eletricista2_col else pd.Series([""] * len(df), index=df.index)

        df_out = pd.DataFrame(
            {
                "ORDEM_DE_SERVICO": ordem_values,
                "GRUPO_NOTA": df["GRUPO_NOTA"],
                "RECURSO": recurso_values,
                "FINALIZACAO": df["FINALIZACAO"],
                "DATA_ENCERRAMENTO": df["DATA_ENCERRAMENTO"],
                "ELETRICISTA1": eletricista1_values,
                "ELETRICISTA2": eletricista2_values,
                "RECUSA": df["RECUSA"],
            }
        )

        df_list.append(df_out)

    if not df_list:
        raise RuntimeError("Nenhum dado válido foi processado dos CSVs copiados.")

    final = pd.concat(df_list, ignore_index=True)
    final = final[final["GRUPO_NOTA"].notna() & (final["GRUPO_NOTA"].str.strip() != "")]

    final["ORDEM_DE_SERVICO"] = (
        final["ORDEM_DE_SERVICO"]
        .astype(str)
        .str.strip()
        .str.replace(r"[^0-9]", "", regex=True)
    )
    final["ELETRICISTA1"] = final["ELETRICISTA1"].fillna("").astype(str).str.strip()
    final["ELETRICISTA2"] = final["ELETRICISTA2"].fillna("").astype(str).str.strip()
    final = final[final["ORDEM_DE_SERVICO"].str.strip() != ""]

    duplicadas = final[final.duplicated(subset=["ORDEM_DE_SERVICO"], keep=False)]
    if not duplicadas.empty:
        logger(f"⚠️ {len(duplicadas)} linhas duplicadas removidas (ordens repetidas).")
        final = final.drop_duplicates(subset=["ORDEM_DE_SERVICO"], keep="first")

    def _detectar_alertas_stc(df_):
        if df_.empty:
            return []

        df_alerta = df_.copy()
        df_alerta["RECURSO"] = df_alerta["RECURSO"].fillna("").astype(str).str.strip()
        df_alerta["ELETRICISTA1"] = df_alerta["ELETRICISTA1"].fillna("").astype(str).str.strip()
        df_alerta["ELETRICISTA2"] = df_alerta["ELETRICISTA2"].fillna("").astype(str).str.strip()

        mask_stc = df_alerta["RECURSO"].str.upper().str.startswith("JUN58")
        mask_um_executor = (df_alerta["ELETRICISTA1"] != "") & (df_alerta["ELETRICISTA2"] == "")
        df_alerta = df_alerta[mask_stc & mask_um_executor].copy()

        if df_alerta.empty:
            return []

        resumo = (
            df_alerta.groupby("RECURSO", dropna=False)
            .agg(QTD_NOTAS=("ORDEM_DE_SERVICO", "count"))
            .reset_index()
            .sort_values(["QTD_NOTAS", "RECURSO"], ascending=[False, True])
        )

        alertas = []
        for row in resumo.itertuples(index=False):
            alertas.append({
                "equipe": row.RECURSO,
                "qtd_notas": int(row.QTD_NOTAS),
            })
        return alertas

    alertas_stc_um_executor = _detectar_alertas_stc(final)
    if alertas_stc_um_executor:
        total_notas_alerta = sum(item["qtd_notas"] for item in alertas_stc_um_executor)
        logger(
            f"⚠️ Alerta STC: {total_notas_alerta} nota(s) do contrato carro com apenas 1 executor, "
            f"em {len(alertas_stc_um_executor)} equipe(s)."
        )
        for item in alertas_stc_um_executor:
            logger(f"   • {item['equipe']}: {item['qtd_notas']} nota(s)")

    final.to_csv(saidas["csv"], sep=";", index=False, encoding="utf-8-sig")
    final.to_csv(saidas["txt"], sep="\t", index=False, header=False, encoding="utf-8")

    logger(f"✅ CSV salvo: {saidas['csv']}")
    logger(f"✅ TXT salvo: {saidas['txt']}")

    bases_dashboard = gerar_bases_dashboard(
        final,
        tipo_periodo,
        data_inicio,
        data_fim,
        logger=logger,
    )

    github_dashboard = enviar_dashboard_para_github(bases_dashboard, logger=logger)

    # Depois de enviar os CSVs, gera o SQLite leve e envia o gzus_dashboard.db automaticamente.
    pos_extrator_gzus = executar_pos_extrator_gzus(bases_dashboard, logger=logger)

    def tempo_inatividade(df_):
        df_ = df_.copy()

        df_["RECURSO"] = df_["RECURSO"].fillna("SEM_RECURSO").astype(str)
        df_.loc[df_["RECURSO"].str.strip() == "", "RECURSO"] = "SEM_RECURSO"

        df_["DATA_ENCERRAMENTO_DT"] = pd.to_datetime(
            df_["DATA_ENCERRAMENTO"], errors="coerce"
        )

        ultimos = df_.groupby("RECURSO")["DATA_ENCERRAMENTO_DT"].max().reset_index()
        agora = datetime.now()

        def fmt_delta(x):
            if pd.isna(x):
                return "N/A"
            delta = agora - x
            total_horas = delta.days * 24 + delta.seconds // 3600
            mins = (delta.seconds // 60) % 60
            return f"{total_horas}h {mins}min"

        ultimos["TEMPO_INATIVIDADE"] = ultimos["DATA_ENCERRAMENTO_DT"].apply(fmt_delta)
        return ultimos[["RECURSO", "TEMPO_INATIVIDADE"]]

    inatividade = tempo_inatividade(final)
    inatividade.to_excel(saidas["inatividade"], index=False)

    medicao_disjuntor_jundiai = None
    if tipo_periodo in ("mes", "mes_anterior"):
        medicao_disjuntor_jundiai = gerar_planilha_medicao_disjuntor_jundiai(
            final,
            obter_saida_medicao_disjuntor_jundiai(tipo_periodo),
            data_inicio.strftime("%d/%m/%Y"),
            data_fim.strftime("%d/%m/%Y"),
            logger=logger,
        )

    logger(f"✅ Planilha de inatividade: {saidas['inatividade']}")
    if medicao_disjuntor_jundiai:
        logger(f"✅ Planilha de medição: {medicao_disjuntor_jundiai}")
    logger(f"📊 Linhas lidas: {total_rows_in}")
    logger(f"📊 Linhas finais: {len(final)}")

    return {
        "tipo_periodo": tipo_periodo,
        "periodo_inicio": data_inicio.strftime("%d/%m/%Y"),
        "periodo_fim": data_fim.strftime("%d/%m/%Y"),
        "csv": str(saidas["csv"]),
        "txt": str(saidas["txt"]),
        "inatividade": str(saidas["inatividade"]),
        "medicao_disjuntor_jundiai": medicao_disjuntor_jundiai,
        "bases_dashboard": bases_dashboard,
        "github_dashboard": github_dashboard,
        "alertas_stc_um_executor": alertas_stc_um_executor,
        "total_alertas_stc_um_executor": int(sum(item["qtd_notas"] for item in alertas_stc_um_executor)),
        "linhas_lidas": total_rows_in,
        "linhas_finais": int(len(final)),
        "arquivos_copiados": [str(p) for p in arquivos_csv],
    }


def processar_arquivos_baixados(eas_list=None, logger=None):
    return processar_arquivos_por_periodo(
        tipo_periodo="dia",
        eas_list=eas_list,
        logger=logger,
    )




COR_BG = "#020617"
COR_BG_2 = "#0b1220"
COR_CARD = "#0f172a"
COR_CARD_2 = "#111827"
COR_BORDA = "#22304a"
COR_TEXTO = "#f8fafc"
COR_TEXTO_2 = "#94a3b8"
COR_DESTAQUE = "#3b82f6"
COR_DESTAQUE_2 = "#2563eb"
COR_OK = "#16a34a"
COR_ALERTA = "#d97706"
COR_ERRO = "#dc2626"
COR_LOG = "#0f172a"


DEFAULT_EAS = [
    "AVARE",
    "ITAPETININGA",
    "JAGUARIUNA",
    "MOCOCA",
    "OURINHOS",
    "PEDREIRA",
    "PIRAJU",
    "SANTA CRUZ DO RIO PARDO",
    "CERQUEIRA CESAR",
    "CASA BRANCA",
    "CAMPO LIMPO PAULISTA",
    "EA ITUPEVA",
    "EA VINHEDO",
    "EA JUNDIA",
    "EA INDAIATUBA",
    "EA SALTO",
    "SAO JOSE DO RIO PARDO",
    "SARAPUI",
    "SAO MIGUEL ARCAJNO",
]

CURRENT_APP = None


def abrir_arquivo_ou_pasta(caminho: Path):
    try:
        caminho = str(caminho)
        if sys.platform.startswith("win"):
            os.startfile(caminho)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", caminho])
        else:
            subprocess.Popen(["xdg-open", caminho])
    except Exception as e:
        print(f"⚠️ Não foi possível abrir: {e}")


def log(msg: str):
    print(msg)
    if CURRENT_APP is not None:
        try:
            CURRENT_APP.push_log(msg)
        except Exception:
            pass


class QueueWriter:
    def __init__(self, fila: queue.Queue):
        self.fila = fila

    def write(self, texto):
        if texto and texto.strip():
            self.fila.put(("log", texto.rstrip()))

    def flush(self):
        pass


class ExtratorProducao:
    def __init__(self, fila=None, headless=False, eas=None):
        self.fila = fila
        self.headless = headless
        self.eas = eas or list(DEFAULT_EAS)
        self.driver = None
        self.wait = None
        self.total_eas = len(self.eas)
        self.exportadas = 0
        self.configuracao_hierarquica_feita = False

    def emit(self, tipo, valor):
        if self.fila is not None:
            self.fila.put((tipo, valor))

    def set_status(self, texto, subtexto=None):
        self.emit("status", texto)
        if subtexto is not None:
            self.emit("substatus", subtexto)

    def push_progress(self, atual, total, nome=""):
        self.emit("progress", {"atual": atual, "total": total, "nome": nome})

    def update_counter(self):
        self.emit("counter", {"exportadas": self.exportadas, "total": self.total_eas})

    def preparar(self):
        load_dotenv(override=True)
        self.matricula = os.getenv("MATRICULA")
        self.senha = os.getenv("SENHA")
        if not self.matricula or not self.senha:
            raise RuntimeError("Crie ou preencha o arquivo .env com MATRICULA e SENHA antes de iniciar.")

        chrome_options = Options()
        chrome_options.add_argument("--disable-notifications")
        chrome_options.add_argument("--disable-popup-blocking")
        chrome_options.add_argument("--disable-infobars")
        chrome_options.add_argument("--start-maximized")
        chrome_options.add_argument("--window-size=1920,1080")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        if self.headless:
            chrome_options.add_argument("--headless=new")

        self.driver = webdriver.Chrome(options=chrome_options)
        self.wait = WebDriverWait(self.driver, 25)

    def preencher(self, by, sel, txt, timeout=15):
        try:
            el = WebDriverWait(self.driver, timeout).until(EC.element_to_be_clickable((by, sel)))
            try:
                el.clear()
            except Exception:
                pass
            el.send_keys(txt)
            return True
        except Exception:
            return False

    def existe(self, by, sel, timeout=5):
        try:
            WebDriverWait(self.driver, timeout).until(EC.presence_of_element_located((by, sel)))
            return True
        except Exception:
            return False

    def click_js(self, el):
        self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
        time.sleep(0.25)
        self.driver.execute_script("arguments[0].click();", el)

    def click_safe(self, el):
        try:
            el.click()
        except Exception:
            self.click_js(el)

    def tentar_login_oracle(self):
        log("🔐 Tentando login Oracle...")
        usuario_ok = (
            self.preencher(By.ID, "userid", self.matricula, timeout=8)
            or self.preencher(By.ID, "username", self.matricula, timeout=8)
            or self.preencher(By.NAME, "userid", self.matricula, timeout=8)
            or self.preencher(By.NAME, "username", self.matricula, timeout=8)
        )
        if not usuario_ok:
            log("ℹ️ Campos Oracle não encontrados.")
            return False

        senha_ok = (
            self.preencher(By.ID, "password", self.senha, timeout=8)
            or self.preencher(By.NAME, "password", self.senha, timeout=8)
        )
        if senha_ok:
            self.driver.switch_to.active_element.send_keys(Keys.ENTER)
            log("✅ Login Oracle enviado")
            time.sleep(4)
            return True

        log("ℹ️ Usuário Oracle encontrado, mas senha não apareceu.")
        return False

    def tentar_login_adfs(self):
        log("🔐 Tentando login ADFS...")
        usuario_ok = (
            self.preencher(By.ID, "userNameInput", self.matricula, timeout=8)
            or self.preencher(By.NAME, "UserName", self.matricula, timeout=8)
        )
        if not usuario_ok:
            log("ℹ️ Campos ADFS não encontrados.")
            return False

        try:
            campo_senha = WebDriverWait(self.driver, 8).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='password']"))
            )
            try:
                campo_senha.clear()
            except Exception:
                pass
            campo_senha.send_keys(self.senha)
            self.driver.switch_to.active_element.send_keys(Keys.ENTER)
            log("✅ Login ADFS enviado")
            time.sleep(4)
            return True
        except Exception:
            log("ℹ️ Campo de senha ADFS não encontrado.")
            return False

    def fazer_login_duplo(self):
        self.set_status("Abrindo Oracle Cloud", "Iniciando autenticação")
        log("🌐 Abrindo Oracle...")
        self.driver.get("https://cpflcloud.fs.ocs.oraclecloud.com/")
        self.driver.maximize_window()
        time.sleep(3)

        self.tentar_login_oracle()
        self.tentar_login_adfs()

        if self.existe(By.ID, "userNameInput", timeout=5) or self.existe(By.ID, "userid", timeout=5):
            log("🔁 Detectado segundo login. Tentando novamente...")
            self.tentar_login_oracle()
            self.tentar_login_adfs()

        log("⏳ Carregando ambiente...")
        time.sleep(7)

    def abrir_console(self):
        self.set_status("Entrando no Console de Alocação", "Aguardando tela principal")
        log("🔍 Aguardando Console de Alocação...")
        link = self.wait.until(
            EC.presence_of_element_located((By.XPATH, "//*[contains(text(),'Console de Alocação')]"))
        )
        self.click_js(link)
        log("✅ Console aberto!")
        time.sleep(4)

    def expandir_grupo(self, nome):
        try:
            seta = self.wait.until(
                EC.element_to_be_clickable((
                    By.XPATH,
                    f'//span[contains(text(),"{nome}")]/preceding::button[@aria-label="Expandir"][1]'
                ))
            )
            self.click_js(seta)
            log(f"⤵️ {nome} expandido")
            time.sleep(0.8)
        except Exception:
            log(f"ℹ️ {nome} já estava expandido ou não achei a seta")

    def foco_painel_eas(self):
        try:
            filtro = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, "//input[@placeholder='Nome ou ID' or @aria-label='Nome ou ID']"))
            )
            filtro.click()
            filtro.send_keys(Keys.ESCAPE)
            time.sleep(0.2)
        except Exception:
            pass

    def localizar_ea(self, xpaths_variantes, max_scrolls=8):
        self.foco_painel_eas()

        for xp in xpaths_variantes:
            els = self.driver.find_elements(By.XPATH, xp)
            if els:
                return els[0]

        for _ in range(max_scrolls):
            self.driver.switch_to.active_element.send_keys(Keys.PAGE_DOWN)
            time.sleep(0.35)
            for xp in xpaths_variantes:
                els = self.driver.find_elements(By.XPATH, xp)
                if els:
                    return els[0]

        for _ in range(max_scrolls):
            self.driver.switch_to.active_element.send_keys(Keys.PAGE_UP)
            time.sleep(0.25)
            for xp in xpaths_variantes:
                els = self.driver.find_elements(By.XPATH, xp)
                if els:
                    return els[0]

        return None

    def menu_exibir_aberto(self):
        return self.existe(By.XPATH, "//*[contains(normalize-space(.), 'Aplicar de forma hierárquica')]", timeout=3)

    def abrir_menu_exibir(self):
        exibir_btn = self.wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[aria-label='Exibir']")))
        self.click_safe(exibir_btn)
        time.sleep(1.2)

    def obter_popup_exibir(self, timeout=10):
        fim = time.time() + timeout
        while time.time() < fim:
            candidatos = self.driver.find_elements(By.XPATH, "//div[.//*[contains(normalize-space(.), 'Aplicar de forma hierárquica')]]")
            for popup in reversed(candidatos):
                try:
                    if popup.is_displayed():
                        return popup
                except Exception:
                    continue
            time.sleep(0.2)
        raise Exception("Popup do menu Exibir não encontrado.")

    def marcar_checkbox_hierarquica(self, popup):
        log("☑️ Marcando 'Aplicar de forma hierárquica'...")
        checkbox = None
        for xp in [
            ".//input[@type='checkbox' and contains(@name, 'recursively')]",
            ".//label[.//*[contains(normalize-space(.), 'Aplicar de forma hierárquica')]]",
            ".//*[contains(normalize-space(.), 'Aplicar de forma hierárquica')]/ancestor::label[1]",
            ".//*[contains(normalize-space(.), 'Aplicar de forma hierárquica')]/ancestor::*[contains(@class,'oj-choice-item')][1]",
        ]:
            candidatos = popup.find_elements(By.XPATH, xp)
            for el in candidatos:
                try:
                    if el.is_displayed():
                        checkbox = el
                        break
                except Exception:
                    continue
            if checkbox:
                break

        if not checkbox:
            raise Exception("Não encontrei o controle de 'Aplicar de forma hierárquica'.")

        try:
            self.driver.execute_script("arguments[0].scrollIntoView({block:'center', inline:'nearest'});", checkbox)
            time.sleep(0.2)
        except Exception:
            pass

        try:
            checkbox.click()
        except Exception:
            try:
                self.click_js(checkbox)
            except Exception:
                self.driver.execute_script(
                    "arguments[0].dispatchEvent(new MouseEvent('mousedown', {bubbles:true}));"
                    "arguments[0].dispatchEvent(new MouseEvent('mouseup', {bubbles:true}));"
                    "arguments[0].dispatchEvent(new MouseEvent('click', {bubbles:true}));",
                    checkbox,
                )
        time.sleep(0.4)

    def clicar_aplicar_popup(self, popup):
        log("🖱️ Confirmando popup do 'Exibir'...")
        clicou = False

        candidatos_xpath = [
            ".//button[normalize-space()='Aplicar']",
            ".//button[contains(normalize-space(.), 'Aplicar')]",
            ".//*[@role='button' and normalize-space()='Aplicar']",
            ".//*[@role='button' and contains(normalize-space(.), 'Aplicar')]",
            ".//oj-button[.//*[contains(normalize-space(.), 'Aplicar')]]",
            ".//button[normalize-space()='OK']",
            ".//button[contains(normalize-space(.), 'OK')]",
            ".//*[@role='button' and normalize-space()='OK']",
            ".//*[@role='button' and contains(normalize-space(.), 'OK')]",
            ".//button[contains(normalize-space(.), 'Confirmar')]",
            ".//*[@role='button' and contains(normalize-space(.), 'Confirmar')]",
        ]

        for xp in candidatos_xpath:
            try:
                elementos = popup.find_elements(By.XPATH, xp)
            except Exception:
                elementos = []

            for el in elementos:
                try:
                    if not el.is_displayed():
                        continue
                    self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                    time.sleep(0.2)
                    self.click_safe(el)
                    clicou = True
                    log("✅ Botão do popup clicado diretamente.")
                    break
                except Exception:
                    continue

            if clicou:
                break

        if not clicou:
            for xp in [
                "//button[normalize-space()='Aplicar']",
                "//button[contains(normalize-space(.), 'Aplicar')]",
                "//*[@role='button' and normalize-space()='Aplicar']",
                "//*[@role='button' and contains(normalize-space(.), 'Aplicar')]",
                "//button[normalize-space()='OK']",
                "//button[contains(normalize-space(.), 'OK')]",
                "//*[@role='button' and normalize-space()='OK']",
                "//*[@role='button' and contains(normalize-space(.), 'OK')]",
                "//button[contains(normalize-space(.), 'Confirmar')]",
                "//*[@role='button' and contains(normalize-space(.), 'Confirmar')]",
            ]:
                try:
                    elementos = self.driver.find_elements(By.XPATH, xp)
                except Exception:
                    elementos = []

                for el in elementos:
                    try:
                        if not el.is_displayed():
                            continue
                        self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                        time.sleep(0.2)
                        self.click_safe(el)
                        clicou = True
                        log("✅ Botão de confirmação encontrado fora do container do popup.")
                        break
                    except Exception:
                        continue

                if clicou:
                    break

        if not clicou:
            log("ℹ️ Não achei botão visível. Tentando confirmar por teclado...")
            for _ in range(3):
                try:
                    self.driver.switch_to.active_element.send_keys(Keys.TAB)
                    time.sleep(0.2)
                except Exception:
                    pass

            for tecla in (Keys.SPACE, Keys.ENTER):
                try:
                    self.driver.switch_to.active_element.send_keys(tecla)
                    clicou = True
                    log("✅ Popup confirmado por teclado.")
                    break
                except Exception:
                    continue

        if not clicou:
            try:
                body = self.driver.find_element(By.TAG_NAME, "body")
                body.send_keys(Keys.ENTER)
                clicou = True
                log("✅ Popup confirmado com ENTER no body.")
            except Exception:
                pass

        time.sleep(1.2)

        if self.menu_exibir_aberto():
            log("⚠️ O menu ainda parece aberto após confirmar. Tentando fechar novamente...")
            try:
                self.driver.switch_to.active_element.send_keys(Keys.ESCAPE)
                time.sleep(0.6)
            except Exception:
                pass

        log("✅ Fluxo de confirmação do popup concluído.")

    def fechar_menu_exibir_se_aberto(self):
        if self.menu_exibir_aberto():
            try:
                body = self.driver.find_element(By.TAG_NAME, "body")
                self.click_js(body)
                time.sleep(0.5)
            except Exception:
                try:
                    self.driver.switch_to.active_element.send_keys(Keys.ESCAPE)
                    time.sleep(0.5)
                except Exception:
                    pass

    def marcar_aplicar_forma_hierarquica(self):
        self.set_status("Configurando exibição", "Aplicando forma hierárquica uma única vez")
        log("⚙️ Configurando 'Aplicar de forma hierárquica' automaticamente...")
        self.abrir_menu_exibir()
        popup = self.obter_popup_exibir()
        self.marcar_checkbox_hierarquica(popup)
        self.clicar_aplicar_popup(popup)
        self.fechar_menu_exibir_se_aberto()

        if self.menu_exibir_aberto():
            log("⚠️ O menu 'Exibir' ainda ficou aberto. Tentando fechar com ESC...")
            try:
                self.driver.switch_to.active_element.send_keys(Keys.ESCAPE)
                time.sleep(0.8)
            except Exception:
                pass

        log("✅ Configuração automática concluída. O bot não abrirá esse menu novamente.")

    def exportar(self):
        log("💾 Exportando...")

        # Garante que nenhum popup residual do menu Exibir esteja bloqueando a barra.
        try:
            self.fechar_menu_exibir_se_aberto()
            time.sleep(0.4)
        except Exception:
            pass

        try:
            body = self.driver.find_element(By.TAG_NAME, "body")
            body.send_keys(Keys.ESCAPE)
            time.sleep(0.2)
        except Exception:
            pass

        acoes_btn = self.wait.until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "[aria-label='Ações']"))
        )
        self.click_safe(acoes_btn)
        time.sleep(0.8)

        exportar_el = self.wait.until(
            EC.element_to_be_clickable((By.XPATH, "//span[contains(normalize-space(.), 'Exportar')]"))
        )
        self.click_safe(exportar_el)
        log("📦 Exportado!")
        time.sleep(2)

        try:
            self.driver.switch_to.active_element.send_keys(Keys.ESCAPE)
        except Exception:
            pass


    def executar(self):
        try:
            self.preparar()
            self.update_counter()
            self.fazer_login_duplo()
            self.abrir_console()
            self.expandir_grupo("EAS SANTA CRUZ")
            self.expandir_grupo("EAS SUDESTE")

            xpaths_eas = [
                [f"//span[contains(., '{variante}')]" for variante in EA_VARIANTES_PROCESSAMENTO.get(ea, [ea])]
                for ea in self.eas
            ]

            for idx, variantes in enumerate(xpaths_eas, start=1):
                nome_alvo = self.eas[idx - 1]
                self.set_status("Processando EAs", f"{idx}/{self.total_eas} · {nome_alvo}")
                self.push_progress(idx - 1, self.total_eas, nome_alvo)
                log(f"\n➡️ ({idx}/{self.total_eas}) Localizando EA...")
                el = self.localizar_ea(variantes)

                if not el:
                    log(f"⚠️ Não achou esta EA: {nome_alvo}. Pulando...")
                    continue

                self.click_js(el)
                time.sleep(0.5)
                try:
                    self.driver.switch_to.active_element.send_keys(Keys.ENTER)
                except Exception:
                    pass
                time.sleep(0.7)

                try:
                    nome_visivel = el.text.strip() or nome_alvo
                except Exception:
                    nome_visivel = nome_alvo
                log(f"🏢 Selecionada: {nome_visivel}")

                if not self.configuracao_hierarquica_feita:
                    log("⚙️ Configurando 'Aplicar de forma hierárquica' UMA ÚNICA VEZ, após selecionar a primeira EA...")
                    self.marcar_aplicar_forma_hierarquica()
                    self.configuracao_hierarquica_feita = True
                    log("✅ Configuração inicial concluída. Reselecionando a EA e seguindo para exportação...")

                    el = self.localizar_ea(variantes)
                    if not el:
                        log(f"⚠️ Não consegui reencontrar a EA após configurar o menu: {nome_alvo}. Pulando...")
                        continue

                    self.click_js(el)
                    time.sleep(0.5)
                    try:
                        self.driver.switch_to.active_element.send_keys(Keys.ENTER)
                    except Exception:
                        pass
                    time.sleep(0.7)
                    log(f"🔁 EA reselecionada após configurar o menu: {nome_visivel}")

                self.exportar()
                self.exportadas += 1
                self.update_counter()
                self.push_progress(idx, self.total_eas, nome_visivel)

            self.set_status("Finalizado", "Todas as EAs processadas")
            log("\n🎉 FINALIZADO — todas as EAs exportadas!")
            return {"ok": True, "exportadas": self.exportadas, "total": self.total_eas}

        except Exception as e:
            log(f"❌ ERRO GERAL: {e}")
            raise
        finally:
            try:
                if self.driver is not None:
                    log("🔚 Fechando navegador...")
                    self.driver.quit()
            except Exception:
                pass


class ExtratorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title(APP_TITULO)
        try:
            if APP_ICON_PATH.exists():
                self.root.iconbitmap(str(APP_ICON_PATH))
        except Exception:
            pass
        self.root.geometry("1180x690")
        self.root.minsize(1100, 640)
        self.root.configure(bg=COR_BG)

        self.fila = queue.Queue()
        self.executando = False
        self.worker = None
        self.exportadas = 0
        self.total = len(DEFAULT_EAS)
        self.auto_job = None
        self.proxima_execucao = None

        self.status_var = tk.StringVar(value="Pronto para iniciar")
        self.substatus_var = tk.StringVar(value="Configure as credenciais e clique em iniciar")
        self.counter_var = tk.StringVar(value=f"0/{self.total} exportadas")
        self.headless_var = tk.BooleanVar(value=(os.getenv("HEADLESS", "0") == "1"))
        self.auto_interval_var = tk.StringVar(value=os.getenv("AUTO_INTERVAL_LABEL", "Desativado"))
        self.auto_custom_minutes_var = tk.StringVar(value=os.getenv("AUTO_INTERVAL_MINUTES", "90"))
        self.auto_status_var = tk.StringVar(value="Agendamento automático desativado")
        self.matricula_var = tk.StringVar(value=os.getenv("MATRICULA", ""))
        self.senha_var = tk.StringVar(value=os.getenv("SENHA", ""))
        self.resumo_modo_var = tk.StringVar(value="Modo atual: Extração diária")
        self.resumo_periodo_var = tk.StringVar(value="Período padrão: hoje")
        self.resumo_eas_var = tk.StringVar(value=f"EAs selecionadas: {len(DEFAULT_EAS)}")
        self.resumo_saida_var = tk.StringVar(value="Saída pronta: nenhuma")
        self.ultima_acao_rotulo = "Extração diária"
        self.ultimo_txt_gerado = None
        self.ultimo_csv_gerado = None
        self.ultimo_periodo_inicio = None
        self.ultimo_periodo_fim = None
        self.ultima_medicao_gerada = None
        self._acao_medicao_somente = False
        self.ea_vars = {ea: tk.BooleanVar(value=True) for ea in DEFAULT_EAS}
        self.ea_filter_var = tk.StringVar(value="")
        self._carregar_estado_eas()

        self._build_ui()
        self._center()
        self.root.protocol("WM_DELETE_WINDOW", self.destroy)
        self._poll()



    def _carregar_estado_eas(self):
        try:
            if UI_STATE_PATH.exists():
                data = json.loads(UI_STATE_PATH.read_text(encoding="utf-8"))
                eas_salvas = data.get("eas_selecionadas")
                if isinstance(eas_salvas, list) and eas_salvas:
                    conjunto = set(eas_salvas)
                    for ea, var in self.ea_vars.items():
                        # EAs novas que ainda não existiam no ui_state.json antigo
                        # continuam marcadas por padrão, para entrarem no download e no tratamento.
                        if ea not in conjunto and ea not in eas_salvas and ea == "SAO MIGUEL ARCAJNO":
                            var.set(True)
                        else:
                            var.set(ea in conjunto)
        except Exception:
            pass

    def _salvar_estado_eas(self):
        try:
            data = {"eas_selecionadas": self.eas_selecionadas()}
            UI_STATE_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception:
            pass



    def _ea_display_list(self):
        filtro = self.ea_filter_var.get().strip().lower()
        if not filtro:
            return list(DEFAULT_EAS)
        return [ea for ea in DEFAULT_EAS if filtro in ea.lower()]

    def _toggle_ea(self, ea):
        self.ea_vars[ea].set(not self.ea_vars[ea].get())
        self._on_ea_selection_change()

    def _on_ea_filter_change(self, *_):
        self._render_ea_tiles()

    def _render_ea_tiles(self):
        if not hasattr(self, "ea_frame"):
            return

        for child in self.ea_frame.winfo_children():
            child.destroy()

        eas_visiveis = self._ea_display_list()

        for col in range(3):
            self.ea_frame.columnconfigure(col, weight=1)

        if hasattr(self, "ea_count_var"):
            self.ea_count_var.set(f"{len(self.eas_selecionadas())}/{len(DEFAULT_EAS)} selecionadas")

        if not eas_visiveis:
            tk.Label(
                self.ea_frame,
                text="Nenhuma EA encontrada para esse filtro.",
                bg=COR_CARD,
                fg=COR_TEXTO_2,
                font=("Segoe UI", 9),
                anchor="w",
            ).grid(row=0, column=0, columnspan=3, sticky="ew", pady=(6, 0))
            return

        for i, ea in enumerate(eas_visiveis):
            selecionada = self.ea_vars[ea].get()
            tile_bg = "#16233a" if selecionada else COR_CARD_2
            tile_border = "#3b82f6" if selecionada else COR_BORDA

            tile = tk.Frame(
                self.ea_frame,
                bg=tile_bg,
                highlightthickness=1,
                highlightbackground=tile_border,
                cursor="hand2",
            )
            linha = i // 3
            coluna = i % 3
            padx = (0, 8) if coluna < 2 else (0, 0)
            tile.grid(row=linha, column=coluna, sticky="ew", pady=4, padx=padx)

            chk = tk.Checkbutton(
                tile,
                text=ea,
                variable=self.ea_vars[ea],
                bg=tile_bg,
                fg=COR_TEXTO,
                activebackground=tile_bg,
                activeforeground=COR_TEXTO,
                selectcolor="#1d4ed8",
                font=("Segoe UI", 9, "bold"),
                relief="flat",
                bd=0,
                anchor="w",
                justify="left",
                command=self._on_ea_selection_change,
            )
            chk.pack(fill="x", padx=10, pady=8)

            tile.bind("<Button-1>", lambda e, nome=ea: self._toggle_ea(nome))
            chk.bind("<Button-1>", lambda e: None)



    def _build_ui(self):
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass

        style.configure("App.TFrame", background=COR_BG)
        style.configure("Title.TLabel", background=COR_BG, foreground=COR_TEXTO, font=("Segoe UI", 18, "bold"))
        style.configure("Sub.TLabel", background=COR_BG, foreground=COR_TEXTO_2, font=("Segoe UI", 9))
        style.configure("CardTitle.TLabel", background=COR_CARD, foreground=COR_TEXTO, font=("Segoe UI", 10, "bold"))
        style.configure("Info.TLabel", background=COR_CARD, foreground=COR_TEXTO, font=("Segoe UI", 11, "bold"))
        style.configure("Soft.TLabel", background=COR_CARD, foreground=COR_TEXTO_2, font=("Segoe UI", 8))
        style.configure("Primary.TButton", background=COR_DESTAQUE_2, foreground="#ffffff", padding=(14, 10), font=("Segoe UI", 9, "bold"), borderwidth=0)
        style.map("Primary.TButton", background=[("active", COR_DESTAQUE), ("pressed", COR_DESTAQUE_2)])
        style.configure("Secondary.TButton", background="#1e293b", foreground="#f8fafc", padding=(12, 10), font=("Segoe UI", 9, "bold"), borderwidth=0)
        style.map("Secondary.TButton", background=[("active", "#334155"), ("pressed", "#334155")], foreground=[("active", "#f8fafc"), ("pressed", "#f8fafc")])
        style.configure("Modern.Horizontal.TProgressbar", troughcolor="#dbe4f0", background=COR_DESTAQUE, bordercolor="#dbe4f0", lightcolor=COR_DESTAQUE, darkcolor=COR_DESTAQUE_2, thickness=12)

        viewport = tk.Frame(self.root, bg=COR_BG)
        viewport.pack(fill="both", expand=True)

        self.main_canvas = tk.Canvas(
            viewport,
            bg=COR_BG,
            highlightthickness=0,
            bd=0,
        )
        self.main_canvas.pack(side="left", fill="both", expand=True)

        vscroll = tk.Scrollbar(viewport, orient="vertical", command=self.main_canvas.yview)
        vscroll.pack(side="right", fill="y")
        self.main_canvas.configure(yscrollcommand=vscroll.set)

        outer = tk.Frame(self.main_canvas, bg=COR_BG)
        self._outer_window = self.main_canvas.create_window((0, 0), window=outer, anchor="nw")

        def _sync_scrollregion(event=None):
            self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all"))

        def _sync_outer_width(event):
            self.main_canvas.itemconfigure(self._outer_window, width=event.width)

        outer.bind("<Configure>", _sync_scrollregion)
        self.main_canvas.bind("<Configure>", _sync_outer_width)

        def _mousewheel(event):
            try:
                delta = event.delta
                if delta == 0 and getattr(event, "num", None) in (4, 5):
                    delta = 120 if event.num == 4 else -120
                step = int(-1 * (delta / 120)) if delta else 0
                if step:
                    self.main_canvas.yview_scroll(step, "units")
            except Exception:
                pass

        self.main_canvas.bind_all("<MouseWheel>", _mousewheel)
        self.main_canvas.bind_all("<Button-4>", _mousewheel)
        self.main_canvas.bind_all("<Button-5>", _mousewheel)

        sidebar = tk.Frame(outer, bg="#0f172a", width=250)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)

        brand = tk.Frame(sidebar, bg="#0f172a")
        brand.pack(fill="x", padx=18, pady=(18, 14))
        tk.Label(brand, text="CPFL", bg="#0f172a", fg="#93c5fd", font=("Segoe UI", 8, "bold")).pack(anchor="w")
        tk.Label(brand, text="Extrator de\nProdução", bg="#0f172a", fg="#f8fafc", font=("Segoe UI", 20, "bold"), justify="left").pack(anchor="w", pady=(2, 6))
        tk.Label(
            brand,
            text="Nova interface visual com navegação lateral e foco operacional.",
            bg="#0f172a",
            fg="#94a3b8",
            font=("Segoe UI", 9),
            justify="left",
            wraplength=200,
        ).pack(anchor="w")

        nav = tk.Frame(sidebar, bg="#0f172a")
        nav.pack(fill="x", padx=14, pady=(8, 10))

        self.btn_nav_operacao = tk.Button(
            nav,
            text="Operação diária",
            command=lambda: self._show_view("operacao"),
            bg="#1e293b",
            fg="#f8fafc",
            activebackground="#2563eb",
            activeforeground="#ffffff",
            relief="flat",
            bd=0,
            padx=14,
            pady=12,
            anchor="w",
            cursor="hand2",
            font=("Segoe UI", 10, "bold"),
        )
        self.btn_nav_operacao.pack(fill="x", pady=(0, 8))

        self.btn_nav_tratamento = tk.Button(
            nav,
            text="Tratamento e medição",
            command=lambda: self._show_view("tratamento"),
            bg="#0f172a",
            fg="#cbd5e1",
            activebackground="#2563eb",
            activeforeground="#ffffff",
            relief="flat",
            bd=0,
            padx=14,
            pady=12,
            anchor="w",
            cursor="hand2",
            font=("Segoe UI", 10, "bold"),
        )
        self.btn_nav_tratamento.pack(fill="x")

        sidebar_info = tk.Frame(sidebar, bg="#0f172a")
        sidebar_info.pack(fill="x", padx=18, pady=(18, 0))
        for titulo, var in (
            ("Status", self.status_var),
            ("Saída pronta", self.resumo_saida_var),
            ("Seleção", self.resumo_eas_var),
        ):
            box = tk.Frame(sidebar_info, bg="#111c31", highlightthickness=1, highlightbackground="#22304a")
            box.pack(fill="x", pady=(0, 10))
            tk.Label(box, text=titulo, bg="#111c31", fg="#93c5fd", font=("Segoe UI", 8, "bold")).pack(anchor="w", padx=12, pady=(10, 2))
            tk.Label(box, textvariable=var, bg="#111c31", fg="#f8fafc", font=("Segoe UI", 9, "bold"), justify="left", wraplength=180).pack(anchor="w", padx=12, pady=(0, 10))

        main = tk.Frame(outer, bg=COR_BG)
        main.pack(side="left", fill="both", expand=True)

        header = tk.Frame(main, bg=COR_BG)
        header.pack(fill="x", padx=20, pady=(18, 12))
        left_header = tk.Frame(header, bg=COR_BG)
        left_header.pack(side="left", fill="x", expand=True)
        ttk.Label(left_header, text="Painel operacional", style="Title.TLabel").pack(anchor="w")
        ttk.Label(
            left_header,
            text="Mesmo fluxo por trás, mas com um visual mais limpo e separado por contexto.",
            style="Sub.TLabel",
        ).pack(anchor="w", pady=(4, 0))

        badge = tk.Label(
            header,
            text="UI redesenhada",
            bg="#dbeafe",
            fg="#1d4ed8",
            font=("Segoe UI", 9, "bold"),
            padx=12,
            pady=7,
        )
        badge.pack(side="right")

        shell = tk.Frame(main, bg=COR_BG)
        shell.pack(fill="both", expand=True, padx=20, pady=(0, 18))

        self.view_operacao = tk.Frame(shell, bg=COR_BG)
        self.view_tratamento = tk.Frame(shell, bg=COR_BG)

        # ---------------------------
        # Operação diária
        # ---------------------------
        op_top = tk.Frame(self.view_operacao, bg=COR_BG)
        op_top.pack(fill="x", pady=(0, 14))
        op_top.columnconfigure(0, weight=5)
        op_top.columnconfigure(1, weight=4)

        panel_exec = tk.Frame(op_top, bg=COR_CARD, highlightthickness=1, highlightbackground=COR_BORDA)
        panel_exec.grid(row=0, column=0, sticky="nsew", padx=(0, 7))
        exec_wrap = tk.Frame(panel_exec, bg=COR_CARD)
        exec_wrap.pack(fill="both", expand=True, padx=18, pady=16)

        tk.Label(exec_wrap, text="Centro de execução", bg=COR_CARD, fg=COR_TEXTO, font=("Segoe UI", 13, "bold")).pack(anchor="w")
        tk.Label(exec_wrap, textvariable=self.status_var, bg=COR_CARD, fg=COR_TEXTO, font=("Segoe UI", 17, "bold")).pack(anchor="w", pady=(10, 2))
        tk.Label(exec_wrap, textvariable=self.substatus_var, bg=COR_CARD, fg=COR_TEXTO_2, font=("Segoe UI", 10)).pack(anchor="w")

        self.progress = ttk.Progressbar(exec_wrap, style="Modern.Horizontal.TProgressbar", mode="determinate", maximum=100)
        self.progress.pack(fill="x", pady=(14, 8))
        self.progress_label = tk.Label(exec_wrap, textvariable=self.counter_var, bg=COR_CARD, fg=COR_TEXTO_2, font=("Segoe UI", 9, "bold"))
        self.progress_label.pack(anchor="w", pady=(0, 12))

        resumo_grid = tk.Frame(exec_wrap, bg=COR_CARD)
        resumo_grid.pack(fill="x")
        for col in range(2):
            resumo_grid.columnconfigure(col, weight=1)

        self.card_modo = self._mini_card(resumo_grid, "Modo atual", self.resumo_modo_var)
        self.card_modo.grid(row=0, column=0, sticky="nsew", padx=(0, 6), pady=(0, 6))
        self.card_periodo = self._mini_card(resumo_grid, "Período", self.resumo_periodo_var)
        self.card_periodo.grid(row=0, column=1, sticky="nsew", padx=(6, 0), pady=(0, 6))
        self.card_eas = self._mini_card(resumo_grid, "EAs selecionadas", self.resumo_eas_var)
        self.card_eas.grid(row=1, column=0, sticky="nsew", padx=(0, 6))
        self.card_saida = self._mini_card(resumo_grid, "Arquivo pronto", self.resumo_saida_var)
        self.card_saida.grid(row=1, column=1, sticky="nsew", padx=(6, 0))

        panel_control = tk.Frame(op_top, bg=COR_CARD, highlightthickness=1, highlightbackground=COR_BORDA)
        panel_control.grid(row=0, column=1, sticky="nsew", padx=(7, 0))
        ctrl_wrap = tk.Frame(panel_control, bg=COR_CARD)
        ctrl_wrap.pack(fill="both", expand=True, padx=18, pady=16)

        tk.Label(ctrl_wrap, text="Acesso rápido", bg=COR_CARD, fg=COR_TEXTO, font=("Segoe UI", 13, "bold")).pack(anchor="w")
        tk.Label(ctrl_wrap, text="Credenciais, modo headless e disparo da rotina diária.", bg=COR_CARD, fg=COR_TEXTO_2, font=("Segoe UI", 10), justify="left", wraplength=360).pack(anchor="w", pady=(4, 12))

        cred_grid = tk.Frame(ctrl_wrap, bg=COR_CARD)
        cred_grid.pack(fill="x")
        self._entry_line(cred_grid, "Matrícula", self.matricula_var).pack(fill="x", pady=(0, 10))
        self._entry_line(cred_grid, "Senha", self.senha_var, password=True).pack(fill="x", pady=(0, 12))

        opts_row = tk.Frame(ctrl_wrap, bg=COR_CARD)
        opts_row.pack(fill="x", pady=(0, 10))
        tk.Checkbutton(
            opts_row,
            text="Executar em headless",
            variable=self.headless_var,
            bg=COR_CARD,
            fg=COR_TEXTO,
            activebackground=COR_CARD,
            activeforeground=COR_TEXTO,
            selectcolor="#1d4ed8",
            font=("Segoe UI", 9),
            relief="flat",
            bd=0,
        ).pack(anchor="w")

        agenda = tk.Frame(ctrl_wrap, bg=COR_CARD_2, highlightthickness=1, highlightbackground=COR_BORDA)
        agenda.pack(fill="x", pady=(4, 12))
        tk.Label(agenda, text="Agendamento automático", bg=COR_CARD_2, fg=COR_TEXTO, font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=12, pady=(10, 2))
        tk.Label(agenda, text="Escolha uma recorrência para disparar a extração diária.", bg=COR_CARD_2, fg=COR_TEXTO_2, font=("Segoe UI", 8)).pack(anchor="w", padx=12, pady=(0, 8))
        self.combo_agendamento = ttk.Combobox(
            agenda,
            textvariable=self.auto_interval_var,
            values=[
                "Desativado",
                "A cada 15 minutos",
                "A cada 30 minutos",
                "A cada 45 minutos",
                "A cada 1 hora",
                "A cada 2 horas",
                "A cada 3 horas",
                "Personalizado",
            ],
            state="readonly",
            font=("Segoe UI", 10),
        )
        self.combo_agendamento.pack(fill="x", padx=12, ipady=5)
        self.combo_agendamento.bind("<<ComboboxSelected>>", lambda e: self.on_schedule_change())

        personalizado = tk.Frame(agenda, bg=COR_CARD_2)
        personalizado.pack(fill="x", padx=12, pady=(8, 0))
        tk.Label(
            personalizado,
            text="Tempo personalizado (minutos):",
            bg=COR_CARD_2,
            fg=COR_TEXTO_2,
            font=("Segoe UI", 8),
        ).pack(side="left")
        self.spin_agendamento_minutos = tk.Spinbox(
            personalizado,
            from_=5,
            to=1440,
            increment=5,
            textvariable=self.auto_custom_minutes_var,
            width=8,
            font=("Segoe UI", 9),
            bg="#0f172a",
            fg=COR_TEXTO,
            insertbackground=COR_TEXTO,
            relief="flat",
            command=self.on_schedule_change,
        )
        self.spin_agendamento_minutos.pack(side="right")
        self.spin_agendamento_minutos.bind("<KeyRelease>", lambda e: self.on_schedule_change())
        self.spin_agendamento_minutos.bind("<FocusOut>", lambda e: self.on_schedule_change())

        tk.Label(agenda, textvariable=self.auto_status_var, bg=COR_CARD_2, fg=COR_TEXTO_2, font=("Segoe UI", 8), justify="left", wraplength=330).pack(anchor="w", padx=12, pady=(8, 10))

        actions = tk.Frame(ctrl_wrap, bg=COR_CARD)
        actions.pack(fill="x")
        actions.columnconfigure(0, weight=1)
        actions.columnconfigure(1, weight=1)
        self.btn_iniciar = ttk.Button(actions, text="▶ Iniciar extração diária", command=self.iniciar, style="Primary.TButton")
        self.btn_iniciar.grid(row=0, column=0, sticky="ew", padx=(0, 5))
        self.btn_copiar = ttk.Button(actions, text="📋 Copiar último TXT", command=self.copiar_dados_txt, style="Secondary.TButton")
        self.btn_copiar.grid(row=0, column=1, sticky="ew", padx=(5, 0))

        op_bottom = tk.Frame(self.view_operacao, bg=COR_BG)
        op_bottom.pack(fill="both", expand=True)
        op_bottom.columnconfigure(0, weight=2)
        op_bottom.columnconfigure(1, weight=4)
        op_bottom.rowconfigure(0, weight=1)

        log_panel = tk.Frame(op_bottom, bg=COR_CARD, highlightthickness=1, highlightbackground=COR_BORDA)
        log_panel.grid(row=0, column=0, sticky="nsew", padx=(0, 7))
        log_head = tk.Frame(log_panel, bg=COR_CARD)
        log_head.pack(fill="x", padx=18, pady=(14, 8))
        tk.Label(log_head, text="Linha do tempo operacional", bg=COR_CARD, fg=COR_TEXTO, font=("Segoe UI", 12, "bold")).pack(side="left")
        tk.Label(log_head, text="ao vivo", bg=COR_CARD, fg=COR_OK, font=("Segoe UI", 9, "bold")).pack(side="right")

        log_area = tk.Frame(log_panel, bg=COR_CARD)
        log_area.pack(fill="both", expand=True, padx=18, pady=(0, 16))
        self.txt_log = tk.Text(log_area, bg=COR_LOG, fg="#e2e8f0", insertbackground="#e2e8f0", relief="flat", bd=0, font=("Consolas", 9), wrap="none", padx=12, pady=10)
        self.txt_log.pack(side="left", fill="both", expand=True)
        yscroll = tk.Scrollbar(log_area, command=self.txt_log.yview)
        yscroll.pack(side="right", fill="y")
        self.txt_log.configure(yscrollcommand=yscroll.set)
        self.txt_log.configure(state="disabled")

        ea_panel = tk.Frame(op_bottom, bg=COR_CARD, highlightthickness=1, highlightbackground=COR_BORDA)
        ea_panel.grid(row=0, column=1, sticky="nsew", padx=(7, 0))
        ea_top = tk.Frame(ea_panel, bg=COR_CARD)
        ea_top.pack(fill="x", padx=18, pady=(14, 10))
        tk.Label(ea_top, text="Seleção de EAs", bg=COR_CARD, fg=COR_TEXTO, font=("Segoe UI", 12, "bold")).pack(side="left")

        tools = tk.Frame(ea_top, bg=COR_CARD)
        tools.pack(side="right")
        tk.Button(tools, text="Todas", command=self.marcar_todas, bg="#1d4ed8", fg="#f8fafc", activebackground="#2563eb", activeforeground="#ffffff", relief="flat", bd=0, padx=10, pady=5, cursor="hand2", font=("Segoe UI", 8, "bold")).pack(side="left", padx=(0, 6))
        tk.Button(tools, text="Nenhuma", command=self.desmarcar_todas, bg="#1e293b", fg="#f8fafc", activebackground="#334155", activeforeground="#ffffff", relief="flat", bd=0, padx=10, pady=5, cursor="hand2", font=("Segoe UI", 8, "bold")).pack(side="left")

        info_row = tk.Frame(ea_panel, bg=COR_CARD)
        info_row.pack(fill="x", padx=18, pady=(0, 8))
        tk.Label(info_row, text="Escolha as bases que entram no lote diário.", bg=COR_CARD, fg=COR_TEXTO_2, font=("Segoe UI", 9)).pack(side="left")

        self.ea_count_var = tk.StringVar(value="")
        tk.Label(info_row, textvariable=self.ea_count_var, bg=COR_CARD, fg="#93c5fd", font=("Segoe UI", 8, "bold")).pack(side="right")

        filtro_row = tk.Frame(ea_panel, bg=COR_CARD)
        filtro_row.pack(fill="x", padx=18, pady=(0, 10))
        tk.Label(filtro_row, text="Filtrar:", bg=COR_CARD, fg=COR_TEXTO_2, font=("Segoe UI", 8, "bold")).pack(side="left", padx=(0, 8))

        filtro_entry = tk.Entry(
            filtro_row,
            textvariable=self.ea_filter_var,
            bg=COR_BG_2,
            fg=COR_TEXTO,
            insertbackground=COR_TEXTO,
            relief="flat",
            bd=0,
            font=("Segoe UI", 9),
        )
        filtro_entry.pack(side="left", fill="x", expand=True, ipady=6)
        self.ea_filter_var.trace_add("write", self._on_ea_filter_change)

        ea_wrap = tk.Frame(ea_panel, bg=COR_CARD)
        ea_wrap.pack(fill="both", expand=True, padx=18, pady=(0, 16))

        self.ea_frame = tk.Frame(ea_wrap, bg=COR_CARD)
        self.ea_frame.pack(fill="both", expand=True)

        self._render_ea_tiles()

        # ---------------------------
        # Tratamento e medição
        # ---------------------------
        tr_top = tk.Frame(self.view_tratamento, bg=COR_BG)
        tr_top.pack(fill="x", pady=(0, 14))
        tr_top.columnconfigure(0, weight=4)
        tr_top.columnconfigure(1, weight=3)

        tr_summary = tk.Frame(tr_top, bg=COR_CARD, highlightthickness=1, highlightbackground=COR_BORDA)
        tr_summary.grid(row=0, column=0, sticky="nsew", padx=(0, 7))
        tr_summary_wrap = tk.Frame(tr_summary, bg=COR_CARD)
        tr_summary_wrap.pack(fill="both", expand=True, padx=18, pady=16)
        tk.Label(tr_summary_wrap, text="Tratamento de dados", bg=COR_CARD, fg=COR_TEXTO, font=("Segoe UI", 13, "bold")).pack(anchor="w")
        tk.Label(tr_summary_wrap, text="Parciais, fechamento do mês anterior e medição usando a última base tratada.", bg=COR_CARD, fg=COR_TEXTO_2, font=("Segoe UI", 10), justify="left", wraplength=520).pack(anchor="w", pady=(4, 12))

        tr_cards = tk.Frame(tr_summary_wrap, bg=COR_CARD)
        tr_cards.pack(fill="x")
        for col in range(2):
            tr_cards.columnconfigure(col, weight=1)
        self.card_modo_trat = self._mini_card(tr_cards, "Modo atual", self.resumo_modo_var)
        self.card_modo_trat.grid(row=0, column=0, sticky="nsew", padx=(0, 6), pady=(0, 6))
        self.card_periodo_trat = self._mini_card(tr_cards, "Período", self.resumo_periodo_var)
        self.card_periodo_trat.grid(row=0, column=1, sticky="nsew", padx=(6, 0), pady=(0, 6))
        self.card_saida_trat = self._mini_card(tr_cards, "Saída pronta", self.resumo_saida_var)
        self.card_saida_trat.grid(row=1, column=0, sticky="nsew", padx=(0, 6))
        self.card_eas_trat = self._mini_card(tr_cards, "Seleção atual", self.resumo_eas_var)
        self.card_eas_trat.grid(row=1, column=1, sticky="nsew", padx=(6, 0))

        tr_actions_panel = tk.Frame(tr_top, bg=COR_CARD, highlightthickness=1, highlightbackground=COR_BORDA)
        tr_actions_panel.grid(row=0, column=1, sticky="nsew", padx=(7, 0))
        tr_actions_wrap = tk.Frame(tr_actions_panel, bg=COR_CARD)
        tr_actions_wrap.pack(fill="both", expand=True, padx=18, pady=16)
        tk.Label(tr_actions_wrap, text="Ações disponíveis", bg=COR_CARD, fg=COR_TEXTO, font=("Segoe UI", 13, "bold")).pack(anchor="w")
        tk.Label(tr_actions_wrap, text="Os parciais usam as EAs marcadas. O cálculo de medição usa a última saída tratada existente.", bg=COR_CARD, fg=COR_TEXTO_2, font=("Segoe UI", 10), justify="left", wraplength=360).pack(anchor="w", pady=(4, 12))

        tratamento_actions = tk.Frame(tr_actions_wrap, bg=COR_CARD)
        tratamento_actions.pack(fill="x")
        for col in range(2):
            tratamento_actions.columnconfigure(col, weight=1)

        self.btn_semana = ttk.Button(tratamento_actions, text="🗓 Parcial semana", command=self.processar_parcial_semana, style="Secondary.TButton")
        self.btn_semana.grid(row=0, column=0, sticky="ew", padx=(0, 5), pady=(0, 8))
        self.btn_mes = ttk.Button(tratamento_actions, text="📅 Parcial mês", command=self.processar_parcial_mes, style="Secondary.TButton")
        self.btn_mes.grid(row=0, column=1, sticky="ew", padx=(5, 0), pady=(0, 8))
        self.btn_mes_anterior = ttk.Button(tratamento_actions, text="🗂 Fechamento mês anterior", command=self.processar_fechamento_mes_anterior, style="Secondary.TButton")
        self.btn_mes_anterior.grid(row=1, column=0, sticky="ew", padx=(0, 5), pady=(0, 8))
        self.btn_medicao_jundiai = ttk.Button(tratamento_actions, text="💰 Cálculo medição", command=self.processar_medicao_disjuntor_jundiai, style="Primary.TButton")
        self.btn_medicao_jundiai.grid(row=1, column=1, sticky="ew", padx=(5, 0), pady=(0, 8))

        self.lbl_ultima_saida = tk.Label(
            tr_actions_wrap,
            text="Última saída: nenhuma ainda",
            bg="#eff6ff",
            fg="#1d4ed8",
            font=("Segoe UI", 9, "bold"),
            justify="left",
            wraplength=360,
            anchor="w",
            padx=12,
            pady=10,
        )
        self.lbl_ultima_saida.pack(fill="x", pady=(10, 0))

        tr_bottom = tk.Frame(self.view_tratamento, bg=COR_BG)
        tr_bottom.pack(fill="both", expand=True)
        tr_bottom.columnconfigure(0, weight=2)
        tr_bottom.columnconfigure(1, weight=3)
        tr_bottom.rowconfigure(0, weight=1)

        help_panel = tk.Frame(tr_bottom, bg=COR_CARD, highlightthickness=1, highlightbackground=COR_BORDA)
        help_panel.grid(row=0, column=0, sticky="nsew", padx=(0, 7))
        help_wrap = tk.Frame(help_panel, bg=COR_CARD)
        help_wrap.pack(fill="both", expand=True, padx=18, pady=16)
        tk.Label(help_wrap, text="Guia rápido", bg=COR_CARD, fg=COR_TEXTO, font=("Segoe UI", 12, "bold")).pack(anchor="w")
        ajuda = (
            "• Parcial semana: segunda a sábado\n"
            "• Parcial mês: mês atual completo\n"
            "• Fechamento mês anterior: mês anterior completo\n"
            "• Cálculo medição: usa o último CSV tratado\n"
            "• Cache: reaproveita CSVs se o arquivo não mudou"
        )
        tk.Label(help_wrap, text=ajuda, bg=COR_CARD, fg=COR_TEXTO_2, font=("Segoe UI", 9), justify="left", anchor="nw").pack(anchor="w", pady=(8, 14))

        info_box = tk.Frame(help_wrap, bg=COR_CARD_2, highlightthickness=1, highlightbackground=COR_BORDA)
        info_box.pack(fill="both", expand=True)
        tk.Label(info_box, text="Última saída disponível", bg=COR_CARD_2, fg=COR_TEXTO, font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=12, pady=(12, 4))
        self.lbl_tratamento_saida = tk.Label(
            info_box,
            textvariable=self.resumo_saida_var,
            bg=COR_CARD_2,
            fg="#1d4ed8",
            font=("Segoe UI", 10, "bold"),
            justify="left",
            anchor="w",
            wraplength=260,
        )
        self.lbl_tratamento_saida.pack(anchor="w", padx=12, pady=(0, 12))

        log_panel_tr = tk.Frame(tr_bottom, bg=COR_CARD, highlightthickness=1, highlightbackground=COR_BORDA)
        log_panel_tr.grid(row=0, column=1, sticky="nsew", padx=(7, 0))
        log_head_tr = tk.Frame(log_panel_tr, bg=COR_CARD)
        log_head_tr.pack(fill="x", padx=18, pady=(14, 8))
        tk.Label(log_head_tr, text="Linha do tempo do tratamento", bg=COR_CARD, fg=COR_TEXTO, font=("Segoe UI", 12, "bold")).pack(side="left")
        tk.Label(log_head_tr, text="ao vivo", bg=COR_CARD, fg=COR_OK, font=("Segoe UI", 9, "bold")).pack(side="right")

        log_trat_area = tk.Frame(log_panel_tr, bg=COR_CARD)
        log_trat_area.pack(fill="both", expand=True, padx=18, pady=(0, 16))
        self.txt_log_tratamento = tk.Text(log_trat_area, bg=COR_LOG, fg="#e2e8f0", insertbackground="#e2e8f0", relief="flat", bd=0, font=("Consolas", 9), wrap="none", padx=12, pady=10)
        self.txt_log_tratamento.pack(side="left", fill="both", expand=True)
        yscroll_trat = tk.Scrollbar(log_trat_area, command=self.txt_log_tratamento.yview)
        yscroll_trat.pack(side="right", fill="y")
        self.txt_log_tratamento.configure(yscrollcommand=yscroll_trat.set)
        self.txt_log_tratamento.configure(state="disabled")

        self._show_view("operacao")

    def _show_view(self, view_name):
        for frame in (self.view_operacao, self.view_tratamento):
            frame.pack_forget()

        if view_name == "tratamento":
            self.view_tratamento.pack(fill="both", expand=True)
            self.btn_nav_tratamento.configure(bg="#2563eb", fg="#ffffff")
            self.btn_nav_operacao.configure(bg="#0f172a", fg="#cbd5e1")
        else:
            self.view_operacao.pack(fill="both", expand=True)
            self.btn_nav_operacao.configure(bg="#2563eb", fg="#ffffff")
            self.btn_nav_tratamento.configure(bg="#0f172a", fg="#cbd5e1")

    def _mini_card(self, parent, titulo, textvariable):
        frame = tk.Frame(parent, bg=COR_CARD_2, highlightthickness=1, highlightbackground=COR_BORDA)
        tk.Label(frame, text=titulo, bg=COR_CARD_2, fg=COR_TEXTO_2, font=("Segoe UI", 8, "bold")).pack(anchor="w", padx=12, pady=(10, 3))
        tk.Label(frame, textvariable=textvariable, bg=COR_CARD_2, fg=COR_TEXTO, font=("Segoe UI", 10, "bold"), justify="left", wraplength=260).pack(anchor="w", padx=12, pady=(0, 12))
        return frame

    def atualizar_resumo_ui(self):
        qtd_eas = len(self.eas_selecionadas()) if hasattr(self, "ea_vars") else 0
        self.resumo_eas_var.set(f"EAs selecionadas: {qtd_eas}")
        self.resumo_modo_var.set(f"Modo atual: {self.ultima_acao_rotulo}")
        periodo_texto = "Período padrão: hoje"
        if "Parcial semana" in self.ultima_acao_rotulo:
            inicio, fim = obter_periodo_atual("semana")
            periodo_texto = f"Semana: {inicio.strftime('%d/%m')} a {fim.strftime('%d/%m')}"
        elif "Parcial mês" in self.ultima_acao_rotulo or "Cálculo medição" in self.ultima_acao_rotulo:
            inicio, fim = obter_periodo_atual("mes")
            periodo_texto = f"Mês: {inicio.strftime('%d/%m')} a {fim.strftime('%d/%m')}"
        elif "Fechamento Mês anterior" in self.ultima_acao_rotulo:
            inicio, fim = obter_periodo_atual("mes_anterior")
            periodo_texto = f"Mês anterior: {inicio.strftime('%d/%m')} a {fim.strftime('%d/%m')}"
        self.resumo_periodo_var.set(periodo_texto)
        nome_saida = Path(self.ultimo_txt_gerado).name if self.ultimo_txt_gerado else "nenhuma"
        self.resumo_saida_var.set(f"Saída pronta: {nome_saida}")

    def _entry_line(self, parent, titulo, var, password=False):
        wrap = tk.Frame(parent, bg=COR_CARD)
        tk.Label(wrap, text=titulo, bg=COR_CARD, fg=COR_TEXTO_2, font=("Segoe UI", 9, "bold")).pack(anchor="w", pady=(0, 4))
        entry = tk.Entry(
            wrap,
            textvariable=var,
            show="•" if password else "",
            bg="#0b1220",
            fg=COR_TEXTO,
            insertbackground=COR_TEXTO,
            relief="flat",
            bd=1,
            highlightthickness=1,
            highlightbackground=COR_BORDA,
            highlightcolor=COR_DESTAQUE,
            font=("Segoe UI", 10),
        )
        entry.pack(fill="x", ipady=9)
        return wrap


    def destroy(self):
        try:
            if hasattr(self, "main_canvas"):
                self.main_canvas.unbind_all("<MouseWheel>")
                self.main_canvas.unbind_all("<Button-4>")
                self.main_canvas.unbind_all("<Button-5>")
        except Exception:
            pass
        try:
            self.root.destroy()
        except Exception:
            pass

    def _center(self):
        self.root.update_idletasks()
        w = self.root.winfo_width()
        h = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (w // 2)
        y = (self.root.winfo_screenheight() // 2) - (h // 2)
        self.root.geometry(f"{w}x{h}+{x}+{y}")

    def push_log(self, texto):
        self.fila.put(("log", texto))

    def _append_log(self, texto):
        for widget_name in ("txt_log", "txt_log_tratamento"):
            widget = getattr(self, widget_name, None)
            if widget is None:
                continue
            widget.configure(state="normal")
            widget.insert("end", texto + "\n")
            widget.see("end")
            widget.configure(state="disabled")

    def _poll(self):
        try:
            while True:
                tipo, valor = self.fila.get_nowait()
                if tipo == "log":
                    self._append_log(str(valor))
                elif tipo == "status":
                    self.status_var.set(str(valor))
                elif tipo == "substatus":
                    self.substatus_var.set(str(valor))
                elif tipo == "progress":
                    atual = int(valor.get("atual", 0))
                    total = max(1, int(valor.get("total", 1)))
                    self.progress["value"] = (atual / total) * 100
                elif tipo == "counter":
                    exportadas = int(valor.get("exportadas", 0))
                    total = int(valor.get("total", 0))
                    self.counter_var.set(f"{exportadas}/{total} exportadas")
                elif tipo == "done":
                    self._finalizar(valor)
        except queue.Empty:
            pass

        if not self.executando and self.proxima_execucao:
            restante = max(0, int(self.proxima_execucao - datetime.now().timestamp()))
            mins, segs = divmod(restante, 60)
            horas, mins = divmod(mins, 60)
            self.auto_status_var.set(
                f"Próxima execução automática em {horas:02d}:{mins:02d}:{segs:02d}"
            )

        try:
            self.atualizar_resumo_ui()
        except Exception:
            pass
        self.root.after(120, self._poll)


    def _on_ea_selection_change(self):
        self.atualizar_resumo_ui()
        self._salvar_estado_eas()
        self._render_ea_tiles()

    def marcar_todas(self):
        for var in self.ea_vars.values():
            var.set(True)

    def desmarcar_todas(self):
        for var in self.ea_vars.values():
            var.set(False)




    def atualizar_ultimo_resultado(self, caminho_csv=None, caminho_txt=None, descricao=None, periodo_inicio=None, periodo_fim=None):
        try:
            self.ultimo_csv_gerado = str(caminho_csv) if caminho_csv else self.ultimo_csv_gerado
            self.ultimo_txt_gerado = str(caminho_txt) if caminho_txt else self.ultimo_txt_gerado
            self.ultima_acao_rotulo = descricao or self.ultima_acao_rotulo
            self.ultimo_periodo_inicio = periodo_inicio or self.ultimo_periodo_inicio
            self.ultimo_periodo_fim = periodo_fim or self.ultimo_periodo_fim

            nome = Path(self.ultimo_txt_gerado).name if self.ultimo_txt_gerado else "nenhuma ainda"
            self.lbl_ultima_saida.configure(
                text=f"Última saída: {self.ultima_acao_rotulo}\nArquivo pronto para copiar: {nome}"
            )
            self.atualizar_resumo_ui()
        except Exception:
            pass

    def atualizar_ultima_saida(self, caminho_txt, descricao):
        self.atualizar_ultimo_resultado(caminho_txt=caminho_txt, descricao=descricao)

    def copiar_dados_txt(self, caminho_txt=None):
        try:
            caminho_base = caminho_txt or self.ultimo_txt_gerado
            if not caminho_base:
                messagebox.showwarning("Aviso", "Ainda não existe nenhum resultado gerado para copiar.")
                return

            caminho_txt = Path(caminho_base)
            if not caminho_txt.exists():
                messagebox.showwarning("Aviso", f"Arquivo TXT não encontrado:\n{caminho_txt}")
                return

            conteudo = caminho_txt.read_text(encoding="utf-8")
            self.root.clipboard_clear()
            self.root.clipboard_append(conteudo)
            self.root.update()
            self._append_log(f"📋 Dados copiados para a área de transferência: {caminho_txt.name}")
            messagebox.showinfo(
                "Sucesso",
                f"Dados copiados para a área de transferência!\n\nOrigem: {self.ultima_acao_rotulo}\nArquivo: {caminho_txt.name}",
            )
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao copiar dados:\n{e}")

    def processar_medicao_disjuntor_jundiai(self):
        if self.executando:
            return

        if not self.ultimo_csv_gerado:
            messagebox.showwarning(
                "Aviso",
                "Primeiro gere um resultado tratado (diário, semana, mês ou mês anterior) para então calcular a medição."
            )
            return

        caminho_csv = Path(self.ultimo_csv_gerado)
        if not caminho_csv.exists():
            messagebox.showwarning("Aviso", f"CSV tratado não encontrado:\n{caminho_csv}")
            return

        try:
            self.status_var.set("Calculando medição")
            self.substatus_var.set(f"Base: {caminho_csv.name}")
            self._append_log("=" * 90)
            self._append_log(f"Cálculo medição Disjuntor Jundiaí: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
            self._append_log(f"📄 Base do cálculo: {caminho_csv}")
            self._append_log("=" * 90)

            df_final = carregar_dataframe_resultado(caminho_csv)
            periodo_inicio = self.ultimo_periodo_inicio or "-"
            periodo_fim = self.ultimo_periodo_fim or "-"
            caminho_medicao = obter_saida_medicao_por_csv(caminho_csv)

            gerado = gerar_planilha_medicao_disjuntor_jundiai(
                df_final,
                caminho_medicao,
                periodo_inicio,
                periodo_fim,
                logger=log,
            )

            if not gerado:
                messagebox.showwarning(
                    "Aviso",
                    "Nenhuma nota de Disjuntor Jundiaí (JUN55, JUN59 ou SAL55) foi encontrada no último resultado tratado."
                )
                return

            self.ultima_medicao_gerada = gerado
            self.status_var.set("Medição concluída")
            self.substatus_var.set(f"Planilha: {Path(gerado).name}")
            self._append_log(f"✅ Medição gerada com sucesso: {gerado}")
            messagebox.showinfo(
                "Concluído",
                f"Cálculo de medição gerado com sucesso.\n\nBase: {caminho_csv.name}\nPeríodo: {periodo_inicio} até {periodo_fim}\nPlanilha: {gerado}"
            )
        except Exception as e:
            self.status_var.set("Erro na medição")
            self.substatus_var.set("Verifique o log operacional")
            self._append_log(f"❌ Erro ao calcular medição: {e}")
            messagebox.showerror("Erro", f"Falha ao calcular a medição:\n{e}")


    def processar_fechamento_mes_anterior(self):
        self._iniciar_processamento_periodo("mes_anterior")

    def processar_parcial_semana(self):

        self._iniciar_processamento_periodo("semana")

    def processar_parcial_mes(self):
        self._iniciar_processamento_periodo("mes")

    def _iniciar_processamento_periodo(self, tipo_periodo):
        if self.executando:
            return

        self._salvar_estado_eas()
        eas = self.eas_selecionadas()
        if not eas:
            messagebox.showwarning("Aviso", "Selecione pelo menos uma EA.")
            return

        nomes = {"semana": "Parcial semana", "mes": "Parcial mês", "mes_anterior": "Fechamento Mês anterior"}
        self.ultima_acao_rotulo = nomes.get(tipo_periodo, "Processamento parcial")
        self.executando = True
        self.total = len(eas)
        self.progress["value"] = 0
        self.counter_var.set(f"0/{self.total} EAs selecionadas")
        self.btn_iniciar.configure(state="disabled")
        self.status_var.set(nomes.get(tipo_periodo, "Processamento parcial"))
        self.substatus_var.set("Buscando arquivos já baixados para o período")
        self.atualizar_resumo_ui()
        self._append_log("=" * 90)
        self._append_log(f"{nomes.get(tipo_periodo, 'Processamento parcial')}: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        self._append_log("=" * 90)

        self.worker = threading.Thread(
            target=self._run_period_worker,
            kwargs={"tipo_periodo": tipo_periodo, "eas": eas},
            daemon=True,
        )
        self.worker.start()

    def _run_period_worker(self, tipo_periodo, eas):
        global CURRENT_APP
        stdout_original = sys.stdout
        stderr_original = sys.stderr
        writer = QueueWriter(self.fila)
        CURRENT_APP = self
        try:
            sys.stdout = writer
            sys.stderr = writer

            nomes = {"semana": "Parcial semana", "mes": "Parcial mês", "mes_anterior": "Fechamento Mês anterior"}
            self.fila.put(("status", nomes.get(tipo_periodo, "Processamento parcial")))
            self.fila.put(("substatus", "Processando arquivos do período selecionado"))
            self.fila.put(("counter", {"exportadas": 0, "total": len(eas)}))
            self.fila.put(("progress", {"atual": 0, "total": max(1, len(eas)), "nome": nomes.get(tipo_periodo, "")}))

            processamento = processar_arquivos_por_periodo(
                tipo_periodo=tipo_periodo,
                eas_list=eas,
                logger=log,
            )

            self.fila.put((
                "done",
                {
                    "ok": True,
                    "acao": "parcial",
                    "tipo_periodo": tipo_periodo,
                    "processamento": processamento,
                },
            ))
        except Exception as e:
            erro = "".join(traceback.format_exception(type(e), e, e.__traceback__))
            self.fila.put(("log", erro))
            self.fila.put(("done", {"ok": False, "erro": str(e), "acao": "parcial", "tipo_periodo": tipo_periodo}))
        finally:
            CURRENT_APP = None
            sys.stdout = stdout_original
            sys.stderr = stderr_original


    def get_schedule_interval_ms(self):
        valor = self.auto_interval_var.get()

        opcoes_fixas = {
            "A cada 15 minutos": 15,
            "A cada 30 minutos": 30,
            "A cada 45 minutos": 45,
            "A cada 1 hora": 60,
            "A cada 2 horas": 120,
            "A cada 3 horas": 180,
        }

        if valor in opcoes_fixas:
            return opcoes_fixas[valor] * 60 * 1000

        if valor == "Personalizado":
            try:
                minutos = int(str(self.auto_custom_minutes_var.get()).strip())
            except Exception:
                self.auto_status_var.set("Informe um tempo personalizado válido em minutos.")
                return None

            if minutos < 5:
                minutos = 5
                self.auto_custom_minutes_var.set("5")
            elif minutos > 1440:
                minutos = 1440
                self.auto_custom_minutes_var.set("1440")

            return minutos * 60 * 1000

        return None

    def on_schedule_change(self):
        intervalo = self.get_schedule_interval_ms()
        if intervalo is None:
            self.cancelar_agendamento()
            self.auto_status_var.set("Agendamento automático desativado")
            self._append_log("⏹️ Agendamento automático desativado.")
            return

        minutos = intervalo // 60000
        self.auto_status_var.set(
            f"Agendamento ativo: nova execução {minutos} min após o término da atual."
        )
        self._append_log(
            f"⏰ Agendamento configurado para a cada {minutos} minutos."
        )

        if not self.executando:
            self.agendar_proxima_execucao()

    def cancelar_agendamento(self):
        if self.auto_job is not None:
            try:
                self.root.after_cancel(self.auto_job)
            except Exception:
                pass
            self.auto_job = None
        self.proxima_execucao = None

    def agendar_proxima_execucao(self):
        intervalo = self.get_schedule_interval_ms()
        self.cancelar_agendamento()

        if intervalo is None:
            self.auto_status_var.set("Agendamento automático desativado")
            return

        self.proxima_execucao = datetime.now().timestamp() + (intervalo / 1000)
        proxima = datetime.fromtimestamp(self.proxima_execucao).strftime("%H:%M:%S")
        self.auto_status_var.set(f"Próxima execução automática às {proxima}")
        self._append_log(f"⏰ Próxima execução automática agendada para {proxima}.")
        self.auto_job = self.root.after(intervalo, self.iniciar_agendado)

    def iniciar_agendado(self):
        self.auto_job = None
        self.proxima_execucao = None
        if self.executando:
            self._append_log("ℹ️ Execução automática adiada porque já existe uma execução em andamento.")
            self.agendar_proxima_execucao()
            return

        self._append_log("🚀 Iniciando execução automática agendada.")
        self.iniciar(modo_agendado=True)

    def salvar_env(self, silencioso=False):
        matricula = self.matricula_var.get().strip()
        senha = self.senha_var.get().strip()
        headless = "1" if self.headless_var.get() else "0"

        if not matricula or not senha:
            if not silencioso:
                messagebox.showwarning("Aviso", "Preencha matrícula e senha antes de salvar o .env.")
            return False

        if not ENV_PATH.exists():
            ENV_PATH.write_text("", encoding="utf-8")
        set_key(str(ENV_PATH), "MATRICULA", matricula)
        set_key(str(ENV_PATH), "SENHA", senha)
        set_key(str(ENV_PATH), "HEADLESS", headless)
        set_key(str(ENV_PATH), "AUTO_INTERVAL_LABEL", self.auto_interval_var.get())
        set_key(str(ENV_PATH), "AUTO_INTERVAL_MINUTES", self.auto_custom_minutes_var.get().strip() or "90")
        os.environ["MATRICULA"] = matricula
        os.environ["SENHA"] = senha
        os.environ["HEADLESS"] = headless
        if not silencioso:
            messagebox.showinfo("Pronto", f"Arquivo .env salvo em:\n{ENV_PATH}")
        return True

    def eas_selecionadas(self):
        return [ea for ea, var in self.ea_vars.items() if var.get()]

    def iniciar(self, modo_agendado=False):
        if self.executando:
            return

        eas = self.eas_selecionadas()
        if not eas:
            messagebox.showwarning("Aviso", "Selecione pelo menos uma EA.")
            return

        if not self.matricula_var.get().strip() or not self.senha_var.get().strip():
            messagebox.showwarning("Aviso", "Preencha matrícula e senha.")
            return

        self.executando = True
        self.total = len(eas)
        self.progress["value"] = 0
        self.counter_var.set(f"0/{self.total} exportadas")
        self.btn_iniciar.configure(state="disabled")
        self.status_var.set("Preparando execução")
        self.substatus_var.set("Iniciando thread do robô")
        self._append_log("=" * 90)
        self._append_log(f"Início: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        self._append_log("=" * 90)

        worker_args = {
            "headless": self.headless_var.get(),
            "eas": eas,
        }
        self.worker = threading.Thread(target=self._run_worker, kwargs={**worker_args, "modo_agendado": modo_agendado}, daemon=True)
        self.worker.start()

    def _run_worker(self, headless, eas, modo_agendado=False):
        global CURRENT_APP
        stdout_original = sys.stdout
        stderr_original = sys.stderr
        writer = QueueWriter(self.fila)
        CURRENT_APP = self
        try:
            self.salvar_env(silencioso=True)
            sys.stdout = writer
            sys.stderr = writer
            bot = ExtratorProducao(fila=self.fila, headless=headless, eas=eas)
            resultado = bot.executar()

            self.fila.put(("status", "Processando arquivos"))
            self.fila.put(("substatus", "Navegador fechado; iniciando tratamento dos dados"))
            processamento = processar_arquivos_baixados(eas_list=eas, logger=log)

            resultado["processamento"] = processamento
            self.fila.put(("done", {"ok": True, "resultado": resultado, "modo_agendado": modo_agendado}))
        except Exception as e:
            erro = "".join(traceback.format_exception(type(e), e, e.__traceback__))
            self.fila.put(("log", erro))
            self.fila.put(("done", {"ok": False, "erro": str(e), "modo_agendado": modo_agendado}))
        finally:
            CURRENT_APP = None
            sys.stdout = stdout_original
            sys.stderr = stderr_original


    def _finalizar(self, payload):
        self.executando = False
        self.btn_iniciar.configure(state="normal")
        modo_agendado = bool(payload.get("modo_agendado"))
        acao = payload.get("acao", "extracao")

        if payload.get("ok"):
            self.progress["value"] = 100

            if acao == "parcial":
                processamento = payload.get("processamento") or {}
                tipo_periodo = payload.get("tipo_periodo", processamento.get("tipo_periodo", ""))
                nomes = {"semana": "Parcial semana", "mes": "Parcial mês", "mes_anterior": "Fechamento Mês anterior"}
                titulo = nomes.get(tipo_periodo, "Processamento parcial")

                self.counter_var.set(f"{self.total}/{self.total} EAs selecionadas")
                self.status_var.set(f"{titulo} concluído")
                self.substatus_var.set(
                    f"Período: {processamento.get('periodo_inicio', '-')} até {processamento.get('periodo_fim', '-')}"
                )
                self.atualizar_resumo_ui()

                msg = (
                    f"{titulo} finalizado com sucesso.\n\n"
                    f"Período: {processamento.get('periodo_inicio', '-')} até {processamento.get('periodo_fim', '-')}\n"
                    f"Linhas lidas: {processamento.get('linhas_lidas', 0)}\n"
                    f"Linhas finais: {processamento.get('linhas_finais', 0)}\n\n"
                    f"TXT: {processamento.get('txt', '')}\n"
                    f"CSV: {processamento.get('csv', '')}\n"
                    f"Inatividade: {processamento.get('inatividade', '')}"
                )
                if processamento.get("medicao_disjuntor_jundiai"):
                    msg += f"\nMedição Disjuntor Jundiaí: {processamento.get('medicao_disjuntor_jundiai')}" 

                txt_gerado = processamento.get("txt")
                csv_gerado = processamento.get("csv")
                if txt_gerado or csv_gerado:
                    self.atualizar_ultimo_resultado(
                        caminho_csv=csv_gerado,
                        caminho_txt=txt_gerado,
                        descricao=titulo,
                        periodo_inicio=processamento.get("periodo_inicio"),
                        periodo_fim=processamento.get("periodo_fim"),
                    )

                messagebox.showinfo("Concluído", msg)

                alertas_stc = processamento.get("alertas_stc_um_executor") or []
                if alertas_stc:
                    total_alertas = processamento.get("total_alertas_stc_um_executor", 0)
                    detalhes = "\n".join(
                        f"• {item['equipe']}: {item['qtd_notas']} nota(s)" for item in alertas_stc
                    )
                    messagebox.showwarning(
                        "Alerta STC - 1 executor",
                        f"Foram identificadas {total_alertas} nota(s) STC com apenas 1 executor.\n\n{detalhes}"
                    )
                    self._append_log("⚠️ Alerta visual STC exibido ao usuário.")

                self._append_log(f"✅ {titulo} concluído com sucesso.")
            else:
                resultado = payload.get("resultado") or {}
                total = resultado.get("total", 0)
                exportadas = resultado.get("exportadas", 0)
                processamento = resultado.get("processamento") or {}
                alertas_stc = processamento.get("alertas_stc_um_executor") or []
                total_alertas = processamento.get("total_alertas_stc_um_executor", 0)
                self.counter_var.set(f"{exportadas}/{total} exportadas")
                self.status_var.set("Execução concluída")
                self.substatus_var.set("Extração e tratamento finalizados com sucesso")

                msg = (
                    f"Extração finalizada com sucesso.\n\n"
                    f"Exportadas: {exportadas}/{total}\n"
                    f"Linhas lidas: {processamento.get('linhas_lidas', 0)}\n"
                    f"Linhas finais: {processamento.get('linhas_finais', 0)}\n\n"
                    f"Arquivos gerados em:\n{PROCESS_OUTPUT_FOLDER}"
                )
                txt_gerado = processamento.get("txt")
                csv_gerado = processamento.get("csv")
                if txt_gerado or csv_gerado:
                    self.atualizar_ultimo_resultado(
                        caminho_csv=csv_gerado,
                        caminho_txt=txt_gerado,
                        descricao="Extração diária",
                        periodo_inicio=processamento.get("periodo_inicio"),
                        periodo_fim=processamento.get("periodo_fim"),
                    )
                if not modo_agendado:
                    messagebox.showinfo("Concluído", msg)
                    if alertas_stc:
                        detalhes = "\n".join(
                            f"• {item['equipe']}: {item['qtd_notas']} nota(s)" for item in alertas_stc
                        )
                        messagebox.showwarning(
                            "Alerta STC - 1 executor",
                            f"Foram identificadas {total_alertas} nota(s) STC com apenas 1 executor.\n\n{detalhes}"
                        )
                        self._append_log("⚠️ Alerta visual STC exibido ao usuário.")
                else:
                    if alertas_stc:
                        detalhes = "; ".join(
                            f"{item['equipe']}: {item['qtd_notas']} nota(s)" for item in alertas_stc
                        )
                        self._append_log(
                            f"⚠️ Alerta STC automático: {total_alertas} nota(s) com 1 executor. {detalhes}"
                        )
                    self._append_log("✅ Execução automática concluída com sucesso.")
        else:
            self.status_var.set("Erro na execução")
            self.substatus_var.set("Verifique o log operacional")
            if not modo_agendado:
                messagebox.showerror("Erro", payload.get("erro", "Ocorreu um erro inesperado."))
            else:
                self._append_log(f"❌ Erro na execução automática: {payload.get('erro', 'Ocorreu um erro inesperado.')}")

        self._acao_medicao_somente = False

        if acao != "parcial" and self.get_schedule_interval_ms() is not None:
            self.agendar_proxima_execucao()



def mostrar_splash():
    splash = tk.Tk()
    splash.overrideredirect(True)
    splash.configure(bg=COR_BG)

    largura = 460
    altura = 280
    x = (splash.winfo_screenwidth() // 2) - (largura // 2)
    y = (splash.winfo_screenheight() // 2) - (altura // 2)
    splash.geometry(f"{largura}x{altura}+{x}+{y}")
    splash.lift()
    try:
        splash.wm_attributes("-topmost", True)
    except Exception:
        pass

    frame = tk.Frame(
        splash,
        bg=COR_CARD,
        highlightthickness=1,
        highlightbackground=COR_BORDA,
    )
    frame.pack(fill="both", expand=True, padx=1, pady=1)

    logo_label = tk.Label(frame, bg=COR_CARD)
    logo_label.pack(pady=(32, 16))

    if APP_LOGO_PATH.exists():
        try:
            img = Image.open(APP_LOGO_PATH).convert("RGBA")
            img = img.resize((140, 140))
            logo_img = ImageTk.PhotoImage(img)
            logo_label.configure(image=logo_img)
            logo_label.image = logo_img
        except Exception:
            logo_label.configure(
                text="SCS",
                fg=COR_TEXTO,
                bg=COR_CARD,
                font=("Segoe UI", 34, "bold"),
            )
    else:
        logo_label.configure(
            text="SCS",
            fg=COR_TEXTO,
            bg=COR_CARD,
            font=("Segoe UI", 34, "bold"),
        )

    tk.Label(
        frame,
        text="Extrator de Produção",
        bg=COR_CARD,
        fg=COR_TEXTO,
        font=("Segoe UI", 18, "bold"),
    ).pack()

    tk.Label(
        frame,
        text="Desenvolvido por Gabriel Irense",
        bg=COR_CARD,
        fg="#f8fafc",
        font=("Segoe UI", 10),
    ).pack(pady=(8, 18))

    progress = ttk.Progressbar(frame, mode="indeterminate", length=220)
    progress.pack(pady=(0, 26))
    progress.start(12)

    splash.update()
    splash.after(2200, splash.destroy)
    splash.mainloop()


def iniciar_interface():
    load_dotenv(override=True)
    root = tk.Tk()
    try:
        if APP_ICON_PATH.exists():
            root.iconbitmap(str(APP_ICON_PATH))
    except Exception:
        pass
    ExtratorGUI(root)
    root.mainloop()


if __name__ == "__main__":
    mostrar_splash()
    iniciar_interface()
